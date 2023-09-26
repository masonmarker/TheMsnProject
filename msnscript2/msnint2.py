# Interpreters MSNScript2
#
# See documentation for more information,
# documentation could lack functions or
# capabilities in this interpreter, as
# this is a work in progress.
#
# docs: masonmarker.com/#/msn2docs
# run 'python msn2.py help' for help
#
# Author : Mason Marker
# Start date : 09/15/2022


# TODO
# speed up function interpretation
# by determining obj and func by argument count first
# as opposed to iterating through all functions.
# do this and/or determine closest function after each character
# consumption.
# this also entails testing and reordering function priority to
# maximize speed.
#
# TODO
# implement string parsing of the same character at which
# it was defined. ex: "hello \"world\"" -> hello "world"
# currently, this is not possible
#
# TODO (less important)
# split interpreter up into multiple files
# for better readability
#
# TODO
# implement warnings and warning handling, as this
# language was designed to be safe yet flexible
#
# TODO
# implement linear interpretation in areas of heavy logic, this applies
# easily non linear approaches in several blocks
# such as <<>> or system calls such as script()
#
# TODO
# implement an interpretation for block syntax
# that permits the existence of whitespace / tabs / carriage returns
# in the multilined block to interpret
#
# TODO
# ensure no code repetition
# (it definitely exists)

# TODO
# 2.0.388
#
# * CHROME BROWSER
# - finish lib/auto/chrome class extension
#
# * PRIVATE FUNCTIONS
# - complete all of their work similarly to a function in any other language,
#   where the function can access external variables, and variables local
#   to the function will never be accessable to any external context unless specified.

# the current logical implementation is conceptual,
# deoptimized, and exists to prove functionality as speed can
# be enhanced later

# NOTE
# Python runner's alias are determined by
# global python_alias (defined below), defaulting to 'python'
#
# console operations will not run successfully
# if the python runner is incorrect,
# is includes all console() calls, process splitting,
# etc.

# importing necessary dependencies
# for all lines of execution
import os
import threading

# remove warnings for calling of integers: "10()"
import warnings
warnings.filterwarnings("ignore", category=SyntaxWarning)

# variables


class Var:
    # constructs a new Var
    def __init__(self, _msn2_reserved_varname, _msn2_reserved_varvalue):
        self.name = _msn2_reserved_varname
        self.value = _msn2_reserved_varvalue

    # determines equality of another Var
    def __eq__(self, other):
        if isinstance(other, Var):
            return other.name == self.name

    # string representation of Var
    def __str__(self):
        return f"(Var {self.name}={self.value})"


# path to the common settings file
settings_path = 'msn2_settings.json'
# latest version of the interpreter'
# set later
latest_version = None
# if settings does not exist
if not os.path.exists(settings_path):
    import json
    # get the latest version number from system/latest.json
    with open('system/latest.json') as f:
        # using the variable declared above
        latest_version = json.load(f)['latest']
    # create settings
    with open(settings_path, 'w') as f:
        # dump default settings
        json.dump({'settings': {'has_ran': False, 'runner_alias': 'python'},
                   'version': latest_version}, f)
# global settings
settings = None
# python alias is in the msn2 settings json
python_alias = 'python'
# obtains the python alias
with open('msn2_settings.json') as f:
    import json
    settings = json.load(f)
    python_alias = settings['settings']['runner_alias']
# msn2 implementation of None
msn2_none = '___msn2_None_'
# thread serial
thread_serial = 0
# global vars
lock = threading.Lock()
auxlock = threading.Lock()
# lock for pointer controls
pointer_lock = threading.Lock()
# pywinauto automation lock
auto_lock = threading.Lock()
# timings_set switch
timings_set = False
# user defined syntax
syntax = {}
# user defined enclosing syntax
enclosed = {}
# user defined macros
macros = {}
# user defined post macros
# aka macros that are defined that the end of a line
postmacros = {}
# user defined inline syntax
inlines = {}
# OpenAI model presets
# these are the latest models at the time of creation
models = None
# accounting information
inst_tree = {}
lines_ran = []
total_ints = 0
# automation
apps = {}
# colors for colored printing,
# defaults to nothing until assigned
colors = ''

# interprets MSNScript2, should create a new interpreter for each execution iteration


class Interpreter:
    # initializer
    def __init__(self):
        # check if this environment has executed
        # a script before
        if not settings['settings']['has_ran']:
            # if it has not ran, install
            # dependencies
            # check if pip is installed
            try:
                import pip
            except:
                # pip not found
                print('[MSN2] pip dependency installer not found, installing pip...')
                # if not, install pip
                os.system(f'{python_alias} -m pip install --upgrade pip')
            # install dependencies
            os.system(f"{python_alias} install_deps.py")
            # finished
            # set has_ran in the json file to true
            settings['settings']['has_ran'] = True
            # write to the json file
            with open(settings_path, 'w') as f:
                import json
                json.dump(settings, f)
        # get the latest version number from system/latest.json
        with open('system/latest.json') as f:
            import json
            global latest_version
            latest_version = json.load(f)['latest']
        # if the version is not the latest
        if settings['version'] != latest_version:
            # update the version
            settings['version'] = latest_version
            # write to the json file
            with open(settings_path, 'w') as f:
                import json
                json.dump(settings, f)
        # set the current settings
        self.settings = settings['settings']
        # get the version in the json file
        self.version = settings['version']
        # basic logging
        self.lines = []
        self.out = ''
        self.log = ''
        self.errors = []
        # environment logging
        self.vars = {}
        self.methods = {}
        self.loggedmethod = []
        self.objects = {}
        self.calledmethod = None
        self.env_max_chars = 200
        # advanced environment logging
        self.current_line = 0
        self.breaking = False
        self.breaking_return = []
        self.redirect = None
        self.redirecting = False
        self.redirect_inside = []
        self.imports = set()
        self.domains = set()
        # threading
        self.thread = None
        self.threads = {}
        self.parent = None
        # ChatGPT connections
        self.openaikey = None
        self.tokens = 100
        self.browser_path = None
        self.serial_1 = 0
        # hosting endpoints
        self.endpoints = {}
        self.endpoint_datas = {}
        self.endpoint_path = 'demos/practical/apidata/apitestdata.csv'
        # multiprocessing
        self.processes = {}
        # global and local scopes for internal Python environment
        self._globals = {}
        self._locals = {}
        # in a try block
        self.trying = False
    # determines if a line is a comment or not

    def is_comment(self, _line):
        return _line.startswith('#') or _line.startswith('::')

    # executes stored script
    def execute(self, script):
        # convert script to lines
        self.lines = []
        # for aggregate syntax support !{}
        inml = False
        ml = ''
        # for block syntax
        inblock = False
        p = 0
        # whether or not to keep
        keep_space = False
        keep_block = ''
        skipping = False
        # for each line of code
        for line in script.split("\n"):
            # add to list of lines
            self.lines.append(line)
            # continue if this line is a comment
            if not line or self.is_comment(line):
                continue
            # for running Python
            if not keep_space and line.endswith('\\\\'):
                # determine if this script should run
                # based on the expression at the beginning of the line
                line = line[:-2]
                line = line.strip()
                # if there is a conditional Python execution
                if line:
                    # execute the conditional line
                    # with import capabilities
                    result = self.interpret(line)
                    keep_space = True
                    # if the conditional evaluated to True
                    if result:
                        # if the conditional line is true, then execute the Python
                        skipping = False
                        keep_block = ''
                    # otherwise, skip the Python
                    else:
                        skipping = True
                        keep_block = ''
                else:
                    skipping = False
                    keep_space = True
                    keep_block = ''
            # if the Python block has been ended
            elif keep_space and line.endswith('\\\\'):
                # if skipping
                if skipping:
                    skipping = False
                    keep_block = ''
                    keep_space = False
                    continue
                # fix line
                line = line[:-2]
                line = line.strip()
                # remove the trailing newline
                # character
                keep_block = keep_block[:-1]
                # modify conditionals
                keep_space = False
                # if setting to a variable
                if line:
                    # set variable to python block
                    self.vars[line] = Var(line, keep_block)
                else:
                    # execute Python
                    self.exec_python(keep_block)
                # reset python block and skipping flag
                keep_block = ''
                skipping = False
            # skipping this Python block
            # due to conditional
            elif keep_space and skipping:
                continue
            # if in a python block
            elif keep_space:
                keep_block += f"{line}\n"
                self.current_line += 1
                continue
            else:
                line = line.strip()
            if self.is_comment(line):
                self.current_line += 1
                continue
            else:
                # aggregate syntax !{} (not recommended for most cases)
                if line.startswith('!{') and line.endswith('}'):
                    ml = line[2:-1]
                    self.interpret(ml)
                    ml = ''
                elif not inml and line.startswith("!{"):
                    inml = True
                    ml += line[2:]
                elif inml and line.endswith("}"):
                    inml = False
                    ml += line[0:len(line) - 1]
                    self.interpret(ml)
                    ml = ''
                elif inml:
                    ml += line
                # block syntax (recommended for most cases)
                elif not inblock and line.endswith('(') or line.endswith(',') \
                    or line.endswith('{') or line.endswith('[') or line.endswith('=') \
                        or line.endswith('{='):
                    for c in line:
                        if c == '(':
                            p += 1
                        if c == ')':
                            p -= 1
                        ml += c
                    inblock = True
                elif inblock:
                    for i in range(len(line)):
                        c = line[i]
                        if c == '(':
                            p += 1
                        if c == ')':
                            p -= 1
                        # end of syntax met
                        if p == 0:
                            ml += line[i:]
                            inter = ml
                            ml = ''
                            inblock = False
                            self.interpret(inter, keep_space=keep_space)
                            break
                        ml += c
                else:
                    self.interpret(line, keep_space=keep_space)
            self.current_line += 1
        return self.out

    def replace_vars(self, line):
        boo = line
        for varname in sorted(self.vars.keys(), key=len, reverse=True):
            try:
                boo = boo.replace(varname, str(
                    self.get_var(eval(f'"{varname}"', {}, {}))))
            except:
                None
        return boo

    # function to attempt to pull an integer out of the function
    # if it is an integer, then it is a loop that runs func times
    def get_int(self, func):
        try:
            return int(func)
        except:
            return None

    def interpret(self, line, block={}, keep_space=False):

        # acquiring globals
        global total_ints
        global lock
        global auxlock
        global auto_lock
        global pointer_lock
        global python_alias
        global inst_tree
        # accounting
        total_ints += 1
        lines_ran.append(line)
        # determine the location of this instruction
        # in the inst_tree depending on lines_ran
        # and the current line
        inst_tree[total_ints] = [self.current_line, line]
        # interpreter is breaking
        if self.breaking:
            return self.breaking_return
        # strip line for interpretation
        try:
            if not keep_space:
                line = line.strip()
        except:
            return
        l = len(line)
        # whether the line should immediately continue or not
        cont = False
        if line == '':
            return
        # the below conditions interpret a line based on initial appearances
        # beneath these conditions will the Interpreter then parse the arguments from the line as a method call

        # method-specific line reached
        if line.startswith('--'):
            line = line[2:]
            try:
                if not self.methods[self.loggedmethod[-1]].ended:
                    self.methods[self.loggedmethod[-1]].add_body(line)
            except:
                None
            return
        # new variable setting and modification syntax as of 12/20/2022
        # iterates to the first '=' sign, capturing the variable name in the
        # process (as it should)
        # msn1 fallback
        if line[0] == '@':
            return self.interpret_msnscript_1(line[1:])
        # python fallback mode specification,
        # both <<>> and
        if line.startswith('<<'):
            # parse all text in the line for text surrounded by |
            funccalls = []
            infunc = False
            func = ''
            for i in range(line.rindex('>>')):
                if line[i] == '|' and not infunc:
                    infunc = True
                elif line[i] == '|' and infunc:
                    infunc = False
                    funccalls.append(func)
                    func = ''
                elif infunc:
                    func += line[i]
            for function in funccalls:
                ret = self.interpret(function)
                if isinstance(ret, str):
                    line = line.replace(
                        f'|{function}|', f'"{str(ret)}"')
                else:
                    line = line.replace(f'|{function}|', str(ret))
            line = line[2:-2]
            try:
                return eval(line)
            except:
                return line

        # embedded MSN2 interpretation macro
        if line.startswith('<2>'):
            # parse all text in the line for text surrounded by %
            funccalls = []
            infunc = False
            func = ''
            for i in range(3, line.rindex('<2>')):
                if line[i] == '%' and not infunc:
                    infunc = True
                elif line[i] == '%' and infunc:
                    infunc = False
                    funccalls.append(func)
                    func = ''
                elif infunc:
                    func += line[i]

            # for each msn2 evaluation
            for function in funccalls:
                ret = self.interpret(function)
                if isinstance(ret, str):
                    line = line.replace(
                        f'%{function}%', f'"{str(ret)}"')
                else:
                    line = line.replace(f'%{function}%', str(ret))
            line = line[3:-3]
            try:
                return self.interpret(line)
            except:
                try:
                    return eval(line, {}, {})
                except:
                    return line

        # user defined syntax
        for key in syntax:
            if line.startswith(key):
                return self.run_syntax(key, line)
        # user defined macro
        for token in macros:
            if line.startswith(token):
                # if the macro returns a value instead of executing a function
                if len(macros[token]) == 4:
                    return macros[token][3]
                # variable name
                varname = macros[token][1]
                val = line[len(token):]
                # store extended for user defined syntax
                self.vars[varname] = Var(varname, val)
                # execute function
                return self.interpret(macros[token][2])
        # user defined postmacro
        for token in postmacros:
            if line.endswith(token):
                # if the macro returns a value instead of executing a function
                if len(postmacros[token]) == 4:
                    return postmacros[token][3]
                varname = postmacros[token][1]
                val = line[0:len(line) - len(token)]
                self.vars[varname] = Var(varname, val)
                return self.interpret(postmacros[token][2])
        # variable replacement, generally unsafe, but replaces
        # all variable names as they're written the the expression after
        # the '*'
        if line[0] == '*':
            return self.interpret(self.replace_vars(line[1:]))
        # invoking user defined enclosing syntax
        for key in enclosed:
            start = enclosed[key][0]
            end = enclosed[key][1]
            if line.startswith(start) and line.endswith(end):
                if len(enclosed[key]) == 5:
                    return enclosed[key][4]
                varname = enclosed[key][2]
                self.vars[varname] = Var(
                    varname, line[len(start):len(line) - len(end)])
                return self.interpret(enclosed[key][3])
        # checks for active Interpreter redirect request
        # generally slow, due to checking for the redirect
        if self.redirecting and not 'stopredirect()' in line.replace(' ', ''):
            self.redirect_inside.append([self.redirect[1], line])
            return self.redirect

        # try base literal
        try:
            if not line.startswith('--'):
                # try evaluating the line
                _ret = eval(line, {}, {})
                # eval cannot be a python class, because names of variables
                # could result in python classes
                # should also not be a built in function
                if not isinstance(_ret, type) and not isinstance(_ret, type(eval)):
                    return _ret
        except:
            None

        func = ''
        objfunc = ''
        obj = ''
        s = 0
        sp = 0
        for i in range(l):
            if cont:
                continue
            try:
                c = line[i]
            except:
                break
            if c == ' ' and s == 0:
                sp += 1
                continue
            if c == '.':
                obj = func
                func = ''
                objfunc = ''
                continue
            # basic method creation
            if c == '~':
                returnvariable = ''
                self.loggedmethod.append('')
                for j in range(i + 1, len(line)):
                    if line[j] != ' ':
                        if line[j] == '(':
                            args = self.method_args(line, j)
                            for k in range(args[1], len(line)):
                                if line[k] != ' ':
                                    if line[k] == '-' and line[k + 1] == '>':
                                        for l in range(k + 2, len(line)):
                                            if line[l] != ' ':
                                                returnvariable += line[l]
                            break
                        self.loggedmethod[-1] += line[j]
                if self.loggedmethod[-1] not in self.methods.keys():
                    self.methods[self.loggedmethod[-1]
                                 ] = self.Method(self.loggedmethod[-1], self)
                else:
                    break
                for arg in args[0]:
                    if arg != '':
                        self.vars[arg] = None
                        self.methods[self.loggedmethod[-1]].add_arg(arg)
                self.methods[self.loggedmethod[-1]].add_return(returnvariable)
                return self.loggedmethod[-1]

            # interpreting a function
            elif c == '(':
                mergedargs = ''
                p = 1
                for j in range(i + 1, l - 1):
                    c2 = line[j]
                    if p == 0:
                        break
                    if c2 == '(':
                        p += 1
                    if c2 == ')':
                        p -= 1
                    mergedargs += c2
                args = self.get_args(mergedargs)
                f = len(func)
                # clean function for handling
                func = func.strip()
                objfunc = objfunc.strip()

                # retrieving arguments

                # class attribute / method access
                if obj in self.vars:
                    vname = obj
                    try:
                        var = self.get_var(vname)
                    except:
                        var = self.vars[vname]
                    try:
                        object = self.vars[obj].value
                    except:
                        object = self.vars[obj]
                    try:
                        # if the object is a class
                        if objfunc in object:
                            # if the object is a self.Method
                            if type(object[objfunc]) == self.Method:
                                # get the Method object
                                method = object[objfunc]
                                # get the number of arguments to the method
                                num_args = len(method.args)
                                # args to pass to the function
                                to_pass = [vname]
                                # if there is no argument
                                if args[0][0] != '':
                                    # for each parsed argument
                                    for k in range(num_args):
                                        try:
                                            to_pass.append(self.parse(
                                                k, line, f, sp, args)[2])
                                        except:
                                            None
                                # create return variable
                                ret_name = method.returns[0]
                                # if the return variable doesn't exist
                                if ret_name not in self.vars:
                                    self.vars[ret_name] = Var(ret_name, None)
                                # insert vname into args[0]
                                args.insert(0, [vname])
                                # execute method
                                try:
                                    method.run(to_pass, self, args)
                                # Index out of bounds error
                                except IndexError:
                                    self.raise_incorrect_args(str(len(method.args)), str(
                                        self.arg_count(args) - 1), line, lines_ran, method)
                                try:
                                    return eval(str(self.vars[method.returns[0]].value), {}, {})
                                except:
                                    try:
                                        return str(self.vars[method.returns[0]].value)
                                    except:
                                        return str(self.vars[method.returns[0]])
                            # otherwise if we're accessing an attribute
                            # no arguments given
                            if args[0][0] == '':
                                return object[objfunc]
                            # parameter provided, wants to set attribute
                            param = self.parse(0, line, f, sp, args)[2]
                            self.vars[obj].value[objfunc] = param
                            return param
                    except self.MSN2Exception as e:
                        raise e
                    except:
                        None

                    # # as of 2.0.388,
                    # # if the objfunc ends with '!',
                    # # it becomes destructive
                    # # and returns the object
                    if objfunc.endswith('!'):
                        # recreate the line to interpret without the '!'
                        self.vars[vname].value = self.interpret(
                            f"{vname}.{objfunc[:-1]}({mergedargs})"
                        )
                        return self.vars[vname].value
                    

                    # methods available to all types
                    if objfunc == 'copy':
                        try:
                            return object.copy()
                        except:
                            # no attribute copy
                            self.err(
                                'Error copying object.',
                                f'Object "{obj}" does not have attribute "copy".',
                                line, lines_ran
                            )

                    if objfunc == 'print':
                        # if no arguments
                        if args[0][0] == '':
                            print(object)
                            return object
                        # if one argument
                        elif len(args) == 1:
                            # what to print
                            to_print = f"{self.parse(0, line, f, sp, args)[2]}{object}"
                            # print the object
                            print(to_print)
                            return to_print

                        # if two arguments
                        elif len(args) == 2:
                            # what to print
                            to_print = f"{self.parse(0, line, f, sp, args)[2]}{object}{self.parse(1, line, f, sp, args)[2]}"
                            # print the object
                            print(to_print)
                            # return the printed object
                            return to_print
                        return object
                    if objfunc == 'val':
                        return object

                    try:
                        # general information
                        if objfunc == 'type':
                            return type(object)
                        if objfunc == 'len':
                            return len(object)
                        # casting
                        if objfunc == 'str':
                            return str(object)
                        if objfunc == 'int':
                            return int(object)
                        if objfunc == 'float':
                            return float(object)
                        if objfunc == 'complex':
                            return complex(object)
                        if objfunc == 'bool':
                            return bool(object)
                        if objfunc == 'dict':
                            return dict(object)
                    except:
                        # casting error
                        self.err(
                            'Casting error',
                            f'Could not cast object of type {type(object)} to the type specified.',
                            line, lines_ran
                        )

                    # switches the values of the two variables
                    if objfunc == 'switch':
                        # other variable name
                        other_varname = self.parse(0, line, f, sp, args)[2]
                        # check variable name
                        self.check_varname(other_varname, line)
                        # other_varname must exist in self.vars
                        if other_varname not in self.vars:
                            self.err(
                                'Error switching variables.',
                                f'Variable name "{other_varname}" does not exist in this context.',
                                line, lines_ran
                            )
                        # switch the variables
                        self.vars[vname].value, self.vars[other_varname].value = self.vars[other_varname].value, self.vars[vname].value
                        # return the variable
                        return self.vars[vname].value

                    # renames a variable
                    if objfunc == 'rename':
                        # get the variable name
                        varname = self.parse(0, line, f, sp, args)[2]
                        # variable name must be a string
                        self.check_varname(varname, line)
                        # rename the variable
                        self.vars[varname] = Var(varname, object)
                        # delete the old entry
                        del self.vars[vname]
                        # return the variable
                        return self.vars[varname]

                    # gets values from the object if the statement is true for each object
                    # runs the function on each element / kv pair
                    #
                    # method is not destructive
                    if objfunc == 'if':
                        # variable name
                        varname = self.parse(0, line, f, sp, args)[2]
                        # check varname
                        self.check_varname(varname, line)
                        new_list = []
                        # perform logic
                        for el in self.vars[vname].value:
                            self.vars[varname] = Var(varname, el)
                            if self.interpret(args[1][0]):
                                new_list.append(el)
                        return new_list
                    # comparing object types
                    if objfunc == 'is':
                        return object is self.parse(0, line, f, sp, args)[2]
                    # test if the object is equal to all the parameters
                    if objfunc == 'equals':
                        for i in range(len(args)):
                            if object != self.parse(i, line, f, sp, args)[2]:
                                return False
                        return True
                    # obtains a slice of the iterable
                    if objfunc == 'slice':
                        first = self.parse(0, line, f, sp, args)[2]
                        second = self.parse(1, line, f, sp, args)[2]
                        # ensure both arguments are integers or None
                        self.type_err(
                            [(first, (int, type(None))), (second, (int, type(None)))], line, lines_ran)
                        return self.vars[vname].value[first:second]
                    # gets the index of the object
                    if objfunc == 'index':
                        # argument must be of type string
                        return self.vars[vname].value.index(self.parse(0, line, f, sp, args)[2])
                    # exports a variable to the parent context
                    if objfunc == 'export':
                        # if an argument is provided
                        # export as name
                        if args[0][0] != '':
                            vname = self.parse(0, line, f, sp, args)[2]
                            # check vname
                            self.check_varname(vname, line)
                        self.parent.vars[vname] = Var(vname, object)
                        return object
                    # if no objfunc, there has been a request
                    # for repeated method calls/access
                    if objfunc == '':
                        ret = self.vars[vname].value
                        # for each block
                        for arg in args:
                            ret = self.interpret(f"{vname}.{arg[0]}")
                        return ret

                    # string_name() returns the string name of the object
                    if objfunc == 'string_name':
                        return vname
                    # performs a function for each element in the iterable
                    if objfunc == 'each':
                        # get the variable name
                        varname = self.parse(0, line, f, sp, args)[2]
                        # check varname
                        self.check_varname(varname, line)
                        # try an indexable
                        try:
                            for i in range(len(object)):
                                self.vars[varname] = Var(varname, object[i])
                                self.interpret(args[1][0])
                        except:
                            # try a set
                            for i in object:
                                self.vars[varname] = Var(varname, i)
                                self.interpret(args[1][0])
                        return object

                    # rfind and lfind
                    # find the index of the first element that satisfies the condition
                    if objfunc == 'rfind':
                        return self.vars[vname].value.rfind(self.parse(0, line, f, sp, args)[2])
                    if objfunc == 'lfind':
                        return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])
                    # find the index of the first element that satisfies the condition
                    if objfunc == 'find':
                        return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])
                    # filters the iterable
                    if objfunc == 'filter':
                        # get the variable name
                        varname = self.parse(0, line, f, sp, args)[2]
                        # check variable name
                        self.check_varname(varname, line)
                        # filtered
                        filtered = []
                        # filter the iterable
                        for el in object:
                            self.vars[varname] = Var(varname, el)
                            if self.interpret(args[1][0]):
                                filtered.append(el)
                        # set the variable to the filtered list
                        self.vars[vname].value = filtered
                        # return the filtered list
                        return self.vars[vname].value

                    # basic arithmetic, non-destructive
                    # takes any amount of arguments
                    try:
                        if objfunc == '+':
                            ret = object
                            for i in range(len(args)):
                                ret += self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '-':
                            ret = object
                            for i in range(len(args)):
                                ret -= self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '*' or objfunc == 'x':
                            ret = object
                            for i in range(len(args)):
                                ret *= self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '/':
                            ret = object
                            for i in range(len(args)):
                                ret /= self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '%':
                            ret = object
                            for i in range(len(args)):
                                ret %= self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '**':
                            ret = object
                            for i in range(len(args)):
                                ret **= self.parse(i, line, f, sp, args)[2]
                            return ret
                        if objfunc == '//':
                            ret = object
                            for i in range(len(args)):
                                ret //= self.parse(i, line, f, sp, args)[2]
                            return ret
                    except:
                        self.raise_operation_err(vname, objfunc, line)
                    # applies methods to to the object, considering
                    # the method takes one argument
                    if objfunc == 'func':
                        ret = object
                        # apply the function to the object
                        for arg in args:
                            ret = self.interpret(f"{arg[0]}({ret})")
                        return ret
                    # reverses the iterable
                    if objfunc == 'reverse':
                        # only types that can be reversed:
                        # list, tuple, string
                        # check types
                        self.type_err(
                            [(object, (list, tuple, str))], line, lines_ran)
                        self.vars[vname].value = self.vars[vname].value[::-1]
                        return self.vars[vname].value
                    # determines if this object is in the object passed
                    if objfunc == 'in':
                        in_var = self.parse(0, line, f, sp, args)[2]
                        # in_var must be of type iterable
                        self.check_iterable(in_var, line)
                        # if object is of type string
                        if isinstance(object, str):
                            # in_var must be a string also
                            self.type_err([(in_var, (str,))], line, lines_ran)
                            return object in in_var
                        else:
                            # otherwise, in_var cannot be a string
                            if isinstance(in_var, str):
                                self.err(
                                    'Type error', f'Cannot search for type {type(object)} in type {type(in_var)}\nConsider casting variable "{vname}" to {str}, \nor change in() argument to a different iterable', line, lines_ran)
                        return object in in_var

                    # variable type specific methods
                    # the isinstance branches below indicate mostly  DESCTRUCTIVE methods!
                    # so be sure to read the code
                    # number specific functions
                    if isinstance(object, int) or isinstance(object, float) or isinstance(object, complex):
                        # increases the value of the variable by 1
                        if objfunc == '++' or objfunc == 'inc':
                            self.vars[vname].value += 1
                            return self.vars[vname].value
                        elif objfunc == '--' or objfunc == 'dec':
                            self.vars[vname].value -= 1
                            return self.vars[vname].value
                        # determines if the number is even
                        elif objfunc == 'even':
                            return self.vars[vname].value % 2 == 0
                        elif objfunc == 'odd':
                            return self.vars[vname].value % 2 != 0
                        # all of the below methods take any amount of arguments
                        # and perform the operation on the variable

                        try:
                            if objfunc == 'add':
                                for i in range(len(args)):
                                    self.vars[vname].value += self.parse(
                                        i, line, f, sp, args)[2]
                                return self.vars[vname].value
                            elif objfunc == 'sub':
                                for i in range(len(args)):
                                    self.vars[vname].value -= self.parse(
                                        i, line, f, sp, args)[2]
                                return self.vars[vname].value
                            elif objfunc == 'mul':
                                for i in range(len(args)):
                                    self.vars[vname].value *= self.parse(
                                        i, line, f, sp, args)[2]
                                return self.vars[vname].value
                            elif objfunc == 'div':
                                for i in range(len(args)):
                                    self.vars[vname].value /= self.parse(
                                        i, line, f, sp, args)[2]
                                return self.vars[vname].value
                        except:
                            return self.raise_operation_err(vname, objfunc, line)

                        # computes the absolute value of the number
                        if objfunc == 'abs':
                            # can only be performed on numbers
                            self.vars[vname].value = abs(
                                self.vars[vname].value)
                            return self.vars[vname].value
                        # rounds this number to the nearest place specified by the first argument
                        elif objfunc == 'round':
                            # round to the nearest decimal place
                            if args[0][0] != '':
                                # decimal place to round to
                                place = self.parse(0, line, f, sp, args)[2]
                                # place type must be int
                                self.type_err([(place, (int,))],
                                              line, lines_ran)
                                self.vars[vname].value = round(
                                    self.vars[vname].value, place)
                            else:
                                self.vars[vname].value = round(
                                    self.vars[vname].value)
                            return self.vars[vname].value
                        elif objfunc == 'floor':
                            # using math
                            import math
                            self.vars[vname].value = math.floor(
                                self.vars[vname].value)
                            return self.vars[vname].value
                        elif objfunc == 'ceil':
                            # using math
                            import math
                            self.vars[vname].value = math.ceil(
                                self.vars[vname].value)
                            return self.vars[vname].value
                        # negates the value, if positive the value becomes negative and vice versa
                        elif objfunc == 'neg':
                            self.vars[vname].value = -self.vars[vname].value
                            return self.vars[vname].value
                        # all of the below methods should take any amount of arguments
                        # test if the variable is greater than all arguments
                        try:
                            if objfunc == 'greater' or objfunc == 'greaterthan' or objfunc == 'g':
                                return all(self.vars[vname].value > self.parse(i, line, f, sp, args)[2] for i in range(len(args)))
                            elif objfunc == 'less' or objfunc == 'lessthan' or objfunc == 'l':
                                return all(self.vars[vname].value < self.parse(i, line, f, sp, args)[2] for i in range(len(args)))
                            elif objfunc == 'greaterequal' or objfunc == 'ge':
                                return all(self.vars[vname].value >= self.parse(i, line, f, sp, args)[2] for i in range(len(args)))
                            elif objfunc == 'lessequal' or objfunc == 'le':
                                return all(self.vars[vname].value <= self.parse(i, line, f, sp, args)[2] for i in range(len(args)))
                        except:
                            self.raise_comp(objfunc, vname, line)
                        # more basic functions
                        return self.vars[vname].value

                    # set based functions
                    elif isinstance(object, set):
                        # adds all arguments to the set object
                        if objfunc == 'add' or objfunc == 'put':
                            for i in range(len(args)):
                                self.vars[vname].value.add(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value
                        if objfunc == 'pop':
                            return self.vars[vname].value.pop()
                        # removes all arguments from the set object
                        if objfunc == 'remove':
                            for i in range(len(args)):
                                self.vars[vname].value.remove(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value
                        # converts this set to a list
                        if objfunc == 'list':
                            return list(self.vars[vname].value)
                        # gets at an index in the set
                        # sets are not indexable
                        if objfunc == 'get':
                            # index to get at
                            ind = self.parse(0, line, f, sp, args)[2]
                            # index must be an int
                            self.type_err([(ind, (int,))], line, lines_ran)
                            # get the index
                            for i in object:
                                if ind == 0:
                                    return i
                                ind -= 1
                    # array based functions
                    elif isinstance(object, list):
                        # adds all arguments to the first argument which should be a variable name
                        # as a string
                        if objfunc == 'push' or objfunc == 'append' or objfunc == 'add':
                            for i in range(len(args)):
                                self.vars[vname].value.append(
                                    self.parse(i, line, f, sp, args)[2])
                            return self.vars[vname].value
                        # pops a value from the array
                        elif objfunc == 'pop':
                            return self.vars[vname].value.pop()
                        # getting at an index
                        elif objfunc == 'get':
                            ind = self.parse(0, line, f, sp, args)[2]
                            # type of ind must be int
                            self.type_err([(ind, (int,))], line, lines_ran)
                            return self.vars[vname].value[ind]
                        # sets at an index
                        elif objfunc == 'set':
                            ind = self.parse(0, line, f, sp, args)[2]
                            # type of ind must be int
                            self.type_err([(ind, (int,))], line, lines_ran)
                            self.vars[vname].value[ind] = self.parse(
                                1, line, f, sp, args)[2]
                            return self.vars[vname].value
                        # gets the average of this array
                        elif objfunc == 'avg' or objfunc == 'average':
                            # try to compute the average
                            try:
                                return sum(self.vars[vname].value) / len(self.vars[vname].value)
                            except:
                                if not self.vars[vname].value:
                                    return self.raise_empty_array(line)
                                else:
                                    return self.raise_avg(line)
                                return raise_empty_array(self.vars[vname].value, line)
                                # if there's an error,
                        # inserts all values at an index
                        if objfunc == 'insert':
                            # inserts the rest of the arguments, one at a time
                            for i in range(len(args)):
                                ind = self.parse(i, line, f, sp, args)[2]
                                # index must be an int
                                self.type_err([(ind, (int,))], line, lines_ran)
                                val = self.parse(i, line, f, sp, args)[2]
                                self.vars[vname].value.insert(ind, val)
                            return self.vars[vname].value
                        # removes a certain amount of all arguments supplied
                        if objfunc == 'removen':
                            # removes count amount of the rest of the arguments from the object
                            count = self.parse(0, line, f, sp, args)[2]
                            # count must be an int
                            self.type_err([(count, (int,))], line, lines_ran)
                            for i in range(1, len(args)):
                                for j in range(count):
                                    val = self.parse(i, line, f, sp, args)[2]
                                    try:
                                        del var[var.index(val)]
                                    except ValueError:
                                        self.raise_value(val, line)
                            return object

                        # removes all occurances of each argument from the list
                        if objfunc == 'remove':
                            for i in range(len(args)):
                                while self.parse(i, line, f, sp, args)[2] in var:
                                    try:
                                        del var[(_v := var.index(self.parse(
                                            i, line, f, sp, args)[2]))]
                                    except ValueError:
                                        self.raise_value(_v, line)
                            return object
                        # gets a sorted copy of this array
                        if objfunc == 'sorted':
                            return sorted(self.vars[vname].value)
                        # sorts this array
                        if objfunc == 'sort':
                            self.vars[vname].value.sort()
                            return self.vars[vname].value
                        # gets the length of the array
                        if objfunc == 'len':
                            return len(self.vars[vname].value)
                        # determines if a list is empty
                        if objfunc == 'empty':
                            return len(self.vars[vname].value) == 0
                        # determines if this list contains an element
                        if objfunc == 'contains' or objfunc == 'has' or objfunc == 'includes':
                            return self.parse(0, line, f, sp, args)[2] in self.vars[vname].value
                        # finds an element in a list
                        # unlike index(), find returns -1 instead of throwing an
                        # error
                        if objfunc == 'find':
                            return self.vars[vname].value.find(self.parse(0, line, f, sp, args)[2])
                        # shuffles a list
                        if objfunc == 'shuffle':
                            import random
                            random.shuffle(self.vars[vname].value)
                            return self.vars[vname].value
                        # performs a function for each element in the iterable
                        # map() is a destructive method
                        if objfunc == 'map':
                            # get the variable name
                            varname = self.parse(0, line, f, sp, args)[2]
                            # check varname
                            self.check_varname(varname, line)
                            for i in range(len(object)):
                                self.vars[varname] = Var(varname, object[i])
                                object[i] = self.interpret(args[1][0])
                            del self.vars[varname]
                            return object
                        # joins the array by the first argument
                        if objfunc == 'join' or objfunc == 'delimit':
                            delimiter = self.parse(0, line, f, sp, args)[2]
                            # delimiter must be a string
                            self.type_err([(delimiter, (str,))],
                                          line, lines_ran)
                            # join the array
                            return delimiter.join(map(str, self.vars[vname].value))
                        # converts this list to a set
                        if objfunc == 'toset':
                            return set(self.vars[vname].value)
                    # if the object is a string
                    elif isinstance(object, str):
                        if objfunc == 'add':
                            for i in range(len(args)):
                                adding = self.parse(i, line, f, sp, args)[2]
                                # adding must be a string
                                self.type_err(
                                    [(adding, (str,))], line, lines_ran)
                                self.vars[vname].value += adding
                            return self.vars[vname].value
                        if objfunc == 'split':
                            splitting_by = self.parse(0, line, f, sp, args)[2]
                            # splitting_by must be a string
                            self.type_err(
                                [(splitting_by, (str,))], line, lines_ran)
                            return self.vars[vname].value.split(splitting_by)
                        # gets the lines of this string
                        if objfunc == 'lines':
                            return self.vars[vname].value.split('\n')
                        # determines if the string is a digit
                        if objfunc == 'isdigit':
                            return self.vars[vname].value.isdigit()
                        # determines if the string is a letter
                        if objfunc == 'isalpha':
                            return self.vars[vname].value.isalpha()
                        # replaces all instances of the first argument with the second argument
                        if objfunc == 'replace':
                            # what to replace
                            replacing = self.parse(0, line, f, sp, args)[2]
                            wth = self.parse(1, line, f, sp, args)[2]
                            # both must be strings
                            self.type_err(
                                [(replacing, (str,)), (wth, (str,))], line, lines_ran)
                            # replacing with
                            if len(args) == 2:
                                # replaces all instances of replacing with wth
                                self.vars[vname].value = self.vars[vname].value.replace(
                                    replacing, wth)
                            elif len(args) == 3:
                                third = self.parse(2, line, f, sp, args)[2]
                                # third must be a string
                                self.type_err([(third, (str,))],
                                              line, lines_ran)
                                self.vars[vname].value = self.vars[vname].value.replace(
                                    replacing, wth, third)
                            # returns the new string
                            return self.vars[vname].value
                        # strips the value at the variable name
                        if objfunc == 'strip':
                            self.vars[vname].value = self.vars[vname].value.strip()
                            return self.vars[vname].value
                        # obtains a stripped version of itself
                        if objfunc == 'stripped':
                            return self.vars[vname].value.strip()
                        # obtains itself
                        if objfunc == 'self':
                            try:
                                return self.vars[vname].value
                            except:
                                return self.vars[vname]
                        # sets a character in this string
                        if objfunc == 'set':
                            # index to set
                            index = self.parse(0, line, f, sp, args)[2]
                            # index must be an int
                            self.type_err([(index, (int,))], line, lines_ran)
                            # create a new string with the new character
                            self.vars[
                                vname].value = f"{self.vars[vname].value[:index]}{self.parse(1, line, f, sp, args)[2]}{self.vars[vname].value[index + 1:]}"
                            # returns the new string
                            return self.vars[vname].value
                        # gets a character in this string
                        if objfunc == 'get':
                            ind = self.parse(0, line, f, sp, args)[2]
                            # ind must be an int
                            self.type_err([(ind, (int,))], line, lines_ran)
                            return self.vars[vname].value[ind]
                        # uppercases the string
                        if objfunc == 'upper':
                            self.vars[vname].value = self.vars[vname].value.upper()
                            return self.vars[vname].value
                        # lowercases the string
                        if objfunc == 'lower':
                            self.vars[vname].value = self.vars[vname].value.lower()
                        # cuts a string to the two indices passed
                        if objfunc == 'cut':
                            start = self.parse(0, line, f, sp, args)[2]
                            end = self.parse(1, line, f, sp, args)[2]
                            # both start and end must be ints
                            self.type_err(
                                [(start, (int,)), (end, (int,))], line, lines_ran)
                            self.vars[vname].value = self.vars[vname].value[start:end]
                            return self.vars[vname].value
                        # inserts a string into a position
                        if objfunc == 'shove':
                            inserting = self.parse(0, line, f, sp, args)[2]
                            index = self.parse(1, line, f, sp, args)[2]
                            # index must be an int
                            self.type_err([(index, (int,))], line, lines_ran)
                            # do the above concatination but with f strings
                            self.vars[
                                vname].value = f"{self.vars[vname].value[:index]}{inserting}{self.vars[vname].value[index:]}"
                            return self.vars[vname].value
                        # gets a string containing a certain amount of characters left and
                        # right of the first occurance of the string inside of a string
                        if objfunc == 'around':
                            # keyword to search for
                            keyword = self.parse(0, line, f, sp, args)[2]
                            # keyword must be a string
                            self.type_err([(keyword, (str,))], line, lines_ran)
                            # get the index of the keyword
                            index = object.find(keyword)
                            # if not found
                            if index == -1:
                                # f"around(): Keyword '{keyword}' not found in string"
                                # raise an msn2 error
                                self.err(
                                    f"around(): Keyword '{keyword}' not found in string", line, lines_ran, f)
                            # get the string
                            return object[index-self.parse(1, line, f, sp, args)[2]:index+len(keyword)+self.parse(2, line, f, sp, args)[2]]
                        # startswith
                        if objfunc == 'startswith':
                            st = self.parse(0, line, f, sp, args)[2]
                            # st must be a string
                            self.type_err([(st, (str,))], line, lines_ran)
                            return object.startswith(st)
                        # endswith
                        if objfunc == 'endswith':
                            st = self.parse(0, line, f, sp, args)[2]
                            # st must be a string
                            self.type_err([(st, (str,))], line, lines_ran)
                            return object.endswith(st)
                    # working with the HTML session
                    # object, checks string type to avoid importing before needed
                    elif str(type(object)) == "<class 'requests_html.HTMLSession'>":
                        # gets information from a website
                        if objfunc == 'all':
                            return object.get(self.parse(0, line, f, sp, args)[2]).html
                        # renders the html session
                        if objfunc == 'render':
                            return object.render(retries=3)
                        return object
                    # working with requests_html.HTML
                    # 2.0.384
                    elif str(type(object)) == "<class 'requests_html.HTML'>":
                        # finds elements in the HTML
                        if objfunc == 'gather':
                            # if one argument
                            if args[0][0] == '':
                                return object.find()
                            else:
                                return object.find(self.parse(0, line, f, sp, args)[2])
                        # return the attribute offered by the object itself
                        return getattr(object, objfunc)
                    # working with Excel sheets
                    elif isinstance(object, self.Sheet):
                        # active elements
                        title = object.title
                        workbook = object.workbook
                        path = object.path
                        sheet = object.sheet
                        # gets the value of a cell
                        if objfunc == 'get':
                            # column of the cell
                            column = self.parse(0, line, f, sp, args)[2]
                            # row of the cell
                            row = self.parse(1, line, f, sp, args)[2]
                            # returns the value of the cell
                            return sheet.cell(row + 1, column + 1).value
                        # sets the value of a cell
                        if objfunc == 'set':
                            # column of the cell
                            column = self.parse(0, line, f, sp, args)[2]
                            # row of the cell
                            row = self.parse(1, line, f, sp, args)[2]
                            # row and column must be int
                            self.type_err(
                                [(column, (int,)), (row, (int,))], line, lines_ran)
                            # value to set the cell to
                            value = self.parse(2, line, f, sp, args)[2]
                            # sets the value of the cell
                            sheet.cell(row + 1, column + 1, value)
                            # returns the sheet
                            return value
                        # clears the sheet
                        if objfunc == 'clear':
                            # clears the sheet
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cell.value = None
                            # returns the sheet
                            return object
                        # gets the populated cell values of a column
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'column':
                            # column, either an integer or string
                            col = self.parse(0, line, f, sp, args)[2]
                            # col must be int or str
                            self.type_err([(col, (int, str))], line, lines_ran)
                            column_values = []
                            # if number
                            if isinstance(col, int):
                                col += 1
                                column_values = [row.value for row in sheet.iter_cols(
                                    min_col=col, max_col=col) for row in cell if row.value != None]
                            # otherwise, get column by title
                            elif isinstance(col, str):
                                # for each column
                                column_values = [row.value for row in sheet.iter_cols(
                                ) if row[0].value == col for row in cell if row.value != None]
                            return column_values
                        # gets the populated cell values of a row
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'row':
                            # row, either an integer or string
                            r = self.parse(0, line, f, sp, args)[2]
                            # r must be int or str
                            self.type_err([(r, (int, str))], line, lines_ran)
                            row_values = []
                            # if number
                            if isinstance(r, int):
                                r += 1
                                row_values = [row.value for row in sheet.iter_rows(
                                    min_row=r, max_row=r) for row in cell if row.value != None]
                            # otherwise, get row by title
                            elif isinstance(r, str):
                                # for each row
                                row_values = [row.value for row in sheet.iter_rows(
                                ) if row[0].value == r for row in cell if row.value != None]
                            return row_values

                        # gets the index of a column with a given title
                        def get_column_index(title):
                            for cell in sheet.iter_cols():
                                if cell[0].value == title:
                                    return cell[0].column
                            return None

                        def get_row_index(title):
                            for cell in sheet.iter_rows():
                                if cell[0].value == title:
                                    return cell[0].row
                            return None
                        # rewrite the above method, but with
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        if objfunc == 'set_column':
                            # column, either an integer or string
                            col = self.parse(0, line, f, sp, args)[2]
                            # col must be int or str
                            self.type_err([(col, (int, str))], line, lines_ran)
                            # iterable of values
                            values = self.parse(1, line, f, sp, args)[2]
                            # check that values is an iterable
                            self.check_iterable(values, line)
                            # if number
                            if isinstance(col, int):
                                col += 1
                                for i in range(len(values)):
                                    sheet.cell(i + 1, col, values[i])
                            # otherwise, get column by title
                            elif isinstance(col, str):
                                # for each column
                                for cell in sheet.iter_cols():
                                    # if the title matches
                                    if cell[0].value == col:
                                        # get the column values
                                        for i in range(len(values)):
                                            sheet.cell(
                                                i + 1, get_column_index(col), values[i])
                            return values
                        # sets a row to an array of values
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'set_row':
                            # row, either an integer or string
                            r = self.parse(0, line, f, sp, args)[2]
                            # must be int or str
                            self.type_err([(r, (int, str))], line, lines_ran)
                            # array of values
                            values = self.parse(1, line, f, sp, args)[2]
                            # check that values is an iterable
                            self.check_iterable(values, line)
                            # if number
                            if isinstance(r, int):
                                r += 1
                                for i in range(len(values)):
                                    sheet.cell(r, i + 1, values[i])
                            # otherwise, get row by title
                            elif isinstance(r, str):
                                # for each row
                                for cell in sheet.iter_rows():
                                    # if the title matches
                                    if cell[0].value == r:
                                        # get the row values
                                        for i in range(len(values)):
                                            sheet.cell(get_row_index(
                                                r), i + 1, values[i])
                            return values
                        # reqrite the above method, but with
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'add_to_column':
                            # column
                            column = self.parse(0, line, f, sp, args)[2]
                            # column should be an int or str
                            self.type_err(
                                [(column, (int, str))], line, lines_ran)
                            # value to add
                            value = self.parse(1, line, f, sp, args)[2]
                            # if number
                            if isinstance(column, int):
                                column += 1
                                # find the first empty cell in the column
                                for i in range(sheet.max_row + 1):
                                    if sheet.cell(i + 1, column).value == None:
                                        sheet.cell(i + 1, column, value)
                                        return value
                                return value
                            # otherwise, get column by title
                            elif isinstance(column, str):
                                column_index = get_column_index(column)
                                # find the first empty cell in the column
                                for i in range(sheet.max_row + 1):
                                    if sheet.cell(i + 1, column_index).value == None:
                                        sheet.cell(i + 1, column_index, value)
                                        return value
                            return value
                        # adds a value to a row
                        # if the argument is a number, it gets the value of that column
                        # if the argument is a string, it gets the value of the column with that title
                        elif objfunc == 'add_to_row':
                            # row
                            row = self.parse(0, line, f, sp, args)[2]
                            # row should be an int or str
                            self.type_err([(row, (int, str))], line, lines_ran)
                            # value to add
                            value = self.parse(1, line, f, sp, args)[2]
                            # if number
                            if isinstance(row, int):
                                row += 1
                                # find the first empty cell in the row
                                for i in range(sheet.max_column):
                                    if sheet.cell(row, i + 1).value == None:
                                        sheet.cell(row, i + 1, value)
                                        return value
                                return value
                            # otherwise, get row by title
                            elif isinstance(row, str):
                                row_index = get_row_index(row)
                                # find the first empty cell in the row
                                for i in range(sheet.max_column):
                                    if sheet.cell(row_index, i + 1).value == None:
                                        sheet.cell(row_index, i + 1, value)
                                        return value
                            return value
                        # performs code for each row
                        # takes 2 arguments, a string for the row variable
                        # and a function to perform on each row
                        elif objfunc == 'each_row':
                            # variable
                            row_var = self.parse(0, line, f, sp, args)[2]
                            # check varname
                            self.check_varname(row_var, line)
                            ret = None
                            # for each cell
                            for cell in sheet.iter_rows():
                                # set the variable to the row
                                self.vars[row_var] = Var(
                                    row_var, [cell[i].value for i in range(len(cell))])
                                # execute the function
                                ret = self.interpret(args[1][0])
                            return ret
                        # writes a matrix (2D array) to this Excel Worksheet
                        # with the offsets given by the second and third arguments,
                        # if no second or third arguments, then the matrix is written
                        # starting at the first cell
                        elif objfunc == 'import_matrix':
                            # get the 2D list
                            matrix = self.parse(0, line, f, sp, args)[2]
                            # matrix must be a list
                            self.type_err([(matrix, (list,))], line, lines_ran)
                            # default offset
                            offsetx = 0
                            offsety = 0
                            # if there is a second argument
                            if len(args) == 2:
                                offsetx = self.parse(1, line, f, sp, args)[2]
                            if len(args) == 3:
                                offsetx = self.parse(1, line, f, sp, args)[2]
                                offsety = self.parse(2, line, f, sp, args)[2]
                            # for each row
                            for i in range(len(matrix)):
                                # for each column
                                for j in range(len(matrix[i])):
                                    w = matrix[i][j]
                                    # if w is an AppElement, write its name
                                    if 'name' in dir(w):
                                        w = w.name
                                    sheet.cell(i + offsety + 1,
                                               j + offsetx + 1, w)
                            return matrix
                        # if nothing else, return the object
                        return object
                    # working with Excel
                    elif isinstance(object, self.Workbook):
                        # active workbook
                        workbook = object.workbook
                        path = object.path
                        # gets or creates a sheet in the workbook
                        if objfunc == 'sheet':
                            # title of the new sheet
                            title = self.parse(0, line, f, sp, args)[2]
                            # title should be a string or int
                            self.type_err([(title, (str, int))],
                                          line, lines_ran)
                            # if title is a string
                            if isinstance(title, str):
                                # if the sheet has already been created,
                                # return the created sheet
                                for name in workbook.sheetnames:
                                    if name.lower() == title.lower():
                                        return self.Sheet(workbook[name], name, workbook, path)
                                # creates the sheet
                                sheet = workbook.create_sheet(title)
                                # returns the sheet
                                return self.Sheet(sheet, title, workbook, path)
                            # title is integer,
                            # return the sheet at that index
                            else:
                                for i, sheet in enumerate(workbook.sheetnames):
                                    if i == title:
                                        return self.Sheet(workbook[sheet], sheet, workbook, path)
                            return None
                        # saves the workbook
                        if objfunc == 'save':
                            workbook.save(path)
                            return object
                        # closes the workbook
                        if objfunc == 'close' or objfunc == 'stop' or objfunc == 'kill':
                            workbook.close()
                            return object
                        # otherwise return the object
                        return object
                    # GENERAL METHODS
                    # gets the immediate children of the parent window

                    def children(parent_window):
                        return [self.AppElement(child, child.window_text()) for child in window.children()]
                    # gets a child at an index
                    # prints the children

                    def child(parent_window, index):
                        child = children(parent_window)[index]
                        return self.AppElement(child, child.window_text())
                    # finds a child with subtext in its name

                    def find_children(parent_window, subtext):
                        subtext = subtext.lower()
                        return [self.AppElement(child, child.window_text()) for child in window.children()
                                if subtext in child.window_text().lower()]

                    # recursively searches the child tree for a certain object type
                    # dont allow ElementAmbiguousError

                    def recursive_search(parent_window, type, as_type, object_string_endswith=None):
                        found = []
                        # get the children
                        # use kwargs to avoid ElementAmbiguousError
                        # kwargs is a criteria to reduce a list by process, class_name, control_type, content_only and/or title.
                        kwargs = {'process': parent_window.process_id()}
                        c = parent_window.children(**kwargs)
                        for child in c:
                            if isinstance(child, type) or (object_string_endswith and str(child).endswith(object_string_endswith)):
                                found += [as_type(child, child.window_text())]
                            found += recursive_search(child, type,
                                                      as_type, object_string_endswith)
                        return found
                    # prints all elements

                    def print_elements(parent_window, retrieve_elements):
                        for i, element in enumerate(retrieve_elements(parent_window)):
                            print(i, ":")
                            print(element)
                        return None
                    # finds an element containing the substring specified

                    def find_elements(parent_window, subtext, retrieve_elements):
                        elements = []
                        subtext = subtext.lower()
                        for element in retrieve_elements(parent_window):
                            if subtext in element.name.lower():
                                elements.append(self.AppElement(
                                    element, element.window_text()))
                        return elements
                    # finds the exact elements specified

                    def find_elements_exact(parent_window, text, retrieve_elements):
                        elements = []
                        for element in retrieve_elements(parent_window):
                            if text == element.name:
                                elements.append(self.AppElement(
                                    element, element.window_text()))
                        return elements
                    # waits for the first element to appear containing the substring specified
                    # is not case sensitive

                    def wait_for_element_subtext(parent_window, retrieve_elements, subtext, timeout=None):
                        subtext = subtext.lower()
                        # subfunction for locating the element

                        def find_element_():
                            try:
                                for element in retrieve_elements(parent_window):
                                    if subtext in element.name.lower():
                                        return self.AppElement(element, element.window_text())
                            except:
                                pass
                        if not timeout:
                            while True:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                        else:
                            import time
                            # get the current time
                            start_time = time.time()
                            # while the time elapsed is less than the timeout
                            while time.time() - start_time < timeout:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                    # waits for the first element to appear with the exact text specified

                    def wait_for_element_exact(parent_window, retrieve_elements, text, timeout=None):
                        # subfunction for locating the element
                        def find_element_():
                            try:
                                for element in retrieve_elements(parent_window):
                                    if text == element.name:
                                        return self.AppElement(element, element.window_text())
                            except:
                                pass
                        if not timeout:
                            while True:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                        else:
                            import time
                            # get the current time
                            start_time = time.time()
                            # while the time elapsed is less than the timeout
                            while time.time() - start_time < timeout:
                                if (_ret := find_element_()) is not None:
                                    return _ret
                    # waits for the first element to appear in all children containing the substring specified with the type specified

                    def wait_for_type_subtext_all(parent_window, type, as_type, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, lambda parent_window: recursive_search(parent_window, type, as_type), subtext, timeout=timeout)
                    # wait for the first element to appear in all children with the exact text specified with the type specified

                    def wait_for_type_exact_all(parent_window, type, as_type, text, timeout=None):
                        return wait_for_element_exact(parent_window, lambda parent_window: recursive_search(parent_window, type, as_type), text, timeout=timeout)

                    # waits for a child to exist with text containing subtext

                    def wait_for_text(parent_window, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, children, subtext, timeout=timeout)
                    # waits for a child to exist in the entire child tree containing subtext

                    def wait_for_text_all(parent_window, subtext, timeout=None):
                        return wait_for_element_subtext(parent_window, all_children, subtext, timeout=timeout)
                    # waits for a child to exist with text exactly equal to text

                    def wait_for_text_exact(parent_window, text, timeout=None):
                        return wait_for_element_exact(parent_window, children, text, timeout=timeout)
                    # waits for a child to exist in the entire child tree with text exactly equal to text

                    def wait_for_text_exact_all(parent_window, text, timeout=None):
                        return wait_for_element_exact(parent_window, all_children, text, timeout=timeout)
                    # prints all children of a parent window

                    def print_children(parent_window):
                        return print_elements(parent_window, children)

                    # gets all children in the child tree
                    def all_children(parent_window):
                        found = []
                        for child in parent_window.children():
                            found.append(self.AppElement(
                                child, child.window_text()))
                            found += all_children(child)
                        return found
                    # prints the child tree of a parent window

                    def print_all_children(parent_window):
                        return print_elements(parent_window, all_children)
                    # gets from all children at an index

                    def all_child(parent_window, index):
                        return all_children(parent_window)[index]
                    # finds all children with subtext in their name

                    def find_all_children(parent_window, subtext):
                        return find_elements(parent_window, subtext, all_children)
                    # finds all children from an exact text

                    def find_all_children_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, all_children)

                    # ---------------------------

                    # NARROWING GENERAL METHODS
                    # recursively gets all menus existing in the parent_window tree
                    # accumulates all instances of pywinauto.controls.uia_controls.MenuWrapper

                    def menus(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.MenuWrapper, self.Menu)
                    # gets a single menu

                    def menu(parent_window, index):
                        return menus(parent_window)[index]
                    # prints all the menus

                    def print_menus(parent_window):
                        return print_elements(parent_window, menus)
                    # finds a menu with subtext in its name

                    def find_menus(parent_window, subtext):
                        return find_elements(parent_window, subtext, menus)

                    # gets all toolbars
                    def toolbars(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ToolbarWrapper, self.ToolBar)

                    def print_toolbars(parent_window):
                        return print_elements(parent_window, toolbars)

                    def toolbar(parent_window, index):
                        return toolbars(parent_window)[index]

                    def find_toolbars(parent_window, subtext):
                        return find_elements(parent_window, subtext, toolbars)

                    # recursively gets all instances of pywinauto.controls.uia_controls.ButtonWrapper
                    def buttons(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ButtonWrapper, self.Button)

                    def button(parent_window, index):
                        return buttons(parent_window)[index]

                    def print_buttons(parent_window):
                        return print_elements(parent_window, buttons)

                    def find_buttons(parent_window, subtext):
                        return find_elements(parent_window, subtext, buttons)

                    # for hyperlinks
                    def links(parent_window):
                        return recursive_search(parent_window, int, self.Link, object_string_endswith="Hyperlink")

                    def link(parent_window, index):
                        return links(parent_window)[index]

                    def print_links(parent_window):
                        return print_elements(parent_window, links)

                    def find_links(parent_window, subtext):
                        return find_elements(parent_window, subtext, links)

                    def find_links_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, links)

                    # for tabitems
                    def tabitems(parent_window):
                        return recursive_search(parent_window, int, self.TabItem, object_string_endswith="TabItem")

                    def tabitem(parent_window, index):
                        return tabitems(parent_window)[index]

                    def print_tabitems(parent_window):
                        return print_elements(parent_window, tabitems)

                    def find_tabitems(parent_window, subtext):
                        return find_elements(parent_window, subtext, tabitems)

                    def find_tabitems_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tabitems)

                    # for tabcontrols
                    def tabcontrols(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="TabControl")

                    def tabcontrol(parent_window, index):
                        return tabcontrols(parent_window)[index]

                    def print_tabcontrols(parent_window):
                        return print_elements(parent_window, tabcontrols)

                    def find_tabcontrols(parent_window, subtext):
                        return find_elements(parent_window, subtext, tabcontrols)

                    def find_tabcontrols_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tabcontrols)

                    # for EditWrapper
                    def inputs(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.EditWrapper, self.Input)

                    def input(parent_window, index):
                        return inputs(parent_window)[index]

                    def print_inputs(parent_window):
                        return print_elements(parent_window, inputs)

                    def find_inputs(parent_window, subtext):
                        return find_elements(parent_window, subtext, inputs)

                    def find_inputs_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, inputs)

                    # for ButtonWrapper but endswith CheckBox
                    def checkboxes(parent_window):
                        return recursive_search(parent_window, int, self.Button, object_string_endswith="CheckBox")

                    def checkbox(parent_window, index):
                        return checkboxes(parent_window)[index]

                    def print_checkboxes(parent_window):
                        return print_elements(parent_window, checkboxes)

                    def find_checkboxes(parent_window, subtext):
                        return find_elements(parent_window, subtext, checkboxes)

                    def find_checkboxes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, checkboxes)

                    # for Image
                    def images(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Image")

                    def image(parent_window, index):
                        return images(parent_window)[index]

                    def print_images(parent_window):
                        return print_elements(parent_window, images)

                    def find_images(parent_window, subtext):
                        return find_elements(parent_window, subtext, images)

                    def find_images_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, images)

                    # for Tables
                    def tables(parent_window):
                        return recursive_search(parent_window, int, self.Table, object_string_endswith="Table")

                    def table(parent_window, index):
                        return tables(parent_window)[index]

                    def print_tables(parent_window):
                        return print_elements(parent_window, tables)

                    def find_tables(parent_window, subtext):
                        return find_elements(parent_window, subtext, tables)

                    def find_tables_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, tables)

                    # for GroupBoxes
                    def groupboxes(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="GroupBox")

                    def groupbox(parent_window, index):
                        return groupboxes(parent_window)[index]

                    def print_groupboxes(parent_window):
                        return print_elements(parent_window, groupboxes)

                    def find_groupboxes(parent_window, subtext):
                        return find_elements(parent_window, subtext, groupboxes)

                    def find_groupboxes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, groupboxes)

                    # for Panes
                    def panes(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Pane")

                    def pane(parent_window, index):
                        return panes(parent_window)[index]

                    def print_panes(parent_window):
                        return print_elements(parent_window, panes)

                    def find_panes(parent_window, subtext):
                        return find_elements(parent_window, subtext, panes)

                    def find_panes_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, panes)

                    # for ListItems
                    def listitems(parent_window):
                        return recursive_search(parent_window, pywinauto.controls.uia_controls.ListItemWrapper, self.AppElement, object_string_endswith="ListItem")

                    def listitem(parent_window, index):
                        return listitems(parent_window)[index]

                    def print_listitems(parent_window):
                        return print_elements(parent_window, listitems)

                    def find_listitems(parent_window, subtext):
                        return find_elements(parent_window, subtext, listitems)

                    def find_listitems_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, listitems)

                    # for documents
                    def documents(parent_window):
                        return recursive_search(parent_window, int, self.AppElement, object_string_endswith="Document")

                    def document(parent_window, index):
                        return documents(parent_window)[index]

                    def print_documents(parent_window):
                        return print_elements(parent_window, documents)

                    def find_documents(parent_window, subtext):
                        return find_elements(parent_window, subtext, documents)

                    def find_documents_exact(parent_window, text):
                        return find_elements_exact(parent_window, text, documents)

                    # for decendants
                    def descendants(parent_window):
                        return recursive_search(parent_window, int, self.AppElement)
                    # ---------------------------
                    # GENERALIZING METHOD CALLS FOR ELEMENT DISCOVERY

                    def callables(window,
                                  # array elements
                                  objfunc1, objfunc1_method,
                                  # print the elements
                                  objfunc2, objfunc2_method,
                                  # get a certain element
                                  objfunc3, objfunc3_method,
                                  # find elements with subtext in their names
                                  objfunc4, objfunc4_method,
                                  # find elements with exact text in their names
                                  objfunc5=None, objfunc5_method=None,
                                  # waits for the first element of a certain type with subtext in name
                                  objfunc6=None, objfunc6_method=None, type1=None, as_type1=None,
                                  # waits for the first element of a certain type with exact text in name
                                  objfunc7=None, objfunc7_method=None, type2=None, as_type2=None,
                                  ):

                        # RETRIEVING CHILDREN
                        # gets the available child reference keywords
                        if objfunc == objfunc1:
                            return objfunc1_method(window)
                        # prints the children
                        if objfunc == objfunc2:
                            return objfunc2_method(window)
                        # gets a certain child
                        # first argument is the index of the child
                        if objfunc == objfunc3:
                            return objfunc3_method(window, self.parse(0, line, f, sp, args)[2])
                        # finds children with subtext in their names
                        if objfunc == objfunc4:
                            return objfunc4_method(window, self.parse(0, line, f, sp, args)[2])
                        if objfunc == objfunc5:
                            return objfunc5_method(window, self.parse(0, line, f, sp, args)[2])

                        # waits for the first child of a certain type with exact text in its name
                        if objfunc == objfunc6:
                            # if 1 argument, there is no timeout
                            if len(args) == 1:
                                return wait_for_type_exact_all(window, type1, as_type1, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                return wait_for_type_exact_all(window, type1, as_type1, self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2])
                        # waits for the first child of a certain type with subtext in its name
                        if objfunc == objfunc7:
                            # if 1 argument, there is no timeout
                            if len(args) == 1:
                                return wait_for_type_subtext_all(window, type2, as_type2, self.parse(0, line, f, sp, args)[2])
                            elif len(args) == 2:
                                return wait_for_type_subtext_all(window, type2, as_type2, self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2])

                        return '<msnint2 no callable>'

                    # ---------------------------

                    # moves the mouse to the center of an element, and clicks it

                    def clk(window, button='left', waittime=0):
                        import time
                        from pywinauto import mouse
                        # set the focus to this element
                        window.set_focus()
                        # wait for the element to be ready
                        time.sleep(waittime)
                        # get the new coordinates of this element after the focus
                        coords = window.get_properties()[
                            'rectangle'].mid_point()
                        # click the mouse
                        mouse.click(button=button, coords=coords)
                        # return the object
                        return object
                    # determines if a point is visible within a rectangle

                    def has_point(object, x, y):
                        try:
                            rect = object.get_properties()['rectangle']
                            # if implemented
                            return rect.top <= y <= rect.bottom and rect.left <= x <= rect.right
                        except:
                            print(str(object))
                            return True
                    # recursively get the first object that has the point
                    # the first object that has the point and no children

                    def rec(root, x, y):
                        # if the root has children
                        if root.children():
                            # for each child
                            for child in root.children():
                                # if the child has the point
                                if has_point(child, x, y):
                                    # return the child
                                    return rec(child, x, y)
                        # if the root has no children
                        else:
                            # return the root
                            return self.AppElement(root, root.window_text())

                    # get all objects that have the point
                    def get_all(root, x, y):
                        all = []
                        # if the root has children
                        if root.children():
                            # for each child
                            for child in root.children():
                                # if the child has the point
                                if has_point(child, x, y):
                                    # add the child to the list
                                    all.append(self.AppElement(
                                        child, child.window_text()))
                                    # get all of the child's children
                                    all += get_all(child, x, y)
                        # return the list
                        return all

                    # presses multiple keys at the same time
                    def press_simul(kys):
                        sending = ''
                        # keys down
                        for key in kys:
                            sending += '{' + key + ' down}'
                        # keys up
                        for key in kys:
                            sending += '{' + key + ' up}'
                        return sending

                    # function for converting keys requiring a shift press
                    #   example: a '3' should be converted to {VK_SHIFT down}3{VK_SHIFT up}
                    #   example: a '"' should be converted to {VK_SHIFT down}'{VK_SHIFT up}
                    #   example: a 'E' should be converted to {VK_SHIFT down}e{VK_SHIFT up}
                    # this function is mainly for converting an exerpt of code to a typable
                    # string for pywinauto to type
                    def convert_keys(keystrokes):
                        new = ''
                        special = {
                            '!': '1',
                            '@': '2',
                            '#': '3',
                            '$': '4',
                            '%': '5',
                            '^': '6',
                            '&': '7',
                            '*': '8',
                            '(': '9',
                            ')': '0',
                            '_': '-',
                            '+': '=',
                            '{': '[',
                            '}': ']',
                            '|': '\\',
                            ':': ';',
                            '"': "'",
                            '<': ',',
                            '>': '.',
                            '?': '/',
                            '~': '`',
                            ' ': ' '
                        }
                        # for each keystroke
                        for key in keystrokes:
                            if key == ' ':
                                # if the key is a space
                                new += ' '
                            elif key in special:
                                # if the key is a special character
                                new += '{VK_SHIFT down}' + \
                                    special[key] + '{VK_SHIFT up}'
                            elif key.isupper():
                                # if the key is uppercase
                                new += '{VK_SHIFT down}' + \
                                    key.lower() + '{VK_SHIFT up}'
                            else:
                                # if the key is not a special character
                                new += key
                        return new
                    # types keys with a delay between each key

                    def type_keys_with_delay(window, text, delay):
                        e = False
                        import time
                        import pywinauto
                        for char in text:
                            try:
                                window.type_keys(char, with_spaces=True)
                            except:
                                if not e:
                                    window.set_focus()
                                    e = True
                                pywinauto.keyboard.send_keys(char)
                            time.sleep(delay)
                    # parses object functions for discovering types
                    # of elements

                    def search(window):
                        import pywinauto
                        ret = '<msnint2 no callable>'
                        # RETRIEVING CHILDREN
                        # gets the available child reference keywords
                        if (chldrn := callables(window,
                                                'children', children,
                                                'print_children', print_children,
                                                'child', child,
                                                'find_children', find_children)) != '<msnint2 no callable>':
                            ret = chldrn
                        # working with the entire child tree
                        elif (all_chldrn := callables(window,
                                                      'all_children', all_children,
                                                      'print_all_children', print_all_children,
                                                      'all_child', all_child,
                                                      'find_all_children', find_all_children,
                                                      'find_all_children_exact', find_all_children_exact,
                                                      objfunc6='wait_for_child', objfunc6_method=wait_for_type_exact_all,
                                                      type1=pywinauto.controls.uiawrapper.UIAWrapper,
                                                      as_type1=self.AppElement,
                                                      objfunc7='wait_for_child_exact', objfunc7_method=wait_for_type_subtext_all,
                                                      type2=pywinauto.controls.uiawrapper.UIAWrapper,
                                                      as_type2=self.AppElement

                                                      )) != '<msnint2 no callable>':
                            ret = all_chldrn
                        # getting all menus
                        elif (mns := callables(window,
                                               'menus', menus,
                                               'print_menus', print_menus,
                                               'menu', menu,
                                               'find_menus', find_menus,
                                               objfunc5=None, objfunc5_method=None,
                                               objfunc6='wait_for_menu_exact', objfunc6_method=wait_for_type_exact_all,
                                               type1=pywinauto.controls.uia_controls.MenuWrapper,
                                               as_type1=self.Menu,
                                               objfunc7='wait_for_menu', objfunc7_method=wait_for_type_subtext_all,
                                               type2=pywinauto.controls.uia_controls.MenuWrapper,
                                               as_type2=self.Menu
                                               )) != '<msnint2 no callable>':
                            ret = mns
                        # gets all toolbars
                        elif (tbrs := callables(window,
                                                'toolbars', toolbars,
                                                'print_toolbars', print_toolbars,
                                                'toolbar', toolbar,
                                                'find_toolbars', find_toolbars,
                                                objfunc5=None, objfunc5_method=None,
                                                objfunc6='wait_for_toolbar_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=pywinauto.controls.uia_controls.ToolbarWrapper,
                                                as_type1=self.ToolBar,
                                                objfunc7='wait_for_toolbar', objfunc7_method=wait_for_type_subtext_all,
                                                type2=pywinauto.controls.uia_controls.ToolbarWrapper,
                                                as_type2=self.ToolBar
                                                )) != '<msnint2 no callable>':
                            ret = tbrs
                        # gets all buttons
                        elif (btns := callables(window,
                                                'buttons', buttons,
                                                'print_buttons', print_buttons,
                                                'button', button,
                                                'find_buttons', find_buttons,
                                                objfunc5=None, objfunc5_method=None,
                                                objfunc6='wait_for_button_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=pywinauto.controls.uia_controls.ButtonWrapper,
                                                as_type1=self.Button,
                                                objfunc7='wait_for_button', objfunc7_method=wait_for_type_subtext_all,
                                                type2=pywinauto.controls.uia_controls.ButtonWrapper,
                                                as_type2=self.Button
                                                )) != '<msnint2 no callable>':
                            ret = btns
                        # gets all tabitems
                        elif (tbs := callables(window,
                                               'tabitems', tabitems,
                                               'print_tabitems', print_tabitems,
                                               'tabitem', tabitem,
                                               'find_tabitems', find_tabitems,
                                               objfunc5=None, objfunc5_method=None,
                                               objfunc6='wait_for_tabitem_exact', objfunc6_method=wait_for_type_exact_all,
                                               type1=int,
                                               as_type1=self.TabItem,
                                               objfunc7='wait_for_tabitem', objfunc7_method=wait_for_type_subtext_all,
                                               type2=int,
                                               as_type2=self.TabItem
                                               )) != '<msnint2 no callable>':
                            ret = tbs
                        # gets all links
                        elif (lnks := callables(window,
                                                'links', links,
                                                'print_links', print_links,
                                                'link', link,
                                                'find_links', find_links,
                                                objfunc5=None, objfunc5_method=None,
                                                objfunc6='wait_for_link_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=int,
                                                as_type1=self.Hyperlink,
                                                objfunc7='wait_for_link', objfunc7_method=wait_for_type_subtext_all,
                                                type2=int,
                                                as_type2=self.Hyperlink
                                                )) != '<msnint2 no callable>':
                            ret = lnks
                        # gets all Inputs
                        elif (inpts := callables(window,
                                                 'inputs', inputs,
                                                 'print_inputs', print_inputs,
                                                 'input', input,
                                                 'find_inputs', find_inputs,
                                                 objfunc6='wait_for_input_exact', objfunc6_method=wait_for_type_exact_all,
                                                 type1=pywinauto.controls.uia_controls.EditWrapper,
                                                 as_type1=self.Input,
                                                 objfunc7='wait_for_input', objfunc7_method=wait_for_type_subtext_all,
                                                 type2=pywinauto.controls.uia_controls.EditWrapper,
                                                 as_type2=self.Input
                                                 )) != '<msnint2 no callable>':
                            ret = inpts
                        # gets all checkboxes
                        elif (chks := callables(window,
                                                'checkboxes', checkboxes,
                                                'print_checkboxes', print_checkboxes,
                                                'checkbox', checkbox,
                                                'find_checkboxes', find_checkboxes,
                                                objfunc6='wait_for_checkbox_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=pywinauto.controls.uia_controls.ButtonWrapper,
                                                as_type1=self.Button,
                                                objfunc7='wait_for_checkbox', objfunc7_method=wait_for_type_subtext_all,
                                                type2=pywinauto.controls.uia_controls.ButtonWrapper,
                                                as_type2=self.Button
                                                )) != '<msnint2 no callable>':
                            ret = chks
                        # gets all images
                        elif (imgs := callables(window,
                                                'images', images,
                                                'print_images', print_images,
                                                'image', image,
                                                'find_images', find_images)) != '<msnint2 no callable>':
                            ret = imgs
                        # gets all tables
                        elif (tbls := callables(window,
                                                'tables', tables,
                                                'print_tables', print_tables,
                                                'table', table,
                                                'find_tables', find_tables,
                                                objfunc6='wait_for_table_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=pywinauto.controls.uia_controls.ListViewWrapper,
                                                as_type1=self.Table,
                                                objfunc7='wait_for_table', objfunc7_method=wait_for_type_subtext_all,
                                                type2=pywinauto.controls.uia_controls.ListViewWrapper,
                                                as_type2=self.Table
                                                )) != '<msnint2 no callable>':
                            ret = tbls
                        # get all GroupBoxes
                        elif (grps := callables(window,
                                                'groupboxes', groupboxes,
                                                'print_groupboxes', print_groupboxes,
                                                'groupbox', groupbox,
                                                'find_groupboxes', find_groupboxes,
                                                objfunc6='wait_for_groupbox_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=int,
                                                as_type1=self.AppElement,
                                                objfunc7='wait_for_groupbox', objfunc7_method=wait_for_type_subtext_all,
                                                type2=int,
                                                as_type2=self.AppElement
                                                )) != '<msnint2 no callable>':
                            ret = grps
                        # for Panes
                        elif (pns := callables(window,
                                               'panes', panes,
                                               'print_panes', print_panes,
                                               'pane', pane,
                                               'find_panes', find_panes,
                                               objfunc6='wait_for_pane_exact', objfunc6_method=wait_for_type_exact_all,
                                               type1=int,
                                               as_type1=self.AppElement,
                                               objfunc7='wait_for_pane', objfunc7_method=wait_for_type_subtext_all,
                                               type2=int,
                                               as_type2=self.AppElement
                                               )) != '<msnint2 no callable>':
                            ret = pns
                        # for ListItems
                        elif (lsts := callables(window,
                                                'listitems', listitems,
                                                'print_listitems', print_listitems,
                                                'listitem', listitem,
                                                'find_listitems', find_listitems,
                                                objfunc6='wait_for_listitem_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=pywinauto.controls.uia_controls.ListItemWrapper,
                                                as_type1=self.AppElement,
                                                objfunc7='wait_for_listitem', objfunc7_method=wait_for_type_subtext_all,
                                                type2=pywinauto.controls.uia_controls.ListItemWrapper,
                                                as_type2=self.AppElement
                                                )) != '<msnint2 no callable>':
                            ret = lsts
                        # for TabControls
                        elif (tabs := callables(window,
                                                'tabcontrols', tabcontrols,
                                                'print_tabcontrols', print_tabcontrols,
                                                'tabcontrol', tabcontrol,
                                                'find_tabcontrols', find_tabcontrols,
                                                objfunc6='wait_for_tabcontrol_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=int,
                                                as_type1=self.AppElement,
                                                objfunc7='wait_for_tabcontrol', objfunc7_method=wait_for_type_subtext_all,
                                                type2=int,
                                                as_type2=self.AppElement
                                                )) != '<msnint2 no callable>':
                            ret = tabs
                        # for Documents
                        elif (docs := callables(window,
                                                'documents', documents,
                                                'print_documents', print_documents,
                                                'document', document,
                                                'find_documents', find_documents,
                                                objfunc6='wait_for_document_exact', objfunc6_method=wait_for_type_exact_all,
                                                type1=int,
                                                as_type1=self.AppElement,
                                                objfunc7='wait_for_document', objfunc7_method=wait_for_type_subtext_all,
                                                type2=int,
                                                as_type2=self.AppElement
                                                )) != '<msnint2 no callable>':
                            ret = docs
                        return ret

                    # if the object is a pywinauto application
                    # KNOWN ISSUES:
                    #   - I've tested this on a Windows 11 laptop and it doesn't
                    #     work for some reason
                    if isinstance(object, self.App):
                        # return for an app
                        ret = object
                        # path to the application to work with
                        path = object.path
                        # actual pwinauto application object
                        app = object.application
                        # window
                        window = app.window() if app else None
                        # thread based operation
                        p_thread = False
                        if objfunc.endswith(':lock'):
                            p_thread = True
                            objfunc = objfunc[:-5]
                            auto_lock.acquire()
                        # element discovery with search()
                        if (srch := search(window)) != '<msnint2 no callable>':
                            ret = srch
                        # STARTING AND STOPPING APPLICATIONS
                        if objfunc == 'start':
                            from pywinauto.application import Application
                            # create and start the application
                            if not object.application:
                                object.application = Application(
                                    backend="uia").start(path)
                            # add to global apps
                            global apps
                            apps[len(apps) + 1] = object
                            ret = object.application
                        # kills the application
                        elif objfunc == 'stop' or objfunc == 'kill' or objfunc == 'close':
                            # kill the application
                            ret = app.kill()
                        # gets the top_window
                        elif objfunc == 'print_tree':
                            ret = app.dump_tree()
                        # gets a connection to this application
                        elif objfunc == 'connection':
                            from pywinauto.application import Application
                            ret = self.App(object.path, Application(
                                backend="uia").connect(process=object.application.process))
                        # gets information about this application
                        # gets the text of the window
                        elif objfunc == 'text':
                            ret = window.window_text()
                        # gets the window
                        elif objfunc == 'window':
                            ret = window
                        # gets the handle
                        elif objfunc == 'handle':
                            ret = window.handle
                        # chrome based children collection

                        def chrome_children_():
                            chrome_window = app.window(title_re='.*Chrome.')
                            chrome_handle = chrome_window.handle
                            wd = app.window(handle=chrome_handle)
                            document = wd.child_window(
                                found_index=0, class_name='Chrome_RenderWidgetHostHWND')
                            return document.descendants()
                        # GOOGLE CHROME ONLY
                        if objfunc == 'chrome_children':
                            # if not arguments
                            if args[0][0] == '':
                                ret = chrome_children_()
                            # if one argument, check if the first argument is contained
                            elif len(args) == 1:
                                subtext = self.parse(0, line, f, sp, args)[
                                    2].lower()
                                # subtext must be str
                                self.type_err(
                                    [(subtext, (str,))], line, lines_ran)
                                ret = [self.AppElement(d, d.window_text()) for d in chrome_children_(
                                ) if subtext in d.window_text().lower()]
                            # if two arguments, check if the first argument is exact
                            elif len(args) == 2:
                                subtext = self.parse(0, line, f, sp, args)[2]
                                # subtext must be str
                                self.type_err(
                                    [(subtext, (str,))], line, lines_ran)
                                ret = [self.AppElement(d, d.window_text()) for d in chrome_children_(
                                ) if subtext == d.window_text()]

                        # waits for a child containing text
                        elif objfunc == 'wait_for_text':
                            # if no timeout provided
                            if len(args) == 1:
                                txt = self.parse(0, line, f, sp, args)[2]
                                # text should be str
                                self.type_err([(txt, (str,))], line, lines_ran)
                                ret = wait_for_text(
                                    window, txt)
                            # if timeout provided
                            elif len(args) == 2:
                                txt = self.parse(0, line, f, sp, args)[2]
                                timeout = self.parse(1, line, f, sp, args)[2]
                                # text should be str and timeout should be float or int or complex
                                self.type_err(
                                    [(txt, (str,)), (timeout, (float, int, complex))], line, lines_ran)
                                ret = wait_for_text(window, txt,
                                                    timeout=timeout)
                        # waits for a child containing text in the entire child tree
                        elif objfunc == 'wait_for_text_all':
                            # if no timeout provided
                            if len(args) == 1:
                                txt = self.parse(0, line, f, sp, args)[2]
                                # text should be str
                                self.type_err([(txt, (str,))], line, lines_ran)
                                ret = wait_for_text_all(
                                    window, txt)
                            elif len(args) == 2:
                                txt = self.parse(0, line, f, sp, args)[2]
                                timeout = self.parse(1, line, f, sp, args)[2]
                                # text should be str and timeout should be float or int or complex
                                self.type_err(
                                    [(txt, (str,)), (timeout, (float, int, complex))], line, lines_ran)
                                ret = wait_for_text_all(window, txt,
                                                        timeout=timeout)

                        # waits for a child containing the exact text
                        elif objfunc == 'wait_for_text_exact':
                            # if no timeout provided
                            if len(args) == 1:
                                txt = self.parse(0, line, f, sp, args)[2]
                                # text should be str
                                self.type_err([(txt, (str,))], line, lines_ran)
                                ret = wait_for_text_exact(
                                    window, txt)
                            elif len(args) == 2:
                                txt = self.parse(0, line, f, sp, args)[2]
                                timeout = self.parse(1, line, f, sp, args)[2]
                                # text should be str and timeout should be float or int or complex
                                self.type_err(
                                    [(txt, (str,)), (timeout, (float, int, complex))], line, lines_ran)
                                ret = wait_for_text_exact(window, txt,
                                                          timeout=timeout)
                        # waits for a child containing the exact text in the entire child tree
                        elif objfunc == 'wait_for_text_exact_all':
                            # if no timeout provided
                            if len(args) == 1:
                                txt = self.parse(0, line, f, sp, args)[2]
                                # text should be str
                                self.type_err([(txt, (str,))], line, lines_ran)
                                ret = wait_for_text_exact_all(
                                    window, txt)
                            elif len(args) == 2:
                                txt = self.parse(0, line, f, sp, args)[2]
                                timeout = self.parse(1, line, f, sp, args)[2]
                                # text should be str and timeout should be float or int or complex
                                self.type_err(
                                    [(txt, (str,)), (timeout, (float, int, complex))], line, lines_ran)
                                ret = wait_for_text_exact_all(window, txt,
                                                              timeout=timeout)

                        # APPLICATION ACTIONS
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'write':
                            writing = self.parse(0, line, f, sp, args)[2]
                            # writing should be a str
                            self.type_err([(writing, (str,))], line, lines_ran)
                            try:
                                # sends keystrokes to the application
                                ret = window.type_keys(
                                    writing, with_spaces=True)
                            except:
                                # with_spaces not allowed
                                ret = window.type_keys(
                                    convert_keys(writing), with_spaces=True)
                        # writes special characters into the console
                        # takes one argument, being the special characters to write
                        elif objfunc == 'write_special':
                            # keystrokes
                            keystrokes = self.parse(0, line, f, sp, args)[2]
                            # keystrokes should be a str
                            self.type_err(
                                [(keystrokes, (str,))], line, lines_ran)
                            # convert to special characters
                            ret = window.type_keys(convert_keys(
                                keystrokes), with_spaces=True)

                        # presses keys at the same time
                        elif objfunc == 'press':
                            kys = []
                            for i in range(len(args)):
                                kys.append(self.parse(i, line, f, sp, args)[2])
                            # presses the keys at the same time
                            ret = window.type_keys(press_simul(kys))
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'send_keys':
                            # import pywinauto.keyboard
                            import pywinauto
                            keystrokes = self.parse(0, line, f, sp, args)[2]
                            # keystrokes should be a str
                            self.type_err(
                                [(keystrokes, (str,))], line, lines_ran)
                            # sends keystrokes to the application
                            ret = pywinauto.keyboard.send_keys(convert_keys(
                                keystrokes), with_spaces=True)

                        # gets the element that is currently hovered over
                        # recurses through all children, determining which elements have
                        # the mouses position
                        elif objfunc == 'hovered':
                            # import win32api
                            import win32api
                            # get the root window of this application
                            root = window.top_level_parent()
                            # get the current mouse position
                            x, y = win32api.GetCursorPos()
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = get_all(root, x, y)
                        # opens the developer tools
                        elif objfunc == 'inspect':
                            # presses the shortcut keys to open the developer tools
                            ret = window.type_keys('{F12}')
                            # waits for the inspect window to appear
                            wait_for_text_all(window, 'Console')
                        # closes the developer tools
                        elif objfunc == 'close_inspect':
                            # presses the shortcut keys to close the developer tools
                            ret = window.type_keys('{F12}')
                        # refreshes the page
                        elif objfunc == 'refresh':
                            # presses the shortcut keys to refresh the page
                            ret = window.type_keys('{F5}')
                        # presses the enter key
                        elif objfunc == 'enter':
                            # presses the enter key
                            ret = window.type_keys('{ENTER}')
                        # presses the escape key
                        elif objfunc == 'escape':
                            # presses the escape key
                            ret = window.type_keys('{ESC}')
                        # page down
                        elif objfunc == 'page_down':
                            # presses the page down key
                            ret = window.type_keys('{PGDN}')
                        # page up
                        elif objfunc == 'page_up':
                            # presses the page up key
                            ret = window.type_keys('{PGUP}')
                        # release auto_lock
                        if p_thread:
                            auto_lock.release()
                        # return the object
                        return ret

                    # if the object is a pywinauto window element
                    elif isinstance(object, self.AppElement):
                        # returning
                        ret = object
                        # get the window of the AppElement object
                        window = object.window
                        # get the text of the AppElement object
                        name = object.name
                        # function to move the mouse from start to end,
                        # with a speed of speed
                        def movemouse(start, end, speed):
                            import time
                            from pywinauto import mouse
                            # reverse the speed, so a speed of 50 gives
                            # end_range of 50, and a speed of 75 gives
                            # end_range of 25
                            # dragging the mouse
                            # presses the mouse down at the coordinates
                            mouse.press(coords=start)
                            end_range = 100 - speed
                            for i in range(0, end_range):
                                mouse.move(coords=(int(start[0] + (end[0] - start[0]) / 100 * i),
                                                   int(start[1] + (end[1] - start[1]) / 100 * i)))
                                time.sleep(0.001)

                            # releases the mouse at the end coordinates
                            mouse.release(coords=end)
                        p_thread = False
                        # thread based functions
                        # if the function is a thread based function
                        if objfunc.endswith(':lock'):
                            p_thread = True
                            auto_lock.acquire()
                            objfunc = objfunc[:-5]
                        # OBTAINING DIFFERENT TYPES OF CHILDREN
                        # get the element window
                        if objfunc == 'window':
                            ret = window
                        # element discovery with search()
                        if (srch := search(window)) != '<msnint2 no callable>':
                            ret = srch
                        # getting information about the current window
                        # gets the window text
                        elif objfunc == 'text':
                            ret = window.window_text()
                        # GETTING LOCATION OF THE WINDOW
                        elif objfunc == 'top':
                            ret = window.get_properties()['rectangle'].top
                        elif objfunc == 'bottom':
                            ret = window.get_properties()['rectangle'].bottom
                        elif objfunc == 'left':
                            ret = window.get_properties()['rectangle'].left
                        elif objfunc == 'right':
                            ret = window.get_properties()['rectangle'].right
                        elif objfunc == 'center' or objfunc == 'mid_point':
                            ret = window.get_properties(
                            )['rectangle'].mid_point()
                        # getting the rectangle overall
                        elif objfunc == 'rectangle':
                            ret = [window.get_properties()['rectangle'].top, window.get_properties()[
                                'rectangle'].bottom, window.get_properties()['rectangle'].left, window.get_properties()['rectangle'].right]
                        # computes the diameter of the window
                        elif objfunc == 'width':
                            try:
                                left = window.get_properties()[
                                    'rectangle'].left
                                right = window.get_properties()[
                                    'rectangle'].right
                                ret = right - left
                            except:
                                ret = None
                        # computes the height of the window
                        elif objfunc == 'height':
                            try:
                                top = window.get_properties()['rectangle'].top
                                bottom = window.get_properties()[
                                    'rectangle'].bottom
                                ret = bottom - top
                            except:
                                ret = None
                        # getting adjacent elements
                        # could or could not be decendants
                        # operation is very slow, should be used mainly
                        # for element discovery
                        elif objfunc == 'element_above':
                            from pywinauto import mouse
                            # pixels above
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # pixels should be int
                            self.type_err(
                                [(pixels, (int,))], line, lines_ran)
                            # get the root window of this application
                            root = object.top_level_parent()
                            # get the top middle point of this element
                            top = object.get_properties(
                            )['rectangle'].top - pixels
                            mid = object.get_properties(
                            )['rectangle'].mid_point()[0]
                            # if there exist two arguments, move the mouse to that location
                            if len(args) == 2:
                                mouse.move(coords=(mid, top))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, mid, top)
                        elif objfunc == 'element_below':
                            from pywinauto import mouse
                            # pixels above
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # pixels should be int
                            self.type_err(
                                [(pixels, (int,))], line, lines_ran)
                            # get the root window of this application
                            root = object.top_level_parent()
                            # get the top middle point of this element
                            bottom = object.get_properties(
                            )['rectangle'].bottom + pixels
                            mid = object.get_properties(
                            )['rectangle'].mid_point()[0]
                            if len(args) == 2:
                                mouse.move(coords=(mid, bottom))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, mid, bottom)
                        elif objfunc == 'element_left':
                            from pywinauto import mouse
                            # pixels to the left
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # pixels should be int
                            self.type_err(
                                [(pixels, (int,))], line, lines_ran)
                            # get the root window of this application
                            root = object.top_level_parent()
                            # get the left middle point of this element
                            left = object.get_properties(
                            )['rectangle'].left - pixels
                            mid = object.get_properties(
                            )['rectangle'].mid_point()[1]
                            if len(args) == 2:
                                mouse.move(coords=(left, mid))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, left, mid)
                        elif objfunc == 'element_right':
                            from pywinauto import mouse
                            # pixels to the right
                            pixels = self.parse(0, line, f, sp, args)[2]
                            # pixels should be int
                            self.type_err(
                                [(pixels, (int,))], line, lines_ran)
                            # get the root window of this application
                            root = object.top_level_parent()
                            # get the right middle point of this element
                            right = object.get_properties(
                            )['rectangle'].right + pixels
                            mid = object.get_properties(
                            )['rectangle'].mid_point()[1]
                            if len(args) == 2:
                                mouse.move(coords=(right, mid))
                            # recursively find all children from the root window
                            # that have the point specified
                            ret = rec(root, right, mid)
                        # focus on the window
                        elif objfunc == 'focus':
                            ret = window.set_focus()
                        # scrolls to the window
                        elif objfunc == 'scroll':
                            from pywinauto import mouse
                            ret = mouse.scroll(coords=(window.get_properties()['rectangle'].mid_point()[0],
                                                       window.get_properties()['rectangle'].mid_point()[1]))
                        # drags this element to either another AppElement
                        elif objfunc == 'drag':
                            # if one argument and that argument isinstance(AppElement)
                            first = self.parse(0, line, f, sp, args)[2]
                            # first should be AppElement
                            self.type_err(
                                [(first, (self.AppElement,))], line, lines_ran)
                            # midpoint of the element to drag to
                            start = (window.get_properties()['rectangle'].mid_point()[0],
                                     window.get_properties()['rectangle'].mid_point()[1])
                            end = (first.get_properties()['rectangle'].mid_point()[0],
                                   first.get_properties()['rectangle'].mid_point()[1])

                            # slowly moves the mouse to the end coordinates
                            # this is to prevent the mouse from moving too fast
                            # and not dragging the object
                            # the farther the distance, the longer it takes
                            # to move the mouse
                            speed = 50
                            if len(args) == 2:
                                speed = self.parse(1, line, f, sp, args)[2]
                                # speed should be int
                                self.type_err(
                                    [(speed, (int,))], line, lines_ran)
                            # drags the mouse
                            movemouse(start, end, speed)
                            ret = True
                        # drags this AppElement to coordinates
                        elif objfunc == 'drag_coords':
                            start = (window.get_properties()['rectangle'].mid_point()[0],
                                     window.get_properties()['rectangle'].mid_point()[1])
                            startcoord = self.parse(0, line, f, sp, args)[2]
                            endcoord = self.parse(1, line, f, sp, args)[2]
                            # both startcoord and endcoord should be int
                            self.type_err(
                                [(startcoord, (int,)), (endcoord, (int,))], line, lines_ran)
                            end = (startcoord, endcoord)
                            # gets the speed, if specified
                            speed = 50
                            if len(args) == 3:
                                speed = self.parse(2, line, f, sp, args)[2]
                                # speed should be int
                                self.type_err(
                                    [(speed, (int,))], line, lines_ran)
                            # drags the mouse
                            movemouse(start, end, speed)
                            ret = True
                        # WINDOW ACTIONS
                        # sends keystrokes to the application
                        # takes one argument, being the keystrokes to send
                        elif objfunc == 'write':
                            writing = self.parse(0, line, f, sp, args)[2]
                            # writing should be str
                            self.type_err([(writing, (str,))], line, lines_ran)
                            timeout = False
                            # if a timeout between keystrokes is offered
                            if len(args) == 2:
                                timeout = True
                            if timeout:
                                delay = self.parse(1, line, f, sp, args)[2]
                                # delay should be float or int or complex
                                self.type_err(
                                    [(delay, (float, int, complex))], line, lines_ran)
                                ret = type_keys_with_delay(
                                    window, writing, delay)
                            else:
                                try:
                                    ret = window.type_keys(
                                        writing, with_spaces=True)
                                except:
                                    window.set_focus()
                                    ret = window.type_keys(writing)
                        # presses backspace
                        # if no arguments, presses it one time
                        # else, presses it the first argument many times
                        if objfunc == 'backspace':
                            window.set_focus()
                            # no argument
                            if args[0][0] == '':
                                ret = window.type_keys('{BACKSPACE}')
                            # else, send {BACKSPACE} that many times
                            else:
                                times = self.parse(0, line, f, sp, args)[2]
                                # times should be int
                                self.type_err(
                                    [(times, (int,))], line, lines_ran)
                                ret = window.type_keys('{BACKSPACE}' * times)
                        # presses the enter key
                        elif objfunc == 'enter':
                            ret = window.type_keys('{ENTER}')
                        # hovers over the window
                        elif objfunc == 'hover':
                            from pywinauto import mouse
                            # hovers the mouse over the window, using the mid point of the element
                            ret = mouse.move(
                                coords=(window.get_properties()['rectangle'].mid_point()))
                        # different types of AppElements
                        # if the appelement is a button
                        if isinstance(object, self.Button):
                            # clicks the button
                            if objfunc == 'click' or objfunc == 'left_click':
                                ret = object.click()
                            # left clicks the button
                            elif objfunc == 'right_click':
                                ret = object.right_click()
                            ret = object
                        # working with Links
                        elif isinstance(object, self.Link):
                            waittime = self.parse(0, line, f, sp, args)[
                                2] if args[0][0] != '' else 1

                            # waittime should be float or int or complex
                            self.type_err(
                                [(waittime, (float, int, complex))], line, lines_ran)
                            # clicks the link
                            if objfunc == 'click' or objfunc == 'left_click':
                                ret = clk(window, waittime=waittime)
                            # right clicks the link
                            elif objfunc == 'right_click':
                                ret = clk(window, button='right',
                                          waittime=waittime)
                            ret = object
                        # working with Tables
                        elif isinstance(object, self.Table):
                            # get table
                            table = object.window
                            # gets a row by index, based on the above logic
                            def row(index):
                                row = []
                                items = []
                                try:
                                    cols = table.column_count()
                                except NotImplementedError:
                                    # not implemented
                                    cols = 5
                                    items = table.items()
                                for i in range(cols):
                                    try:
                                        try:
                                            wrapper = table.cell(
                                                row=index, column=i)
                                        except:
                                            # table.items() gets a 1D list of items,
                                            # compute the index of the item
                                            # based on 'i' and 'index'
                                            wrapper = items[i + index * cols]

                                        row.append(self.AppElement(
                                            wrapper, wrapper.window_text()))
                                    except:
                                        break
                                return row
                            # gets a column by index
                            def col(index):
                                col = []
                                for i in range(table.column_count()):
                                    try:
                                        wrapper = table.cell(
                                            row=i, column=index)
                                        col.append(self.AppElement(
                                            wrapper, wrapper.window_text()))
                                    except:
                                        break
                                return col

                            # gets a cell at a row and column
                            if objfunc == 'get':
                                # get column
                                col = self.parse(0, line, f, sp, args)[2]
                                # get row
                                row = self.parse(1, line, f, sp, args)[2]
                                # column and row should be int
                                self.type_err(
                                    [(col, (int,)), (row, (int,))], line, lines_ran)
                                wrapper = table.cell(row=row, column=col)
                                # gets the cell
                                ret = self.AppElement(
                                    wrapper, wrapper.window_text())
                            # try to accumulate all the rows
                            # up to sys.maxsize
                            elif objfunc == 'matrix':
                                import sys
                                matrix = []
                                for i in range(sys.maxsize):
                                    try:
                                        if (_r := row(i)):
                                            matrix.append(_r)
                                        else:
                                            break
                                    except:
                                        break
                                ret = matrix
                            # gets a row
                            elif objfunc == 'row':
                                ind = self.parse(0, line, f, sp, args)[2]
                                # ind should be int
                                self.type_err([(ind, (int,))], line, lines_ran)
                                ret = row(ind)
                            # gets a column
                            elif objfunc == 'column':
                                ind = self.parse(0, line, f, sp, args)[2]
                                # ind should be int
                                self.type_err([(ind, (int,))], line, lines_ran)
                                ret = col(ind)
                        # working with ToolBars
                        elif isinstance(object, self.ToolBar):
                            toolbar_window = object.window
                            # gets the buttons of the toolbar
                            if objfunc == 'buttons':
                                ret = [toolbar_window.button(i) for i in range(
                                    toolbar_window.button_count())]
                            # prints the buttons of this toolbar
                            if objfunc == 'print_buttons':
                                for i in range(toolbar_window.button_count()):
                                    print(i, ':', toolbar_window.button(i))
                                ret = None
                            # gets a button at an index
                            if objfunc == 'button':
                                ret = toolbar_window.button(
                                    self.parse(0, line, f, sp, args)[2])
                            # finds all buttons with subtext in their names
                            if objfunc == 'find_buttons':
                                txt = self.parse(0, line, f, sp, args)[2]
                                # txt should be str
                                self.type_err([(txt, (str,))], line, lines_ran)
                                ret = find_buttons(
                                    toolbar_window, txt)
                            ret = object
                        # working with scrollbars
                        elif isinstance(object, self.ScrollBar):
                            scrollbar_window = object.window
                            if objfunc == 'scroll_down':
                                ret = scrollbar_window.scroll_down(
                                    amount='page', count=1)
                        # extra methods such that this AppElement requires different logic
                        if objfunc == 'click' or objfunc == 'left_click':
                            waittime = self.parse(0, line, f, sp, args)[
                                2] if args[0][0] != '' else 1
                            # waittime must be float or int or complex
                            self.type_err(
                                [(waittime, (float, int, complex))], line, lines_ran)
                            ret = clk(window, waittime=waittime)
                        elif objfunc == 'right_click':
                            waittime = self.parse(0, line, f, sp, args)[
                                2] if args[0][0] != '' else 1
                            # waittime must be float or int or complex
                            self.type_err(
                                [(waittime, (float, int, complex))], line, lines_ran)
                            ret = clk(window, button='right', waittime=waittime)
                        # if thread based, release the lock
                        if p_thread:
                            auto_lock.release()
                        return ret
                    # if the object is a dictionary
                    elif isinstance(object, dict):
                        # allows for repetitive setting on a multiple indexed dictionary
                        if objfunc == 'set':
                            self.vars[vname].value[self.parse(0, line, f, sp, args)[
                                2]] = self.parse(1, line, f, sp, args)[2]
                            return self.vars[vname].value
                        # first argument is what to set, should be called to_set
                        # rest of the arguments are the indices at which to index the object and set to_set
                        if objfunc == 'setn':
                            # what to set
                            to_set = self.parse(0, line, f, sp, args)[2]
                            # the rest of the arguments are the indices
                            # example: dict.setn('im being set', 'index1', 'index2', 'index3', ...)
                            # should equal: dict['index1']['index2']['index3'] = 'im being set'
                            # the object to set
                            obj = self.vars[vname].value
                            # iterates through the indices
                            for i in range(1, len(args)):
                                # if the index is the last one
                                if i == len(args) - 1:
                                    # sets the index to to_set
                                    obj[self.parse(i, line, f, sp, args)[
                                        2]] = to_set
                                # if the index is not the last one
                                else:
                                    # sets the object to the index
                                    obj = obj[self.parse(
                                        i, line, f, sp, args)[2]]
                            # returns the object
                            return self.vars[vname].value
                        # recursively gets a value in a dictionary
                        if objfunc == 'get':
                            # the object to get from
                            obj = self.vars[vname].value
                            # iterates through the indices
                            for i in range(len(args)):
                                ind = self.parse(i, line, f, sp, args)[2]
                                try:
                                    # sets the object to the index
                                    obj = obj[ind]
                                except KeyError:
                                    self.raise_key(ind, line)
                            # returns the object
                            return obj
                        # gets the keys of this dictionary
                        if objfunc == 'keys':
                            return self.vars[vname].value.keys()
                        # gets the values of this dictionary
                        if objfunc == 'values':
                            return self.vars[vname].value.values()
                        # gets the items of this dictionary
                        if objfunc == 'items':
                            return self.vars[vname].value.items()
                        # executes a function for each key-value pair
                        if objfunc == 'foreach':
                            # variable name of the key
                            keyname = self.parse(0, line, f, sp, args)[2]
                            # variable name of the value
                            valuename = self.parse(1, line, f, sp, args)[2]
                            # check both keyname and valuename as varnames
                            self.check_varname(keyname, line)
                            self.check_varname(valuename, line)
                            # function to execute
                            function = args[2][0]
                            # loop through the dictionary
                            for key, value in self.vars[vname].value.items():
                                # set the key and value variables
                                self.vars[keyname] = Var(keyname, key)
                                self.vars[valuename] = Var(valuename, value)
                                # execute the function
                                self.interpret(function)
                            # return the dictionary
                            return self.vars[vname].value
                        # maps each value in the dictionary to the output of the function
                        if objfunc == 'map':
                            # map arguments
                            keyvarname = self.parse(0, line, f, sp, args)[2]
                            valuevarname = self.parse(1, line, f, sp, args)[2]
                            # check both for varnames
                            self.check_varname(keyvarname, line)
                            self.check_varname(valuevarname, line)
                            function = args[2][0]
                            new_dict = {}
                            # loop through the objects items, assigning the key to the key and
                            # value to the value
                            for key, value in self.vars[vname].value.items():
                                # log old key
                                old_key = key
                                # execute the function
                                self.vars[keyvarname] = Var(keyvarname, key)
                                self.vars[valuevarname] = Var(
                                    valuevarname, value)
                                # run the function
                                ret = self.interpret(function)
                                if self.vars[keyvarname].value == old_key:
                                    new_dict[old_key] = self.vars[valuevarname].value
                                else:
                                    new_dict[self.vars[keyvarname].value] = self.vars[valuevarname].value
                            self.vars[vname].value = new_dict
                            return self.vars[vname].value

                # user function execution requested
                # user functions take priority
                # over general msn2 functions
                if func in self.methods:
                    method = self.methods[func]
                    # create func args
                    func_args = []
                    # if arguments supplied
                    if not args[0][0] == '':
                        for i in range(len(args)):
                            # check if we're setting a certain argument
                            as_s = args[i][0].strip()
                            meth_argname = None
                            if as_s[0] == '&':
                                func_args, meth_argname, arg, _ = self.split_named_arg(
                                    as_s, method, func_args)
                            # else, just append the argument
                            else:
                                arg = self.parse(i, line, f, sp, args)[2]
                                func_args.append(self.parse(
                                    i, line, f, sp, args)[2])
                            try:
                                meth_argname = method.args[i]
                            # incorrect amount of function arguments supplied
                            except IndexError:
                                self.raise_incorrect_args(str(len(method.args)), str(
                                    self.arg_count(args)), line, lines_ran, method)
                            try:
                                self.vars[meth_argname] = Var(
                                    meth_argname, arg)
                            # unhashable type:list, this means a named argument
                            # is being requested to be set
                            except TypeError:
                                self.vars[meth_argname[0]] = Var(
                                    meth_argname[0], arg)
                    # create return variable
                    ret_name = method.returns[0]
                    # add the return variable if not exists
                    if ret_name not in self.vars:
                        self.vars[ret_name] = Var(ret_name, None)
                    # execute method
                    try:
                        method.run(func_args, self, args)
                    # index out of bounds error in method run
                    except IndexError:
                        # raise msn2 error
                        self.raise_index_out_of_bounds(line, lines_ran, method)
                    # if its a variable
                    if ret_name in self.vars:
                        return self.vars[ret_name].value
                    try:
                        return eval(str(self.vars[ret_name].value), {}, {})
                    except:
                        pass
                    try:
                        return str(self.vars[ret_name].value)
                    except:
                        return str(self.vars[ret_name])

                # the below conditions interpret a line based on initial appearances
                # beneath these conditions will the Interpreter then parse the arguments from the line as a method call
                # request for Interpreter redirect to a block of code
                # the first argument is
                if func == 'redirect':
                    self.redirect_inside = []
                    # creates redirect for this interpreter
                    # check for type errors
                    linevar = self.parse(0, line, f, sp, args)[2]
                    self.type_err([(linevar, (str,))], line, lines_ran)
                    self.redirect = [linevar, args[1][0]]
                    self.redirecting = True
                    return self.redirect
                # request for Interpreter redirect cancellation
                if func == 'stopredirect':
                    self.redirecting = False
                    return True
                # starts the redirection
                if func == 'startredirect':
                    ret = None
                    for _ins in self.redirect_inside:
                        self.vars[self.redirect[0]] = Var(
                            self.redirect[0], _ins[1])
                        ret = self.interpret(_ins[0])
                    return ret
                # alternative to '~' user defined method syntax
                # example call:
                # function('sum')
                elif func == 'function':
                    # obtain the name of the function
                    fname = self.parse(0, line, f, sp, args)[2]
                    # function name must be a string
                    self.type_err([(fname, (str,))], line, lines_ran)
                    # function arguments
                    # create the new Method
                    new_method = self.Method(fname, self)
                    # add the body
                    new_method.add_body(args[1][0])
                    new_method.add_return(f"{fname}__return__")
                    # obtain the rest of the arguments as method args
                    for i in range(2, len(args)):
                        # adds variable name as an argument
                        # if any function specific argument is None, break
                        val = self.parse(i, line, f, sp, args)[2]
                        # val must be a string
                        self.type_err([(val, (str,))], line, lines_ran)
                        if val == None:
                            break
                        new_method.add_arg(val)
                    self.methods[fname] = new_method
                    return fname
                # simpler way to create a function
                elif func == 'def':
                    # get the name of the new function
                    name = self.parse(0, line, f, sp, args)[2]
                    # function name must be a string
                    self.type_err([(name, (str,))], line, lines_ran)
                    # create the new function
                    new_func = self.Method(name, self)
                    func_args = []
                    __temp = self.Method('', self)
                    # get the args for this function between the name and the body
                    # parse all args between the name and the body
                    for i in range(1, len(args) - 1):
                        # get the stripped string rep of this arg
                        as_s = args[i][0].strip()
                        if as_s[0] == '&':
                            func_args, meth_argname, _, ind = self.split_named_arg(
                                as_s, __temp, func_args)
                            # add argument with default value
                            new_func.add_arg([meth_argname, func_args[ind]])
                        else:
                            # add this argument to the method
                            new_func.add_arg(self.parse(
                                i, line, f, sp, args)[2])
                            new_arg = self.parse(i, line, f, sp, args)[2]
                            self.type_err([(new_arg, (str,))], line, lines_ran)
                    # add the body
                    new_func.add_body(f"ret('{name}',{args[-1][0]})")
                    # return buffer variable name
                    r_name = f"{name}__return__"
                    # if the return buffer doesn't exist, create it
                    if r_name not in self.vars:
                        # create the return variable
                        self.vars[r_name] = Var(r_name, None)
                    # add the return variable
                    new_func.add_return(r_name)
                    # add the function to the methods
                    self.methods[name] = new_func
                    # return the name of the function
                    return name
                # performs modular arithmetic on the two arguments given
                elif func == 'mod':
                    arg1 = self.parse(0, line, f, sp, args)[2]
                    arg2 = self.parse(1, line, f, sp, args)[2]
                    # verify both arguments are either int or float or complex or str
                    self.type_err(
                        [(arg1, (_allowed := (int, float, complex, str))), (arg2, allowed)], line, lines_ran)
                    # both types must be int or float
                    return arg1 % arg2
                # returns a value to a function
                # first argument is the function to return to
                # second argument is the value to return
                # ret() is used to create a return buffer for
                #   multiprogramming.
                elif func == 'ret':
                    name = self.parse(0, line, f, sp, args)[2]
                    # name must be a string
                    self.type_err([(name, (str,))], line, lines_ran)
                    # function to return to
                    vname = f"{name}__return__"
                    # value
                    value = self.parse(1, line, f, sp, args)[2]
                    # if the variable does not exist, we should create it
                    if vname not in self.vars:
                        self.vars[vname] = Var(vname, None)
                    self.vars[vname].value = value
                    return value
                # creating a list
                if func == 'arr' or func == 'from':
                    if args[0][0] == '':
                        return []
                    # return list comprehension
                    return [self.parse(i, line, f, sp, args)[2] for i in range(len(args))]
                # creates a dictionary from its arguments
                # every two arguments
                # an odd quantity of arguments is impossible
                if func == 'dictfrom' or func == 'object':
                    d = {}
                    if args[0][0] == '':
                        return d
                    # cannot have an odd number of arguments
                    if len(args) % 2 == 1:
                        self.err(
                            'Odd number of arguments in creating object',
                            f'An even number of arguments required to have valid key-value pairs.\nYou said: {len(args)} arg(s)',
                            line, lines_ran
                        )
                    # step over arguments in steps of 2
                    for i in range(0, len(args), 2):
                        d[self.parse(i, line, f, sp, args)[2]] = self.parse(
                            i + 1, line, f, sp, args)[2]
                    return d
                # splits the first argument by the second argument
                if func == 'split':
                    first = self.parse(0, line, f, sp, args)[2]
                    second = self.parse(1, line, f, sp, args)[2]
                    # first and second must be str
                    self.type_err([(first, (str,)), (second, (str,))], line, lines_ran)
                    return self.parse(0, line, f, sp, args)[2].split(self.parse(1, line, f, sp, args)[2])
                # obtains text between the first argument of the second argument
                if func == 'between':
                    # surrounding token
                    surrounding = self.parse(0, line, f, sp, args)[2]
                    # surrounding must be str
                    self.type_err([(surrounding, (str,))], line, lines_ran)
                    # string to analyze
                    string = self.parse(1, line, f, sp, args)[2]
                    # string must be str
                    self.type_err([(string, (str,))], line, lines_ran)
                    funccalls = []
                    try:
                        while string.count(surrounding) > 1:
                            string = string[string.index(
                                surrounding) + len(surrounding):]
                            funccalls.append(
                                string[:string.index(surrounding)])
                            string = string[string.index(
                                surrounding) + len(surrounding):]
                    except:
                        None
                    return funccalls
                # determines if the argument passed is of the type specified
                if func == 'isstr':
                    return isinstance(self.parse(0, line, f, sp, args)[2], str)
                elif func == 'islist':
                    return isinstance(self.parse(0, line, f, sp, args)[2], list)
                elif func == 'isfloat':
                    return isinstance(self.parse(0, line, f, sp, args)[2], float)
                elif func == 'isint':
                    return isinstance(self.parse(0, line, f, sp, args)[2], int)
                elif func == 'isdict':
                    return isinstance(self.parse(0, line, f, sp, args)[2], dict)
                elif func == 'isinstance':
                    return isinstance(self.parse(0, line, f, sp, args)[2], self.parse(1, line, f, sp, args)[2])

                # gets the sum of all arguments
                if func == 'sum':
                    total = 0
                    for i in range(len(args)):
                        try:
                            total += sum(self.parse(i, line, f, sp, args)[2])
                        except:
                            try:
                                total += self.parse(i, line, f, sp, args)[2]
                            except Exception as e:
                                self.err(
                                    'Error computing sum',
                                    e, line, lines_ran
                                )
                    return total
                # creates / sets a variable
                if func == 'var':
                    # extract varname
                    varname = self.parse(0, line, f, sp, args)[2]
                    # must be varname
                    self.check_varname(varname, line)
                    # extract value
                    value = self.parse(1, line, f, sp, args)[2]
                    # add / set variable
                    self.vars[varname] = Var(varname, value)
                    return value
                # converts the argument to a list
                elif func == 'list':
                    try:
                        return list(self.parse(0, line, f, sp, args)[2])
                    except:
                        return self.err(
                            'Casting error',
                            'Could not cast arg to a list',
                            line, lines_ran
                        )
                # absolute value
                elif func == 'abs':
                    try:
                        return abs(self.parse(0, line, f, sp, args)[2])
                    except:
                        return self.err(
                            'Error computing absolute value',
                            f'Could not compute absolute value of arg\nConsider changing arg to a number',
                            line, lines_ran
                        )
                # zip
                elif func == 'zip':
                    first = self.parse(0, line, f, sp, args)[2]
                    second = self.parse(1, line, f, sp, args)[2]
                    # verify both are iterable
                    self.check_iterable(first, line)
                    self.check_iterable(second, line)
                    return zip(first, second)
                # next
                elif func == 'next':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be iterable
                    self.check_iterable(arg, line)
                    return next(arg)
                # iter
                elif func == 'iter':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be iterable
                    self.check_iterable(arg, line)
                    return iter(arg)
                # determines if a variable exists or not
                elif func == 'exists':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    return arg in self.vars
                # determine if a function exists or not
                elif func == 'exists:function':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    return arg in self.methods
                # gets the length of the first argument
                elif func == 'len':
                    # get the first argument
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be iterable
                    self.check_iterable(arg, line)
                    return len(arg)
                # asserts each argument is True, prints and logs assertion error
                elif func == 'assert':
                    for i in range(len(args)):
                        assertion = self.parse(i, line, f, sp, args)[2]
                        if not assertion:
                            failed = ''
                            for arg in args:
                                failed += f"{assertion} "
                            self.err(
                                f"Assertion error in '{line}'", assertion, failed, lines_ran)
                    return True
                # asserts each argument throws an error
                elif func == 'assert:err':
                    for i in range(len(args)):
                        thrown = True
                        try:
                            # set self.trying to True
                            self.trying = True
                            # execute the line
                            ret = self.parse(i, line, f, sp, args)[2]
                            thrown = False
                            self.trying = False
                        except:
                            thrown = True
                        if not thrown:
                            self.err(
                                f"Assertion error, expected error", 'No error thrown where one was expected', line, lines_ran)
                    return True
                # trace capabilities
                elif obj == 'trace':
                    if objfunc == 'before':
                        numlines = self.parse(0, line, f, sp, args)[2]
                        # numlines must be int
                        self.type_err([(numlines, (int,))], line, lines_ran)
                        return lines_ran[len(lines_ran) - numlines:]
                    if objfunc == 'this':
                        return lines_ran[-1]
                    if objfunc == 'len':
                        return total_ints
                    return '<msnint2 class>'
                # gets msn2 settings information
                elif func == 'settings':
                    # returns the settings dict
                    return self.settings
                # referencing python variables
                elif obj == 'py':
                    # getting a variable from the current python environment
                    if objfunc == 'get':
                        vn = ''
                        name = self.parse(0, line, f, sp, args)[2]
                        # name must be str
                        self.type_err([(name, (str,))], line, lines_ran)
                        try:
                            return self._locals[name]
                        except KeyError:
                            self.no_var_err(
                                vn, 'local', 'local', self._locals, line)
                    # setting a variable in the current python environment
                    elif objfunc == 'set':
                        # name
                        vn = self.parse(0, line, f, sp, args)[2]
                        # vn must be str
                        self.type_err([(vn, (str,))], line, lines_ran)
                        value = self.parse(1, line, f, sp, args)[2]
                        self._locals[vn] = value
                        return value
                    # get the locals
                    elif objfunc == 'locals':
                        return self._locals
                    # gets a local variable
                    elif objfunc == 'local':
                        vn = self.parse(0, line, f, sp, args)[2]
                        # vn must be str
                        self.type_err([(vn, (str,))], line, lines_ran)
                        try:
                            return self._locals[vn]
                        except KeyError:
                            self.no_var_err(
                                vn, 'local', 'local', self._locals, line)
                    # get the globals
                    elif objfunc == 'globals':
                        return self._globals
                    # gets a global variable
                    elif objfunc == 'global':
                        vn = self.parse(0, line, f, sp, args)[2]
                        # vn must be str
                        self.type_err([(vn, (str,))], line, lines_ran)
                        try:
                            return self._globals[vn]
                        except:
                            self.no_var_err(
                                vn, 'global', 'global', self._globals, line)
                    # runs a stored python script
                    elif objfunc == 'run':
                        scr = self.parse(0, line, f, sp, args)[2]
                        # scr must be str
                        self.type_err([(scr, (str,))], line, lines_ran)
                        # execute the python and return
                        # the snippet with arguments inserted
                        return self.exec_python(scr)
                    # return a local variable within the python
                    # environment
                    else:
                        try:
                            # check in locals
                            return self._locals[objfunc]
                        except KeyError:
                            try:
                                # check in globals
                                return self._globals[objfunc]
                            except:
                                # raise error
                                self.no_var_err(objfunc, 'local or global',
                                                'local and global', self._globals, line)

                # casting
                try:
                    if func == 'int':
                        return int(self.parse(0, line, f, sp, args)[2])
                    elif func == 'float':
                        return float(self.parse(0, line, f, sp, args)[2])
                    elif func == 'str':
                        return str(self.parse(0, line, f, sp, args)[2])
                    elif func == 'bool':
                        return bool(self.parse(0, line, f, sp, args)[2])
                    elif func == 'complex':
                        return complex(self.parse(0, line, f, sp, args)[2])
                    # gets the type of the argument
                    elif func == 'type':
                        return type(self.parse(0, line, f, sp, args)[2])
                    # gets the dir of the argument
                    elif func == 'dir':
                        return dir(self.parse(0, line, f, sp, args)[2])
                    # casting to iterables / sets / dicts
                    elif func == 'set':
                        # creates a set from all arguments
                        if args[0][0] == '':
                            return set()
                        # creates a set from all arguments
                        s = set()
                        # adds all arguments to the set
                        for i in range(len(args)):
                            s.add(self.parse(i, line, f, sp, args)[2])
                        return s
                    elif func == 'dict':
                        return dict(self.parse(0, line, f, sp, args)[2])
                    elif func == 'tuple':
                        return tuple(self.parse(0, line, f, sp, args)[2])
                except Exception as e:
                    self.err(
                        'Casting error',
                        f'Could not cast arg to specified type\n{e}',
                        line, lines_ran
                    )
                # conditional logic
                if func == 'if':
                    # false block is optional
                    try:
                        false_block_s = args[2][0]
                    except:
                        false_block_s = None
                    ifcond = self.parse(0, line, f, sp, args)[2]
                    # if condition is true
                    if ifcond:
                        return self.parse(1, line, f, sp, args)[2]
                    # otherwise false block is executed
                    if false_block_s:
                        return self.parse(2, line, f, sp, args)[2]
                    return False
                # while logic WIPWIPWIPWIPWIP
                elif func == 'while':
                    while (self.interpret(args[0][0])):
                        self.interpret(args[1][0])
                    return True
                 # iteration
                elif func == 'for':
                    # times to loop
                    start = self.parse(0, line, f, sp, args)[2]
                    end = self.parse(1, line, f, sp, args)[2]
                    loopvar = self.parse(2, line, f, sp, args)[2]
                    # start must be int
                    # end must be int
                    # loopvar must be str
                    self.type_err([(start, (int,)), (end, (int,)), (loopvar, (str,))], line, lines_ran)
                    self.vars[loopvar] = Var(loopvar, start)
                    # regular iteration
                    if start < end:
                        for i in range(start, end):
                            if loopvar in self.vars and self.vars[loopvar].value >= end:
                                break
                            self.vars[loopvar] = Var(loopvar, i)
                            self.interpret(args[3][0])
                    # reversed if requested
                    elif start > end:
                        for i in reversed(range(end, start)):
                            if loopvar in self.vars and self.vars[loopvar].value < end:
                                break
                            self.vars[loopvar] = Var(loopvar, i)
                            self.interpret(args[3][0])
                    return self.vars[loopvar].value
                # executes a block of code for each element in an array
                elif func == 'each':
                    # get array argument
                    array = self.parse(0, line, f, sp, args)[2]
                    # get element variable name
                    element_name = self.parse(1, line, f, sp, args)[2]
                    # prepare each element
                    self.vars[element_name] = Var(element_name, 0)
                    # execute block for each element
                    for i in range(len(array)):
                        self.vars[element_name].value = array[i]
                        self.interpret(args[2][0])
                    return array
                # sorting an array by an attribute of each element
                elif func == 'sortby':
                    # iterable to sort
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be an iterable
                    self.check_iterable(iterable, line)
                    # variable name
                    varname = self.parse(1, line, f, sp, args)[2]
                    # check variable name
                    self.check_varname(varname, line)
                    # pairing elements to their interpretations
                    pairing = []
                    for i in range(len(iterable)):
                        self.vars[varname] = Var(varname, iterable[i])
                        pairing.append(
                            (self.interpret(args[2][0]), iterable[i]))
                    # sort the pairing based on the first element of each pair
                    pairing.sort(key=lambda x: x[0])
                    # return the sorted array containing the second element
                    # of each pair
                    return [pair[1] for pair in pairing]
                # performs list comprehension
                elif func == 'comp':
                    lst = []
                    # array to comprehend
                    arr = self.parse(0, line, f, sp, args)[2]
                    # should be an iterable
                    self.check_iterable(arr, line)
                    # varname for the element
                    varname = self.parse(1, line, f, sp, args)[2]
                    # should be a varname
                    self.check_varname(varname, line)
                    # block to perform
                    block = args[2][0]
                    # performs the list comprehension
                    for v in arr:
                        self.vars[varname] = Var(varname, v)
                        r = self.interpret(block)
                        if r != msn2_none:
                            lst.append(r)
                    return lst
                # returns the first argument, then performs the second argument as a block
                elif func == 'do':
                    ret = self.parse(0, line, f, sp, args)[2]
                    self.interpret(args[1][0])
                    return ret
                # special value for msn2 to return None
                elif func == 'None':
                    return msn2_none
                # filters an iterable to retain all elements that satisfy the second argument as
                # a block
                # the first argument is a string for a variable for each element
                # second argument is a block to run for each element
                elif func == 'filter':
                    # iterable to filter
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # check if iterable
                    self.check_iterable(iterable, line)
                    # variable name
                    varname = self.parse(1, line, f, sp, args)[2]
                    # check variable name
                    self.check_varname(varname, line)
                    # new array
                    filtered = []
                    # iterate through each element
                    for v in iterable:
                        # set the variable to the element
                        self.vars[varname] = Var(varname, v)
                        # if the block returns true, add the element to the new array
                        if self.interpret(args[2][0]):
                            filtered.append(v)
                    return filtered
                # unpacks the first argument into any amount of variables
                # specified by the remaining arguments provided as variable names
                # creates the variables if they don't exists
                # 5/22/2023
                elif func == 'unpack':
                    # iterable to unpack
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # check if iterable
                    self.check_iterable(iterable, line)
                    # variable names to unpack into
                    for i in range(1, len(args)):
                        varname = self.parse(i, line, f, sp, args)[2]
                        self.vars[varname] = Var(varname, iterable[i - 1])
                    return iterable
                # determines if the array has all elements specified
                # takes any amount of arguments
                # first argument is the iterable
                # the rest are elements to check for
                elif func == 'has':
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    # optimized code:
                    try:
                        return all(self.parse(i + 1, line, f, sp, args)[2] in iterable for i in range(len(args) - 1))
                    except Exception as e:
                        return self.err(
                            'Error in has()',
                            e, line, lines_ran
                        )
                # gets the first element in the iterable
                elif func == 'first' or func == 'head':
                    try:
                        return self.parse(0, line, f, sp, args)[2][0]
                    except:
                        return None
                # gets the last element in the iterable
                elif func == 'last' or func == 'tail':
                    try:
                        return self.parse(0, line, f, sp, args)[2][-1]
                    except:
                        return None
                # the following provide efficient variable arithmetic
                elif func == 'add':
                    first = self.parse(0, line, f, sp, args)[2]
                    second = self.parse(1, line, f, sp, args)[2]
                    # first must be a varname
                    self.check_varname(first, line)
                    # case array
                    if isinstance(self.vars[first].value, list):
                        self.vars[first].value.append(second)
                    # case string or number
                    else:
                        self.vars[first].value += second
                    return self.vars[first].value
                # performs basic operations on non-variable values
                elif obj == 'op':
                    try:
                        # obtains the first argument
                        arg1 = self.parse(0, line, f, sp, args)[2]
                        # adds all arguments
                        if objfunc == 'append' or objfunc == 'push' or objfunc == 'add' or objfunc == 'plus' or objfunc == '+' or objfunc == 'concat' or objfunc == 'concatenate' or objfunc == 'join' or objfunc == 'merge' or objfunc == 'sum':
                            if isinstance(arg1, list):
                                for i in range(1, len(args)):
                                    arg1.append(self.parse(
                                        i, line, f, sp, args)[2])
                                return arg1
                            else:
                                for i in range(1, len(args)):
                                    arg1 += self.parse(i, line, f, sp, args)[2]
                                return arg1
                        # subtracts all arguments
                        if objfunc == 'sub' or objfunc == 'minus' or objfunc == 'subtract' or objfunc == '-':
                            for i in range(1, len(args)):
                                arg1 -= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        if objfunc == 'mul' or objfunc == 'times' or objfunc == 'x' or objfunc == '*' or objfunc == 'multiply':
                            for i in range(1, len(args)):
                                arg1 *= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        if objfunc == 'div' or objfunc == 'divide' or objfunc == 'over' or objfunc == '/':
                            for i in range(1, len(args)):
                                arg1 /= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        # integer division
                        if objfunc == 'idiv' or objfunc == 'intdiv' or objfunc == 'intdivide' or objfunc == 'intover' or objfunc == '//' or objfunc == '÷÷':
                            for i in range(1, len(args)):
                                arg1 //= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        if objfunc == 'mod' or objfunc == 'modulo' or objfunc == 'modulus' or objfunc == '%' or objfunc == 'remainder':
                            for i in range(1, len(args)):
                                arg1 %= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        if objfunc == 'pow' or objfunc == 'power' or objfunc == 'exponent' or objfunc == '**':
                            for i in range(1, len(args)):
                                arg1 **= self.parse(i, line, f, sp, args)[2]
                            return arg1
                        if objfunc == 'root' or objfunc == 'nthroot' or objfunc == 'nthrt':
                            for i in range(1, len(args)):
                                arg1 **= (1 / self.parse(i, line, f, sp, args)[2])
                            return arg1
                        return '<msnint2 class>'
                    except Exception as e:
                        self.err(
                            'Error in op() class',
                            f'Could not perform operation on arguments\n{e}',
                            line, lines_ran
                        )
                # formats a number to a money string
                elif func == 'USD':
                    # number to format
                    num = self.parse(0, line, f, sp, args)[2]
                    # number must be int or float
                    self.type_err([(num, (int, float))], line, lines_ran)
                    return f"${num:,.2f}"
                # formats a number to a certain number of decimal places
                elif func == 'format':
                    # number to format
                    num = self.parse(0, line, f, sp, args)[2]
                    # number must be int or float
                    self.type_err([(num, (int, float))], line, lines_ran)
                    # number of decimal places
                    places = self.parse(1, line, f, sp, args)[2]
                    # places must be int
                    self.type_err([(places, (int,))], line, lines_ran)
                    return f"{num:.{places}f}"
                # rounds a number
                elif func == 'round':
                    # number to round
                    num = self.parse(0, line, f, sp, args)[2]
                    # number must be int or float
                    self.type_err([(num, (int, float))], line, lines_ran)
                    # number of digits to round to
                    digits = self.parse(1, line, f, sp, args)[2]
                    # digits must be int
                    self.type_err([(digits, (int,))], line, lines_ran)
                    return round(num, digits)
                # computes the maximum value from all arguments
                # takes any amount of arguments, all being
                # either numbers or lists
                elif func == 'maximum':
                    try:
                        maxval = max(_f) if isinstance(
                            (_f := self.parse(0, line, f, sp, args)[2] ), list) else _f
                        for i in range(1, len(args)):
                            val = self.parse(i, line, f, sp, args)[2]
                            # is a list argument
                            if isinstance(val, list):
                                maxval = max(maxval, max(val))
                            # is a number
                            else:
                                maxval = max(maxval, val)
                    except Exception as e:
                        self.err(
                            'Error finding maximum value',
                            e,
                            line, lines_ran
                        )
                    return maxval
                elif func == 'minimum':
                    try:
                        minval = min(_f) if isinstance(
                            (_f := self.parse(0, line, f, sp, args)[2]), list) else _f
                        for i in range(1, len(args)):
                            val = self.parse(i, line, f, sp, args)[2]
                            # is a list argument
                            if isinstance(val, list):
                                minval = min(minval, min(val))
                            # is a number
                            else:
                                minval = min(minval, val)
                    except Exception as e:
                        self.err(
                            'Error finding minimum value',
                            e,
                            line, lines_ran
                        )
                    return minval

                # more support for functions
                elif obj == 'function':
                    fname = self.parse(0, line, f, sp, args)[2]
                    # fname must be a string
                    self.type_err([(fname, (str,))], line, lines_ran)
                    # adds a line of code to a function / method's body
                    if objfunc == 'addbody':
                        self.methods[fname].add_body(
                            self.parse(1, line, f, sp, args)[2])
                        return fname
                    # adds an argument to a function
                    if objfunc == 'addarg':
                        arg = self.parse(1, line, f, sp, args)[2]
                        # arg must be a string
                        self.type_err([(arg, (str,))], line, lines_ran)
                        self.methods[fname].add_arg(
                            arg)
                        return fname
                    # adds a return variable to this function
                    if objfunc == 'addreturn':
                        ret = self.parse(1, line, f, sp, args)[2]
                        # ret must be a string
                        self.type_err([(ret, (str,))], line, lines_ran)
                        self.methods[fname].add_return(
                            ret)
                        return fname
                    if objfunc == 'getbody':
                        return self.methods[fname].body
                    if objfunc == 'getargs':
                        return self.methods[fname].args
                    if objfunc == 'getreturn':
                        return self.methods[fname].returns[0]
                    # removes a function from the working context
                    if objfunc == 'destroy:function':
                        self.methods.pop(fname)
                        return fname
                    # simulates an execution of a function
                    if objfunc == 'run':
                        # form a string that is msn2 of the user defined function
                        args_str = ''
                        for i in range(1, len(args)):
                            arg = self.parse(i, line, f, sp, args)[2]
                            if i != len(args) - 1:
                                args_str += f"{arg},"
                            else:
                                args_str += str(arg)
                        inst = f"{fname}({args_str})"
                        return self.interpret(inst)
                    return '<msnint2 class>'
                elif func == 'sub':
                    vn = self.parse(0, line, f, sp, args)[2]
                    # vn must be a variable name
                    self.check_varname(vn, line)
                    # gets the substring
                    other = self.parse(1, line, f, sp, args)[2]
                    self.vars[vn].value -= other
                    return self.vars[first].value
                elif func == 'mul':
                    vn = self.parse(0, line, f, sp, args)[2]
                    # vn must be a variable name
                    self.check_varname(vn, line)
                    # gets the substring
                    other = self.parse(1, line, f, sp, args)[2]
                    self.vars[vn].value *= other
                    return self.vars[first].value
                elif func == 'div':
                    vn = self.parse(0, line, f, sp, args)[2]
                    # vn must be a variable name
                    self.check_varname(vn, line)
                    # gets the substring
                    other = self.parse(1, line, f, sp, args)[2]
                    self.vars[vn].value /= other
                    return self.vars[first].value
                # appends to an array variable
                elif func == 'append':
                    # variable name
                    vn = self.parse(0, line, f, sp, args)[2]
                    # vn must be a variable name
                    self.check_varname(vn, line)
                    other = self.parse(1, line, f, sp, args)[2]
                    # appends to the array
                    self.vars[vn].value.append(other)
                    return self.vars[vn].value
                # gets at the index specified
                elif func == '->':
                    indexable = self.parse(0, line, f, sp, args)[2]
                    # must be indexable
                    if not isinstance(indexable, (list, str, dict, tuple)):
                        self.err(
                            'Error indexing with ->()',
                            'First argument not indexable, indexable types are lists, strings, dicts, and tuples',
                            line, lines_ran
                        )
                    try:    
                        return indexable[self.parse(1, line, f, sp, args)[2]]
                    except IndexError:
                        return self.raise_index_out_of_bounds(line, lines_ran, self.Method('->', self))
                # gets the MSNScript version of this interpreter
                elif func == 'version':
                    return self.version
                # destroys a function or variable
                elif func == 'destroy':
                    for i in range(len(args)):
                        varname = self.parse(i, line, f, sp, args)[2]
                        # varname must be varname
                        self.check_varname(varname, line)
                        # must be a varname
                        # deletes all variables or methods that start with '__'
                        if varname == '__':
                            for key in list(self.vars):
                                if key.startswith('__'):
                                    del self.vars[key]
                            return True
                        if varname in self.vars:
                            del self.vars[varname]
                        elif varname in self.methods:
                            del self.methods[varname]
                    return True
                # destroys functions
                elif func == 'destroy:function':
                    fname = None
                    for i in range(len(args)):
                        fname = self.parse(i, line, f, sp, args)[2]
                        # must be a varname
                        self.check_varname(fname, line)
                        self.methods.pop(fname)
                    return fname

                # gets a range()
                elif func == 'range':
                    start = self.parse(0, line, f, sp, args)[2]
                    # start must be int
                    self.type_err([(start, (int,))], line, lines_ran)
                    # if one argument
                    if len(args) == 1:
                        return range(start)
                    # get the end of the range
                    end = self.parse(1, line, f, sp, args)[2]
                    # end must be int
                    self.type_err([(end, (int,))], line, lines_ran)
                    # if two arguments
                    if len(args) == 2:
                        return range(start, end)
                    if len(args) == 3:
                        step = self.parse(2, line, f, sp, args)[2]
                        # step must be int
                        self.type_err([(step, (int,))], line, lines_ran)
                        return range(start, end, step)
                    return range()

                # random capabilities
                elif func == 'random':
                    import random
                    # gets a random number between 0 and 1
                    if len(args) == 1:
                        return random.random()
                    # random number in range
                    elif len(args) == 2:
                        arg = self.parse(0, line, f, sp, args)[2]
                        # arg must be int
                        self.type_err([(arg, (int,))], line, lines_ran)
                        arg2 = self.parse(1, line, f, sp, args)[2]
                        # arg2 must be int
                        self.type_err([(arg2, (int,))], line, lines_ran)
                        return (random.random() * (arg2 - arg)) + arg
                    # random int in range
                    elif len(args) == 3:
                        # using math
                        import math
                        arg = self.parse(0, line, f, sp, args)[2]
                        # arg must be int
                        self.type_err([(arg, (int,))], line, lines_ran)
                        arg2 = self.parse(1, line, f, sp, args)[2]
                        # arg must be int
                        self.type_err([(arg2, (int,))], line, lines_ran)
                        return math.floor((random.random() * (arg2 - arg)) + arg)
                    return '<msnint2 class>'

                # html parsing simplified
                elif obj == 'html':
                    url = self.parse(0, line, f, sp, args)[2]
                    # url must be str
                    self.type_err([(url, (str,))], line, lines_ran)
                    # creates a BeautifulSoup object of a url
                    if objfunc == 'soup':
                        from bs4 import BeautifulSoup
                        import requests
                        response = requests.get(url)
                        return BeautifulSoup(response.content, 'html5lib')
                    # scrapes all html elements from a url
                    # WIPWIPWIPWIP
                    if objfunc == 'from':
                        obj_to_add = []
                        all_elem = self.html_all_elements(url)
                        for elem in all_elem:
                            obj_to_add.append(
                                {'tag': elem.name, 'attrs': elem.attrs, 'text': elem.text})
                        return obj_to_add
                    # starts an HTMLSession
                    if objfunc == 'session':
                        from requests_html import HTMLSession
                        return HTMLSession()

                # ai specific usage
                elif obj == 'ai':
                    # ChatGPT API
                    import openai
                    import tiktoken
                    global models
                    # verify existence of openai api key
                    if not openai.api_key:
                        try:
                            openai.api_key = os.environ['OPENAI_API_KEY']
                        except:
                            # msn2 error, no OPENAI env var
                            self.err(
                                f"OpenAI API key not found. Please set your OPENAI_API_KEY environment variable to your OpenAI API key.", True, '', lines_ran)
                    # if models not defined, define them
                    if not models:
                        # determines if a model needs chatcompletion
                        def needs_completion(model):
                            return model.startswith('text') or model == 'gpt-3.5-turbo-instruct'
                        # determines if arguments for 'model' are str and in models
                        def check_model(model):
                            # model must be str
                            self.type_err([(model, (str,))], line, lines_ran)
                            # model must exist
                            if model not in models:
                                self.err(
                                    'Model not found',
                                    f'Model {model} not found. Available models are {", ".join(models.keys())}',
                                    line, lines_ran
                                )
                        # gets responses from the models
                        def response(model, prompt):
                            # enforce model as str and prompt as str
                            self.type_err([(model, (str,)), (prompt, (str,))], line, lines_ran)
                            if needs_completion(model):
                                # return v1/completions endpoint
                                return openai.Completion.create(
                                    model=model,
                                    prompt=prompt,
                                    temperature=0.5,
                                    max_tokens=models[model]['max_tokens'] // 2
                                ).choices[0].text
                            else:
                                return openai.ChatCompletion.create(
                                    model=model,
                                    messages=[{'role': 'user', 'content': prompt}],
                                    temperature=0.5,
                                    max_tokens=models[model]['max_tokens'] // 2
                                )
                        # available models for use by the query
                        models = {
                            'gpt-3.5-turbo-instruct': {
                                "max_tokens": 4097,
                                # compute $0.0015 / 1K tokens
                                "price_per_token": 0.0015 / 1000
                            },
                            'gpt-3.5-turbo': {
                                "max_tokens": 4097,
                                "price_per_token": 0.0015 / 1000
                            },
                            'gpt-3.5-turbo-16k': {
                                "max_tokens": 16384,
                                "price_per_token": 0.003 / 1000
                            },
                            "text-davinci-003": {
                                "max_tokens": 4097,
                                # turbo * 10, as stated by OPENAI model pricing documentation
                                "price_per_token": (0.0015 / 1000) * 10
                            },
                            # gets responses from these models
                            'response': response,
                            'needs_completion': needs_completion,
                            'check_model': check_model
                        }
                    # gets the available models
                    if objfunc == 'models':
                        return models
                    # gets max tokens for a model
                    elif objfunc == 'max_tokens':
                        # get model
                        model = self.parse(0, line, f, sp, args)[2]
                        # check model
                        models['check_model'](model)
                        # return max tokens
                        return models[model]['max_tokens']
                    # gets price per token for a model
                    elif objfunc == 'price_per_token':
                        # get the model
                        model = self.parse(0, line, f, sp, args)[2]
                        # check the model
                        models['check_model'](model)
                        # return price per token
                        return models[model]['price_per_token']
                    # asks openai model a question
                    # simple ai, see top of file for definition
                    # simple ask of the AI without context
                    if objfunc == 'basic':
                        # generates an ai response with the basic model
                        return models['response']('gpt-3.5-turbo', self.parse(0, line, f, sp, args)[2])
                    # asks an advanced model a question, no context
                    elif objfunc == 'advanced':
                        # generates an ai response with the advanced model
                        return models['response']('gpt-3.5-turbo-16k', self.parse(0, line, f, sp, args)[2])
                    # 2.0.388
                    # makes a fully customized query to the openai api
                    elif objfunc == 'query':
                        # model to use
                        model = self.parse(0, line, f, sp, args)[2]
                        # check model
                        models['check_model'](model)
                        # messages
                        messages = self.parse(1, line, f, sp, args)[2]
                        # messages must be list
                        self.type_err([(messages, (list,))], line, lines_ran)
                        # temperature
                        temperature = self.parse(2, line, f, sp, args)[2]
                        # temp must be int or float
                        self.type_err([(temperature, (int, float))], line, lines_ran)
                        # max_tokens
                        max_tokens = self.parse(3, line, f, sp, args)[2]
                        # max_tokens must be int
                        self.type_err([(max_tokens, (int,))], line, lines_ran)
                        # top_p
                        top_p = self.parse(4, line, f, sp, args)[2]
                        # top_p must be int or float
                        self.type_err([(top_p, (int, float))], line, lines_ran)
                        # frequency_penalty
                        frequency_penalty = self.parse(5, line, f, sp, args)[2]
                        # frequency_penalty must be int or float
                        self.type_err([(frequency_penalty, (int, float))], line, lines_ran)
                        # presence_penalty
                        presence_penalty = self.parse(6, line, f, sp, args)[2]
                        # presence_penalty must be int or float
                        self.type_err([(presence_penalty, (int, float))], line, lines_ran)
                        # stop
                        stop = self.parse(7, line, f, sp, args)[2]
                        # stop must be str
                        self.type_err([(stop, (str,))], line, lines_ran)
                        # if model is standard
                        if models['needs_completion'](model):
                            return openai.Completion.create(
                                model=model,
                                prompt=messages,
                                temperature=temperature,
                                max_tokens=max_tokens,
                                top_p=top_p,
                                frequency_penalty=frequency_penalty,
                                presence_penalty=presence_penalty,
                                stop=stop
                            ).choices[0].text
                        else:
                            return openai.ChatCompletion.create(
                                model=model,
                                messages=messages,
                                temperature=temperature,
                                top_p=top_p,
                                frequency_penalty=frequency_penalty,
                                presence_penalty=presence_penalty,
                                stop=stop
                            )
                            
                    # gets the amount of tokens for a string given a ChatGPT model
                    elif objfunc == 'tokens':
                        # string to check
                        prompt = self.parse(0, line, f, sp, args)[2]
                        # prompt must be str
                        self.type_err([(prompt, (str,))], line, lines_ran)
                        # model to check
                        model_name = self.parse(1, line, f, sp, args)[2]
                        # check model
                        models['check_model'](model_name)
                        # get the encoding
                        return len(tiktoken.encoding_for_model(model_name).encode(prompt))
                    return '<msnint2 class>'
                # merges all arguments into one
                elif func == 'merge':
                    # gets the first argument
                    arg1 = self.parse(0, line, f, sp, args)[2]
                    # arg must exist
                    if arg1 is None:
                        self.err(
                            'Error in merge()',
                            'First argument must exist',
                            line, lines_ran
                        )
                    # gets the rest of the arguments
                    for i in range(1, len(args)):
                        _arg = self.parse(i, line, f, sp, args)[2]
                        # arg must exist
                        if _arg is None:
                            self.err(
                                'Error in merge()',
                                'Merging arguments must exist',
                                line, lines_ran
                            )
                        arg1 |= self.parse(i, line, f, sp, args)[2]
                    return arg1
                # raises an msn2 exception
                elif func == 'exception':
                    # get the error
                    error = self.parse(0, line, f, sp, args)[2]
                    # error must be str
                    self.type_err([(error, (str,))], line, lines_ran)
                    # get the message
                    message = self.parse(1, line, f, sp, args)[2]
                    # message must be str
                    self.type_err([(message, (str,))], line, lines_ran)
                    # get the line
                    l = self.parse(2, line, f, sp, args)[2]
                    # line must be str
                    self.type_err([(l, (str,))], line, lines_ran)
                    # set trying to False
                    return self.err(error, message, l, lines_ran)

                # defines new syntax, see tests/validator.msn2 for documentation
                elif func == 'syntax':
                    synt = self.parse(0, line, f, sp, args)[2]
                    # synt must be str
                    self.type_err([(synt, (str,))], line, lines_ran)
                    # add the between
                    between = self.parse(1, line, f, sp, args)[2]
                    # between must be a varname
                    self.check_varname(between, line)
                    return self.add_syntax(synt, between, args[2][0])
                # creates a new enclosed syntax that should execute the block
                # specified on the line by which it was created
                elif func == 'enclosedsyntax':
                    start = self.parse(0, line, f, sp, args)[2]
                    # start must be str
                    self.type_err([(start, (str,))], line, lines_ran)
                    end = self.parse(1, line, f, sp, args)[2]
                    # end must be str
                    self.type_err([(end, (str,))], line, lines_ran)
                    varname = self.parse(2, line, f, sp, args)[2]
                    # varname must be a varname
                    self.check_varname(varname, line)
                    index = f"{start}msnint2_reserved{end}"
                    enclosed[index] = [start, end, varname, args[3][0]]
                    if len(args) == 5:
                        enclosed[index].append(
                            self.parse(4, line, f, sp, args)[2])
                    return enclosed[index]
                # defines a new macro
                elif func == 'macro':
                    token = self.parse(0, line, f, sp, args)[2]
                    # token must be str
                    self.type_err([(token, (str,))], line, lines_ran)
                    # get the symbol
                    symbol = self.parse(1, line, f, sp, args)[2]
                    # symbol must be str
                    self.type_err([(symbol, (str,))], line, lines_ran)
                    macros[token] = [token, symbol, args[2][0]]
                    # 4th argument offered as a return value from that macro
                    # as opposed to a block of code
                    if len(args) == 4:
                        macros[token].append(
                            self.parse(3, line, f, sp, args)[2])
                    return macros[token]
                # creates a macro that should be declared at the end of a line
                elif func == 'postmacro':
                    token = self.parse(0, line, f, sp, args)[2]
                    # token must be str
                    self.type_err([(token, (str,))], line, lines_ran)
                    # get the symbol
                    symbol = self.parse(1, line, f, sp, args)[2]
                    # symbol must be str
                    self.type_err([(symbol, (str,))], line, lines_ran)
                    postmacros[token] = [token, symbol, args[2][0]]
                    # same as macro
                    if len(args) == 4:
                        postmacros[token].append(
                            self.parse(3, line, f, sp, args)[2])
                    return postmacros[token]
                # creates an inline syntax that allows for value replacement
                # within a line
                # this is different from the above system calls because this
                # syntax is invoked and returned as a value which will replace
                # the invocation within the line
                # arguments are the same as enclosedsyntax
                #
                # WIPWIPWIPWIPWIP
                # elif func == 'inlinesyntax':
                #     start = self.parse(0, line, f, sp, args)[2]
                #     end = self.parse(1, line, f, sp, args)[2]
                #     index = f"{start}msnint2_reserved{end}"
                #     inlines[index] = [start, end, self.parse(
                #         2, line, f, sp, args)[2], args[3][0]]
                #     if len(args) == 5:
                #         inlines[index].append(
                #             self.parse(4, line, f, sp, args)[2])
                #     return inlines[index]
                # performs object based operations
                elif obj == 'var':
                    # determines if all variables passed are equal
                    if objfunc == 'equals':
                        firstvar = self.vars[self.parse(
                            0, line, f, sp, args)[2]].value
                        return all(firstvar == self.vars[self.parse(i, line, f, sp, args)[2]].value for i in range(1, len(args)))
                    return '<msnint2 class>'
                # gets the value of a variable
                elif func == 'val':
                    # gets the variable name
                    varname = self.parse(0, line, f, sp, args)[2]
                    # varname must be a varname
                    self.check_varname(varname, line)
                    try:
                        return self.vars[varname].value
                    except:
                        return self.vars[varname]
                # gets a sorted version of the array
                elif func == 'sorted':
                    # iterable
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    return sorted(iterable)
                # gets a copy of the object
                elif func == 'copy':
                    try:
                        return self.parse(0, line, f, sp, args)[2].copy()
                    except Exception as e:
                        self.err(
                            'Error copying object',
                            e, line, lines_ran
                        )
                # performs file-specific operations
                elif obj == 'file':
                    # import shutil
                    import shutil
                    # creates a file
                    if objfunc == 'create':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        open(path, 'w').close()
                        lock.release()
                        return True
                    # reads text from a file
                    if objfunc == 'read':
                        lock.acquire()
                        # get filename
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        file = open(path, "r", encoding='utf-8')
                        contents = file.read()
                        file.close()
                        lock.release()
                        return contents
                    # writes to a file
                    if objfunc == 'write':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        file = open(path, "w")
                        towrite = str(self.parse(1, line, f, sp, args)[2])
                        file.write(towrite)
                        file.close()
                        lock.release()
                        return towrite
                    # writes the argument as code
                    if objfunc == 'writemsn':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        file = open(path, "w")
                        towrite = args[1][0]
                        file.write(towrite)
                        lock.release()
                        return towrite
                    # clears a file of all text
                    if objfunc == 'clear':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        file = open(path, "w")
                        file.write("")
                        file.close()
                        lock.release()
                        return True
                    # appends to a file
                    if objfunc == 'append':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        file = open(path, "a")
                        towrite = str(self.parse(1, line, f, sp, args)[2])
                        file.write(towrite)
                        file.close()
                        lock.release()
                        return towrite
                    # deletes a file
                    if objfunc == 'delete':
                        lock.acquire()
                        deleting = self.parse(0, line, f, sp, args)[2]
                        # deleting must be str
                        self.type_err([(deleting, (str,))], line, lines_ran)
                        try:
                            os.remove(deleting)
                        except:
                            None
                        lock.release()
                        return deleting
                    # renames a file
                    if objfunc == 'rename':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        # old must be str
                        self.type_err([(old, (str,))], line, lines_ran)
                        new = self.parse(1, line, f, sp, args)[2]
                        # new must be str
                        self.type_err([(new, (str,))], line, lines_ran)
                        os.rename(old, new)
                        lock.release()
                        return new
                    # copies a file
                    if objfunc == 'copy':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        # old must be str
                        self.type_err([(old, (str,))], line, lines_ran)
                        new = self.parse(1, line, f, sp, args)[2]
                        # new must be str
                        self.type_err([(new, (str,))], line, lines_ran)
                        shutil.copy2(old, new)
                        lock.release()
                        return new
                    if objfunc == 'copy2':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        # old must be str
                        self.type_err([(old, (str,))], line, lines_ran)
                        new = self.parse(1, line, f, sp, args)[2]
                        # new must be str
                        self.type_err([(new, (str,))], line, lines_ran)
                        shutil.copy2(old, new)
                        lock.release()
                        return new
                    if objfunc == 'copyfile':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        # old must be str
                        self.type_err([(old, (str,))], line, lines_ran)
                        new = self.parse(1, line, f, sp, args)[2]
                        # new must be str
                        self.type_err([(new, (str,))], line, lines_ran)
                        shutil.copyfile(old, new)
                        lock.release()
                        return new
                    if objfunc == 'fullpath':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        fullpath = os.path.abspath(path)
                        lock.release()
                        return fullpath
                    # moves a file
                    if objfunc == 'move':
                        lock.acquire()
                        old = self.parse(0, line, f, sp, args)[2]
                        # old must be str
                        self.type_err([(old, (str,))], line, lines_ran)
                        new = self.parse(1, line, f, sp, args)[2]
                        # new must be str
                        self.type_err([(new, (str,))], line, lines_ran)
                        shutil.move(old, new)
                        lock.release()
                        return new
                    # determines if a file exists
                    if objfunc == 'exists':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        exists = os.path.exists(path)
                        lock.release()
                        return exists
                    # determines if a file is a directory
                    if objfunc == 'isdir':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        isdir = os.path.isdir(path)
                        lock.release()
                        return isdir
                    # determines if a file is a file
                    if objfunc == 'isfile':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        isfile = os.path.isfile(path)
                        lock.release()
                        return isfile
                    # lists files in a directory
                    if objfunc == 'listdir':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        try:
                            listdir = os.listdir(path)
                            lock.release()
                            return listdir
                        except FileNotFoundError:
                            # directory doesn't exist
                            lock.release()
                            return None
                    # makes a directory
                    if objfunc == 'mkdir':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        try:
                            made = os.mkdir(path)
                            lock.release()
                            return made
                        except FileExistsError:
                            lock.release()
                            return False
                    # removes a directory
                    if objfunc == 'rmdir':
                        lock.acquire()
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        try:
                            rm = os.rmdir(path)
                            lock.release()
                            return rm
                        except OSError:
                            lock.release()
                            return None
                    # gets the current working directory
                    if objfunc == 'getcwd':
                        lock.acquire()
                        cwd = os.getcwd()
                        lock.release()
                        return cwd
                    # gets the size of a file
                    if objfunc == 'getsize':
                        lock.acquire()
                        # get filename
                        path = self.parse(0, line, f, sp, args)[2]
                        # path must be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        size = os.path.getsize(path)
                        lock.release()
                        return size
                    # deletes all files and directories within a directory
                    if objfunc == 'emptydir':
                        lock.acquire()
                        directory = self.parse(0, line, f, sp, args)[2]
                        # directory must be str
                        self.type_err([(directory, (str,))], line, lines_ran)
                        try:
                            for file in os.listdir(directory):
                                try:
                                    os.remove(os.path.join(directory, file))
                                except:
                                    shutil.rmtree(os.path.join(
                                        directory, file), ignore_errors=True)
                            lock.release()
                            return directory
                        except FileNotFoundError:
                            # directory doesn't exist
                            lock.release()
                            return None
                elif func == 'fileacquire':
                    lock.acquire()
                    return True
                elif func == 'filerelease':
                    lock.release()
                    return True
                # automation operations
                elif obj == 'auto':
                    # gets the largest element from a list of elements
                    if objfunc == 'largest':
                        elements = self.parse(0, line, f, sp, args)[2]
                        # elements must be iterable
                        self.check_iterable(elements, line)
                        if not elements:
                            return elements
                        largest = elements[0]
                        for element in elements:
                            try:
                                # element has width and height
                                if element.width() > largest.width() and element.height() > largest.height():
                                    largest = element
                            except:
                                # element does not have width and height
                                return element
                        return largest
                    # picks a file from the file chooser
                    # returns the path of the file chosen
                    if objfunc == 'file':
                        # imports
                        from tkinter import Tk
                        from tkinter.filedialog import askopenfilename
                        Tk().withdraw()
                        # you can only run .msn2 scripts
                        return askopenfilename(initialdir=os.getcwd(), filetypes=[("MSN2 Script", "*.msn2")])
                    return '<msnint2 class>'
                # # performs math operations
                try:
                    if obj == 'math':
                        if objfunc == 'add':
                            return self.parse(0, line, f, sp, args)[2] + self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'subtract':
                            return self.parse(0, line, f, sp, args)[2] - self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'multiply':
                            return self.parse(0, line, f, sp, args)[2] * self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'divide':
                            return self.parse(0, line, f, sp, args)[2] / self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'power':
                            return self.parse(0, line, f, sp, args)[2] ** self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'root':
                            return self.parse(0, line, f, sp, args)[2] ** (1 / self.parse(1, line, f, sp, args)[2])
                        if objfunc == 'sqrt':
                            return self.parse(0, line, f, sp, args)[2] ** 0.5
                        if objfunc == 'mod':
                            return self.parse(0, line, f, sp, args)[2] % self.parse(1, line, f, sp, args)[2]
                        if objfunc == 'floor':
                            import math
                            return math.floor(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'ceil':
                            import math
                            return math.ceil(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'round':
                            return round(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'abs':
                            return abs(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'sin':
                            import math
                            return math.sin(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'cos':
                            import math
                            return math.cos(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'tan':
                            import math
                            return math.tan(self.parse(0, line, f, sp, args)[2])
                        if objfunc == 'asin':
                            import math
                            return math.asin(self.parse(0, line, f, sp, args)[2])
                        return '<msnint2 class>'
                except Exception as e:
                    self.err(
                        'Error in math class',
                        f'Unable to perform computation\n{e}',
                        line, lines_ran
                    )

                # inserts a value into the iterable at the specified index
                if func == 'map':
                    # iterable
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    # varname
                    varname = self.parse(1, line, f, sp, args)[2]
                    # varname must be a varname
                    self.check_varname(varname, line)
                    # map the function to each element in the iterable
                    for i, el in enumerate(iterable):
                        self.vars[varname] = Var(varname, el)
                        iterable[i] = self.interpret(args[2][0])
                    return iterable
                # inserts a value into the iterable at the specified index
                elif func == 'insert':
                    # iterable
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    # index
                    index = self.parse(1, line, f, sp, args)[2]
                    # index must be int
                    self.type_err([(index, (int,))], line, lines_ran)
                    # value
                    value = self.parse(2, line, f, sp, args)[2]
                    # insert the value into the iterable
                    return iterable.insert(index, value)
                # gets the type of the first argument passed
                elif func == 'type':
                    return type(self.parse(0, line, f, sp, args)[2])
                # gets the parent context
                elif func == 'parent':
                    return self.parent
                # gets the booting context
                elif func == 'boot':
                    while self.parent != None:
                        self = self.parent
                    return self
                # deletes a variable
                elif func == 'del':
                    for i in range(len(args)):
                        first = self.parse(i, line, f, sp, args)[2]
                        # first must be a varname
                        self.check_varname(first, line)
                        del self.vars[first]
                    return True
                # concatinates two strings
                elif func == 'cat':
                    cat = str(self.parse(0, line, f, sp, args)[2])
                    # concatinate rest of arguments
                    for i in range(1, len(args)):
                        cat += str(self.parse(i, line, f, sp, args)[2])
                    return cat
                # determines equality of all arguments
                elif func == 'equals':
                    arg1 = self.parse(0, line, f, sp, args)[2]
                    return all(self.parse(i, line, f, sp, args)[-1] == arg1 for i in range(1, len(args)))
                # nots a bool
                elif func == 'not':
                    return not self.parse(0, line, f, sp, args)[2]
                # ands two bools
                elif func == 'and':
                    return all(self.parse(i, line, f, sp, args)[2] for i in range(len(args)))
                # ors two bools
                elif func == 'or':
                    return self.parse(0, line, f, sp, args)[2] or self.parse(1, line, f, sp, args)[2]
                try:
                    # comparing numbers
                    if func == 'greater' or func == 'g':
                        return self.parse(0, line, f, sp, args)[2] > self.parse(1, line, f, sp, args)[2]
                    elif func == 'less' or func == 'l':
                        return self.parse(0, line, f, sp, args)[2] < self.parse(1, line, f, sp, args)[2]
                    elif func == 'greaterequal' or func == 'ge':
                        return self.parse(0, line, f, sp, args)[2] >= self.parse(1, line, f, sp, args)[2]
                    elif func == 'lessequal' or func == 'le':
                        return self.parse(0, line, f, sp, args)[2] <= self.parse(1, line, f, sp, args)[2]
                except Exception as e:
                    self.err(
                        'Comparison error',
                        f'Unable to compare values\n{e}',
                        line, lines_ran
                    )
                # data structure for holding multiple items
                if func == 'class':
                    # new interpreter
                    inter = Interpreter()
                    # log self
                    inter.parent = self
                    inter.trying = self.trying
                    # extract class name
                    name = self.parse(0, line, f, sp, args)[2]
                    # name must be a varname
                    self.check_varname(name, line)
                    # execute the block in the private environment
                    inter.execute(args[1][0])
                    # creates a variable out of the new interpreters resources
                    obj_to_add = {}
                    for varname in inter.vars:
                        val = inter.vars[varname].value
                        obj_to_add[varname] = Var(varname, val)
                    for methodname in inter.methods:
                        obj_to_add[methodname] = Var(
                            f"{methodname}#method", inter.methods[methodname])
                    self.vars[name] = Var(name, obj_to_add)
                    return obj_to_add

                # gets the first argument at the second argument
                elif func == 'get':
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    index = self.parse(1, line, f, sp, args)[2]
                    # index must be int or str
                    self.type_err([(index, (int, str))], line, lines_ran)
                    try:
                        return iterable[index]
                    except IndexError:
                        self.err(
                            'Index Error',
                            f'Index out of bounds: {index} (you) > {str(len(iterable))} (max)',
                            line,
                            lines_ran
                        )
                # gets the first argument as an iterable at the indices specified
                # can take unlimited arguments for unlimited indexing
                elif func == 'getn':
                    # get iterable
                    iterable = self.parse(0, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    try:
                        # must have at least one index
                        ind1 = self.parse(1, line, f, sp, args)[2]
                        # get at the index
                        obj = iterable[ind1]
                        # get the rest of the indices
                        for i in range(2, len(args)):
                            # get the index
                            ind = self.parse(i, line, f, sp, args)[2]
                            # index must be int or str
                            self.type_err([(ind, (int, str))], line, lines_ran)
                            # get at the index
                            obj = obj[ind]
                        return obj
                    except:
                        self.err(
                            'Error in getn()',
                            'Could not index the iterable',
                            line, lines_ran
                        )
                # get the keys of the first argument
                elif func == 'keys':
                    arg = None
                    try:
                        return (arg := self.parse(0, line, f, sp, args)[2]).keys()
                    except:
                        self.err(
                            'Error getting keys',
                            f'Argument must be a dictionary\nYou said: {arg}',
                            line, lines_ran
                        )
                # imports resources from another location
                elif func == 'import' or func == 'launch' or func == 'include' or func == 'using':
                    # for each import
                    for i in range(len(args)):
                        if script := self.imp(i, line, f, sp, args, self.imports):
                            self.execute(script)
                    return
                # creating variable domains
                elif func == 'domain':
                    # get the name of the domain to create
                    domain_name = self.parse(0, line, f, sp, args)[2]
                    # domain_name must be a varname
                    self.check_varname(domain_name, line)
                    # domains cannot coexist
                    if domain_name in self.domains:
                        self.err(
                            'Domain Error',
                            f'Domain "{domain_name}" already exists',
                            line,
                            lines_ran
                        )
                    # add the domain_name to the set
                    self.domains.add(domain_name)
                    # interpret the block in a new interpreter
                    # with the domain_name as the parent
                    inter = Interpreter()
                    inter.parent = self
                    inter.interpret(args[1][0])
                    # throws a domain error
                    def domain_err(name, object):
                        self.err(
                            'Domain Error',
                            f'Domain object name already claimed: "{name}", consider renaming the object',
                            line,
                            lines_ran
                        )
                    # for each variable in the interpreter
                    for varname in inter.vars:
                        name = f"{domain_name}:{varname}"
                        # name cannot already exist
                        if name in self.vars:
                            domain_err(name, self.vars)
                        self.vars[name] = inter.vars[varname]
                    # do the same for methods
                    for methodname in inter.methods:
                        name = f"{domain_name}:{methodname}"
                        if name in self.methods:
                            domain_err(name, self.methods)
                        self.methods[name] = inter.methods[methodname]
                    return domain_name
                # utilizing variable domains
                # 2.0.387
                elif func == 'domain:find':
                    # get the domain directory
                    # this directory is foramatted such that
                    # directory:directory:directory
                    domain_dir = self.parse(0, line, f, sp, args)[2]
                    # domain_dir must be a varname
                    self.check_varname(domain_dir, line)
                    # get all variables in the domain
                    # with the domain_dir chopped off
                    domain_vars = {}
                    # for each variable in the domain
                    for varname in self.vars:
                        # if the variable is in the domain
                        if varname.startswith(domain_dir):
                            # get the variable name without the domain
                            varname = varname[len(domain_dir):]
                            # add the variable to the domain_vars
                            domain_vars[varname] = self.vars[varname]
                    # return the domain_vars
                    return domain_vars
                # imports values from an enclosing Python script
                elif func == 'in':
                    reserved_name = '_msn2_reserved_in__'
                    if reserved_name not in self.vars:
                        return None
                    inval = self.vars[reserved_name].value
                    # if no arguments, return the value
                    if args[0][0] == '':
                        return inval
                    # if 1 argument, get index of value from input
                    elif len(args) == 1:
                        ind = self.parse(0, line, f, sp, args)[2]
                        # ind must be int
                        self.type_err([(ind, (int,))], line, lines_ran)
                        return inval[ind]
                    # if 2 arguments, get slice of input
                    elif len(args) == 2:
                        start = self.parse(0, line, f, sp, args)[2]
                        # start must be int
                        self.type_err([(start, (int,))], line, lines_ran)
                        end = self.parse(1, line, f, sp, args)[2]
                        # end must be int
                        self.type_err([(end, (int,))], line, lines_ran)
                        return inval[start:end]
                    return inval
                # exports values to an enclosing Python script
                elif func == 'out':
                    outting = [self.parse(i, line, f, sp, args)[2]
                               for i in range(len(args))]
                    _out = '_msn2_reserved_out__'
                    # create a variable for the enclosing Python script
                    # to access
                    self.vars[_out] = Var(_out, outting)
                    # return outting
                    return outting
                # interpreter printing mechanism
                elif func == 'prnt':
                    for i in range(len(args)):
                        srep = str(self.parse(i, line, f, sp, args)[2])
                        if i != len(args) - 1:
                            self.out += srep + ' '
                        else:
                            self.out += srep + '\n'
                    return first
                # python print
                elif func == 'print':
                    ret = None
                    for i in range(len(args)):
                        ret = self.parse(i, line, f, sp, args)[2]
                        if i != len(args) - 1:
                            print(ret, end=" ", flush=True)
                        else:
                            print(ret, flush=True)
                    return ret
                # print a box with Interpreter.bordered()
                # to get a string representation of the current argument
                # surrounded by a box
                elif func == 'print:box':
                    ret = None
                    for i in range(len(args)):
                        ret = self.parse(i, line, f, sp, args)[2]
                        if i != len(args) - 1:
                            print(Interpreter.bordered(
                                str(ret)), end=" ", flush=True)
                        else:
                            print(Interpreter.bordered(str(ret)), flush=True)
                    return ret
                # prints with styled colors
                # use self.styled_print() to print
                elif func == 'print:color':
                    # obtain a list of each argument
                    print_args = [self.parse(i, line, f, sp, args)[
                        2] for i in range(len(args))]
                    # print each argument
                    return self.styled_print(print_args)
                # sleeps the thread for the first argument amount of seconds
                elif func == "sleep":
                    import time
                    delay = self.parse(0, line, f, sp, args)[2]
                    # delay must be int or float
                    self.type_err([(delay, (int, float))], line, lines_ran)
                    return time.sleep(delay)
                # returns this interpreter
                elif func == 'me':
                    return self.me()
                # provides a representation of the current environment
                elif func == 'env':
                    should_print = False
                    if args[0][0] != '':
                        should_print = True
                    # first argument should be either string or integer
                    first = self.parse(0, line, f, sp, args)[2]
                    strenv = "--------- environment ---------"
                    strenv += f"\nout:\n{self.out}"
                    strenv += "\nvariables:\n"
                    for varname, v in self.vars.items():
                        try:
                            strenv += f"\t{varname} = {self.shortened(v.value)}\n"
                        except:
                            None
                    strenv += "\nmethods:\n"
                    # printing methods
                    for methodname, Method in self.methods.items():
                        strenv += f"\t{methodname}("
                        for i in range(len(Method.args)):
                            arg = Method.args[i]
                            if i != len(Method.args) - 1:
                                strenv += f"{arg}, "
                            else:
                                strenv += str(arg)
                        strenv += f") : {len(Method.body)} inst\n"
                    # printing macros
                    strenv += "\nmacros:\n\t"
                    # adding regular macros
                    if len(macros) > 0:
                        strenv += "premacros:\n\t\t"
                        for macro in macros:
                            # strenv += macro + "\n\t\t"
                            strenv += f"{macro}\n\t\t"
                    if len(postmacros) > 0:
                        strenv += "\n\tpostmacros:\n\t\t"
                        for macro in postmacros:
                            #strenv += macro + "\n\t\t"
                            strenv += f"{macro}\n\t\t"
                    if len(syntax) > 0:
                        strenv += "\n\tsyntax:\n\t\t"
                        for macro in syntax:
                            strenv += f"{macro}\n\t\t"
                    if len(enclosed) > 0:
                        strenv += "\n\tenclosedsyntax:\n\t\t"
                        for macro in enclosed:
                            strenv += f"{macro}\n\t\t"
                    strenv += f"\nlog:\n{self.log}\n-------------------------------"
                    if should_print:
                        self.styled_print([{
                            'text': strenv,
                            'style': 'bold',
                            'fore': 'blue'
                        }])
                    return strenv

                # changes the max chars printed with the env() call
                elif func == 'env:maxchars':
                    # if no arguments, return the current maxchars
                    if args[0][0] == '':
                        return self.env_max_chars
                    # otherwise, a single variable was provided
                    # this will change the max chars
                    new_maxchars = self.parse(0, line, f, sp, args)[2]                
                    # check for type errors
                    self.type_err([(new_maxchars, (int,))], line, lines_ran)
                    # alter the max chars
                    self.env_max_chars = new_maxchars
                    return new_maxchars

                # arithmetic, equivalent to the op class
                # executes MSNScript2 from its string representation
                try:
                    if func == '-':
                        if len(args) == 1:
                            return self.interpret(self.parse(0, line, f, sp, args)[2])
                        # subtracts all arguments from the first argument
                        else:
                            ret = self.parse(0, line, f, sp, args)[2]
                            try:
                                ret = ret.copy()
                            except AttributeError:
                                None
                            for i in range(1, len(args)):
                                ret -= self.parse(i, line, f, sp, args)[2]
                            return ret
                    elif func == '+':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret += self.parse(i, line, f, sp, args)[2]
                        return ret
                    elif func == 'x':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret *= self.parse(i, line, f, sp, args)[2]
                        return ret
                    elif func == '/':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret /= self.parse(i, line, f, sp, args)[2]
                        return ret
                    elif func == '//':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret //= self.parse(i, line, f, sp, args)[2]
                        return ret
                    elif func == '%':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret %= self.parse(i, line, f, sp, args)[2]
                        return ret
                    elif func == '^':
                        ret = self.parse(0, line, f, sp, args)[2]
                        try:
                            ret = ret.copy()
                        except AttributeError:
                            None
                        for i in range(1, len(args)):
                            ret **= self.parse(i, line, f, sp, args)[2]
                        return ret
                except Exception as e:
                    self.err(
                        'Error',
                        e, line, lines_ran
                    )

                # determines if a string is a digit
                if func == 'isdigit':
                    return self.parse(0, line, f, sp, args)[2].isdigit()
                # determines if a string is alpha
                elif func == 'isalpha':
                    return self.parse(0, line, f, sp, args)[2].isalpha()
                # does something with a value as a temporary
                # variable
                elif func == 'as':
                    # temporary variable name
                    varname = self.parse(0, line, f, sp, args)[2]
                    # varname must be a varname
                    self.check_varname(varname, line)
                    # block to execute
                    block = args[2][0]
                    # set the variable
                    self.vars[varname] = Var(
                        varname, self.parse(1, line, f, sp, args)[2])
                    # execute the block
                    ret = self.interpret(block)
                    # delete the variable
                    del self.vars[varname]
                    return ret
                # startswith
                elif func == 'startswith':
                    # arg
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    # prefix
                    prefix = self.parse(1, line, f, sp, args)[2]
                    # prefix must be str
                    self.type_err([(prefix, (str,))], line, lines_ran)
                    return arg.startswith(prefix)
                # endswith
                elif func == 'endswith':
                    # arg
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    # suffix
                    suffix = self.parse(1, line, f, sp, args)[2]
                    # suffix must be str
                    self.type_err([(suffix, (str,))], line, lines_ran)
                    return arg.endswith(suffix)
                # strips a str
                elif func == 'strip':
                    try:
                        return self.parse(0, line, f, sp, args)[2].strip()
                    except Exception as e:
                        self.err(
                            'Error in strip()',
                            e, line, lines_ran
                        )
                # gets a slice
                elif func == 'slice':
                    # first
                    first = self.parse(0, line, f, sp, args)[2]
                    # first must be slicable
                    self.check_iterable(first, line)
                    # second
                    second = self.parse(1, line, f, sp, args)[2]
                    # second must be int
                    self.type_err([(second, (int,))], line, lines_ran)
                    # third
                    third = self.parse(2, line, f, sp, args)[2]
                    # third must be int
                    self.type_err([(third, (int,))], line, lines_ran)
                    return first[second:third]
                # joins a str
                elif func == 'iterable:join':
                    delimiter = self.parse(0, line, f, sp, args)[2]
                    # delimiter must be str
                    self.type_err([(delimiter, (str,))], line, lines_ran)
                    # iterable
                    iterable = self.parse(1, line, f, sp, args)[2]
                    # iterable must be iterable
                    self.check_iterable(iterable, line)
                    return delimiter.join(iterable)
                # returns the MSNScript2 passed as a string
                elif func == 'async' or func == 'script' or func == 'HTML':
                    # inserts key tokens
                    return self.msn2_replace(args[0][0])
                # gets the current time
                elif func == 'now':
                    import time
                    return time.time()
                # creates a private execution enviroment
                # private block will have read access to the enclosing Interpreter's
                # variables and methods
                elif func == 'private' or func == 'inherit:all':
                    inter = Interpreter()
                    inter.parent = self
                    for vname, entry in self.vars.items():
                        try:
                            inter.vars[vname] = Var(vname, entry.value)
                        except:
                            inter.vars[vname] = Var(vname, entry)
                    for mname, entry in self.methods.items():
                        inter.methods[mname] = entry
                    return inter.interpret(args[0][0])
                # breaks out of the working context
                elif func == 'break':
                    self.breaking = True
                    return
                # reverses the first argument
                elif func == 'reverse':
                    # arg
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be iterable
                    self.check_iterable(arg, line)
                    return arg[::-1]
                # performs upper()
                elif func == 'upper':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    return arg.upper()
                # performs lower()
                elif func == 'lower':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    return arg.lower()
                # performs title()
                elif func == 'title':
                    arg = self.parse(0, line, f, sp, args)[2]
                    # arg must be str
                    self.type_err([(arg, (str,))], line, lines_ran)
                    # return title
                    return arg.title()
                # inherits methods only from the parent context
                elif func == 'inherit:methods':
                    for methodname in self.parent.methods:
                        self.methods[methodname] = self.parent.methods[methodname]
                    return True
                # inherits variables from parent context
                elif func == 'inherit:vars':
                    for varname in self.parent.vars:
                        self.vars[varname] = self.parent.vars[varname]
                    return True
                # inherits a single variable or function
                elif func == 'inherit:single':
                    name = self.parse(0, line, f, sp, args)[2]
                    # name must be a varname
                    self.check_varname(name, line)
                    if name in self.parent.vars:
                        self.vars[name] = self.parent.vars[name]
                    elif name in self.parent.methods:
                        self.methods[name] = self.parent.methods[name]
                    return True
                # creates a new execution environment
                # new env is executed by a fresh interpreter
                # nothing is inherited from parent context
                elif func == 'new' or func == 'inherit:none':
                    inter = Interpreter()
                    inter.parent = self
                    return inter.interpret(args[0][0])
                # sets the python alias
                # until the program stops
                elif func == 'alias':
                    new_al = self.parse(0, line, f, sp, args)[2]
                    # new_al must be str
                    self.type_err([(new_al, (str,))], line, lines_ran)
                    return (python_alias := new_al)
                # starts a new process with the first argument as the target
                elif func == 'process':
                    # path to the process to run
                    path = self.parse(0, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    # if windows:
                    if os.name == 'nt':
                        import subprocess
                        # runs the process
                        sub = subprocess.run(
                            args=[python_alias, 'msn2.py', path], shell=True)
                        self.processes[path] = sub
                        return sub
                    # if linux
                    elif os.name == 'posix':
                        self.err(
                            'POSIX not yet implemented',
                            '',
                            line,
                            lines_ran
                        )
                        return None
                    return None
                # starts a process via MSN2 code
                elif func == 'proc':
                    # import lib processes module if not imported
                    if 'lib/processes.msn2' not in self.imports:
                        self.interpret('import("lib/processes")')
                    # import the processes library and
                    # create a new process
                    name = self.parse(0, line, f, sp, args)[2]
                    return self.interpret(
                        f"(import('lib/processes'),processes:fork('{name}',private(async({args[1][0]}))))")
                # gets the pid of the working process
                elif func == 'pid':
                    return os.getpid()
                # creates a new thread to execute the block, thread
                # starts on the same interpreter
                elif func == "thread":
                    # name not provided
                    if len(args) == 1:
                        global thread_serial
                        name = f"__msn2_thread_id_{thread_serial}"
                        block = args[0][0]
                    # name provided (2 arguments provided)
                    else:
                        name = str(self.parse(0, line, f, sp, args)[2])
                        # name must be a varname
                        self.check_varname(name, line)
                        block = args[1][0]
                    thread = threading.Thread(
                        target=self.interpret, args=(block,))
                    thread.name = name
                    self.threads[name] = [thread, self]
                    thread.start()
                    return True
                # creates a thread pool to execute the block
                elif func == "threadpool":
                    import concurrent.futures
                    # get the amount of threads to create
                    # create the thread pool
                    # submit the block to the pool
                    max_workers = self.parse(0, line, f, sp, args)[2]
                    # max_workers must be int
                    self.type_err([(max_workers, (int,))], line, lines_ran)
                    concurrent.futures.ThreadPoolExecutor(max_workers).submit(self.interpret, args[1][0])
                    return True

                # creates or edits thread variable
                elif func == 'tvar':
                    # thread name
                    name = str(self.parse(0, line, f, sp, args)[2])
                    # variable name
                    varname = str(self.parse(1, line, f, sp, args)[2])
                    # variable value
                    val = self.parse(2, line, f, sp, args)[2]
                    # thread var name
                    tvarname = f"_msn2_tvar_{name}_{varname}"
                    # sets a thread specific variable
                    self.vars[tvarname] = Var(varname, val)
                    return val
                # gets a thread variable
                elif func == 'gettvar':
                    # gets the variable
                    return self.vars[f"_msn2_tvar_{self.parse(0, line, f, sp, args)[2]}_{self.parse(1, line, f, sp, args)[2]}"].value
                # creates a string variable name for functions
                # that require a string variable name
                elif func == 'tvarstr':
                    # returns the string
                    return f"_msn2_tvar_{self.parse(0, line, f, sp, args)[2]}_{self.parse(1, line, f, sp, args)[2]}"
                # interprets a variable by variable name
                # a and a variable method
                elif func == 'varmethod':
                    # variable name
                    return self.interpret(f"{self.parse(0, line, f, sp, args)[2]}.{args[1][0]}")
                # acquires the global lock
                elif func == 'acquire':
                    return auxlock.acquire()
                # releases the global lock
                elif func == 'release':
                    return auxlock.release()
                # acquires the pointer lock
                elif func == 'acquire:pointer':
                    return pointer_lock.acquire()
                elif func == 'release:pointer':
                    return pointer_lock.release()
                # joins the current working thread with the thread name specified
                elif func == 'join':
                    for i in range(len(args)):
                        name = self.parse(i, line, f, sp, args)[2]
                        thread = self.thread_by_name(name)
                        while thread == None:
                            None
                        thread.join()
                    return True
                # exits the working thread
                elif func == 'stop':
                    return os._exit(0)
                # tries the first argument, if it fails, code falls to the catch/except block
                # there is no finally implementation
                elif func == 'try':
                    ret = None
                    self.trying = True
                    try:
                        ret = self.interpret(args[0][0])
                    except:
                        self.trying = False
                        # if there is another argument
                        if len(args) == 2:
                            ret = self.interpret(args[1][0])
                    self.trying = False
                    return ret
                # waits for a certain condition to be true
                elif func == 'wait':
                    ret = None
                    # no block per tick provided
                    if len(args) == 1:
                        while not (ret := self.interpret(args[0][0])):
                            None
                    # block provided
                    elif len(args) == 2:
                        while not (ret := self.interpret(args[0][0])):
                            self.interpret(args[1][0])
                    # block with tick provided
                    elif len(args) == 3:
                        import time
                        s = self.parse(2, line, f, sp, args)[2]
                        # s must be int or float
                        self.type_err([(s, (int, float))], line, lines_ran)
                        while not (ret := self.interpret(args[0][0])):
                            self.interpret(args[1][0])
                            time.sleep(s)
                    return ret

                # performs a an action every certain amount of seconds
                # where the amount of seconds is the first argument
                # the block is the second argument
                # third argument is optional, and is the amount of seconds
                # the interval should last for, if not provided, the interval
                # will last forever
                if func == 'interval':
                    import time
                    # amount of seconds
                    seconds = self.parse(0, line, f, sp, args)[2]
                    # seconds must be int or float
                    self.type_err([(seconds, (int, float))], line, lines_ran)
                    # if the interval should last for a certain amount of seconds
                    # should account for the first argument to correctly wait
                    if len(args) == 3:
                        extra = self.parse(2, line, f, sp, args)[2]
                        # extra must be int or float
                        self.type_err([(extra, (int, float))], line, lines_ran)
                        # if time is negative, we set it to infinity
                        if extra == -1:
                            extra = float('inf')
                        end = time.time() + extra
                        while time.time() < end:
                            time.sleep(seconds)
                            self.interpret(args[1][0])
                    else:
                        while True:
                            time.sleep(seconds)
                            self.interpret(args[1][0])
                # exports a quantity of variables or methods from the working context to the parent context,
                # ex private context -> boot context
                elif func == 'export':
                    # if last argument is True,
                    # we add the variables to the parent context's variable
                    last_arg = self.parse(len(args) - 1, line, f, sp, args)[2]
                    for i in range(len(args)):
                        varname = self.parse(i, line, f, sp, args)[2]
                        # varname must be a varname
                        self.check_varname(varname, line)
                        if varname in self.vars:
                            if isinstance(last_arg, bool):
                                # if self.vars[varname].value is any type of number
                                if isinstance(self.vars[varname].value, (int, float, complex)):
                                    self.parent.vars[varname].value += self.vars[varname].value
                                # otherwise add every element to the parent context's variable
                                elif isinstance(self.vars[varname].value, list):
                                    for element in self.vars[varname].value:
                                        self.parent.vars[varname].value.append(
                                            element)
                            else:
                                self.parent.vars[varname] = self.vars[varname]
                        elif varname in self.methods:
                            self.parent.methods[varname] = self.methods[varname]
                    return True
                # exports a single value as the variable name
                # first argument is the new variable name
                # second is the value to export
                elif func == 'exportas':
                    # variable name
                    varname = self.parse(0, line, f, sp, args)[2]
                    # varname must be a varname
                    self.check_varname(varname, line)
                    # value
                    val = self.parse(1, line, f, sp, args)[2]
                    # export to parent context
                    self.parent.vars[varname] = Var(varname, val)
                    return val
                # exports all variables and methods to the parent context
                elif func == 'exportall':
                    for varname in self.vars:
                        self.parent.vars[varname] = self.vars[varname]
                    for methodname in self.methods:
                        self.parent.methods[methodname] = self.methods[methodname]
                    return True
                # sends a command to the console, console type depends on
                # executors software of course
                elif func == 'console':
                    # os.system all arguments,
                    # returns the console output
                    # of the last argument
                    ret = None
                    for i in range(len(args)):
                        command = self.parse(i, line, f, sp, args)[2]
                        # command must be str
                        self.type_err([(command, (str,))], line, lines_ran)
                        ret = os.system(command)
                    return ret
                # Execute the command and capture the output
                # only takes one argument
                elif func == 'console:read':
                    import subprocess
                    # returns the console output
                    # of the last argument
                    command = self.parse(0, line, f, sp, args)[2]
                    # command must be str
                    self.type_err([(command, (str,))], line, lines_ran)
                    process = subprocess.run(command, shell=True, capture_output=True, text=True)
                    if process.returncode == 0:
                        return process.stdout
                    else:
                        return process.stderr

                # performs a get request to an http server
                # first parameter is the URL
                # second parameter is a map of parameters to sent as a request
                elif func == 'request':
                    import requests
                    # get URL to request from
                    url = self.parse(0, line, f, sp, args)[2]
                    # url must be str
                    self.type_err([(url, (str,))], line, lines_ran)
                    try:
                        # get parameters
                        params = self.parse(1, line, f, sp, args)[2]
                    except:
                        params = None
                    # return response
                    return requests.get(url=url, params=params).json()

                # requires thread-safe context, see /demos/protected.msn2
                # simulates returning of the function currently running
                # should be used cautiously, if you dont know whether to use return() or var()
                # to return a value, use var()
                elif func == 'return':
                    method = self.methods[self.loggedmethod[-1]]
                    # evaluate returning literal
                    ret = self.parse(0, line, f, sp, args)[2]
                    # set return variable
                    ret_name = method.returns[0]
                    # if not a variable
                    if ret_name not in self.vars:
                        self.vars[ret_name] = Var(ret_name, None)
                    self.vars[ret_name].value = ret
                    return ret
                # gets the public IP address of the machine
                elif func == 'pubip':
                    import requests
                    # asks an api server for this address
                    return requests.get('https://api.ipify.org').text
                # gets the private ips of this machine
                elif func == 'privips':
                    import socket
                    return socket.gethostbyname_ex(socket.gethostname())[2]
                # starts an api endpoint
                elif func == 'ENDPOINT':
                    # imports
                    from flask import Flask
                    from flask_restful import Api, Resource
                    import logging
                    # initial API endpoint data
                    path = None
                    init_data = {}
                    port = 5000
                    host = '127.0.0.1'
                    last_arg = None
                    # 1 argument, defaults to 127.0.0.1:5000/path = {}
                    if len(args) == 1:
                        # path to endpoint
                        path = self.parse(0, line, f, sp, args)[2]
                        # path should be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        last_arg = path
                    # 2 arguments, defaults to 127.0.0.1:5000/path = init_data
                    elif len(args) == 2:
                        # path to endpoint
                        path = self.parse(0, line, f, sp, args)[2]
                        # path should be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        # json to initialize at the endpoint
                        init_data = self.parse(1, line, f, sp, args)[2]
                        # init_data should be dict
                        self.type_err([(init_data, (dict,))], line, lines_ran)
                        last_arg = init_data
                    # 3 arguments, defaults to host:port/path = init_data
                    else:
                        # host to endpoint as first argument
                        host = self.parse(0, line, f, sp, args)[2]
                        # host should be str
                        self.type_err([(host, (str,))], line, lines_ran)
                        # port to endpoint as second argument
                        port = self.parse(1, line, f, sp, args)[2]
                        # port should be int
                        self.type_err([(port, (int,))], line, lines_ran)
                        # path to endpoint
                        path = self.parse(2, line, f, sp, args)[2]
                        # path should be str
                        self.type_err([(path, (str,))], line, lines_ran)
                        # json to initialize at the endpoint
                        init_data = self.parse(3, line, f, sp, args)[2]
                        # init_data should be dict
                        self.type_err([(init_data, (dict,))], line, lines_ran)
                        last_arg = init_data
                        if len(args) == 5:
                            last_arg = self.parse(4, line, f, sp, args)[2]
                    # prepare endpoint
                    print('serving on http://' + host + ':' + str(port) + path)
                    app = Flask(__name__)
                    cors = False
                    # if the last argument is a string with 'CORS' in it
                    # then enable CORS
                    if isinstance(last_arg, str) and 'CORS' in last_arg:
                        from flask_cors import CORS
                        # enable CORS
                        print('starting with cors')
                        cors = True
                        CORS(app)
                    # disable flask messages that aren't error-related
                    log = logging.getLogger('werkzeug')
                    log.disabled = True
                    app.logger.disabled = True
                    # gets Flask Api
                    api = Api(app)
                    # api endpoint class
                    class EndPoint(Resource):
                        @classmethod
                        def make_api(cls, response):
                            cls.response = response
                            return cls
                        # GET
                        def get(self):
                            return self.response
                        # POST
                        def post(self):
                            from flask import request
                            # obtains current endpoint data
                            current_data = self.response
                            # updates current data with data to post
                            current_data.update(request.get_json())
                            # updates next response
                            self.make_api(current_data)
                            return current_data
                        # DELETE

                        def delete(self):
                            self.make_api({})
                            return self.response
                    curr_endpoint = EndPoint.make_api(init_data)
                    # logs newly created endpoint
                    self.endpoints[path] = api
                    # adds class EndPoint as a Resource to the Api with the specific path
                    # passes arg2 alongside
                    api.add_resource(curr_endpoint, path)
                    # starting flask server
                    try:
                        # if internal
                        app.run(host=host, port=port,
                                debug=False, use_reloader=False)
                    except:
                        # if external
                        try:
                            if __name__ == '__main__':
                                app.run(host='0.0.0.0', port=port,
                                        debug=False, use_reloader=False)
                        except:
                            None
                    return api
                # posts to an api endpoint
                elif func == 'POST':
                    import requests
                    # url to post to, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]
                    # host must be str
                    self.type_err([(host, (str,))], line, lines_ran)
                    # port to post to
                    port = self.parse(1, line, f, sp, args)[2]
                    # port must be int
                    self.type_err([(port, (int,))], line, lines_ran)
                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    # data to post
                    data = self.parse(3, line, f, sp, args)[2]
                    # data must be dict
                    self.type_err([(data, (dict,))], line, lines_ran)
                    # if local network
                    if host == '0.0.0.0':
                        response = requests.post(
                            url=('http://127.0.0.1:' + str(port) + path), json=data)
                    # if localhost
                    else:
                        # post to endpoint
                        response = requests.post(
                            url=('http://' + host + ':' + str(port) + path), json=data)
                    # get response
                    return response.json()
                # gets from an api endpoint
                elif func == 'GET':
                    import requests
                    # url to get from, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]
                    # host must be str
                    self.type_err([(host, (str,))], line, lines_ran)
                    # port to get from
                    port = self.parse(1, line, f, sp, args)[2]
                    # port must be int
                    self.type_err([(port, (int,))], line, lines_ran)
                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    # if local network
                    if host == '0.0.0.0':
                        return requests.get(url=('http://127.0.0.1:' + str(port) + path)).json()
                    # if localhost
                    else:
                        return requests.get(url=('http://' + host + ':' + str(port) + path)).json()
                # deletes from an api endpoint
                elif func == 'DELETE':
                    import requests
                    # url to delete from, defaults to localhost
                    host = self.parse(0, line, f, sp, args)[2]
                    # host must be str
                    self.type_err([(host, (str,))], line, lines_ran)
                    # port to delete from
                    port = self.parse(1, line, f, sp, args)[2]
                    # port must be int
                    self.type_err([(port, (int,))], line, lines_ran)
                    # path after url
                    path = self.parse(2, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    if host == '0.0.0.0':
                        response = requests.delete(
                            url=('http://127.0.0.1:' + str(port) + path))
                    else:
                        # delete from endpoint
                        response = requests.delete(
                            url=('http://' + host + ':' + str(port) + path))
                    return response.json()
                # determines if the system is windows or not
                elif func == 'windows':
                    return os.name == 'nt'
                # determines if system is linux
                elif func == 'linux':
                    return os.name == 'posix'
                # determines if system is mac
                elif func == 'mac':
                    import sys
                    return sys.platform == 'darwin'
                # simulates function closure
                elif func == 'end':
                    method = self.methods[self.loggedmethod[-1]]
                    self.loggedmethod.pop()
                    method.ended = True
                    return True
                # gets the value from a Var object
                elif func == 'static':
                    try:
                        return self.parse(0, line, f, sp, args)[2].value
                    except Exception as e:
                        self.err(
                            'Error in static()',
                            e, line, lines_ran
                        )
                # object instance requested
                # if the function is in the variables
                # and the variable is a class
                elif func in self.vars and isinstance(self.vars[func].value, dict):
                    # get classname to create
                    classname = func
                    # template Var obj to create from
                    var_obj = self.vars[classname].value
                    # new class instance
                    instance = {}
                    curr_arg_num = 0
                    # attributes to apply
                    for name in var_obj:
                        # if attribute is a method
                        if isinstance(var_obj[name].value, self.Method):
                            # add the method to the instance
                            instance[name] = var_obj[name].value
                            # if the method's name is 'const'
                            if var_obj[name].value.name == 'const':
                                # run the function with the argument being
                                # this instance
                                var_obj[name].value.run(
                                    [instance], self, actual_args=args[1:])
                            continue
                        # if attribute is a variable
                        # value can be None
                        try:
                            instance[name] = self.parse(
                                curr_arg_num, line, f, sp, args)[2]
                            if instance[name] == None:
                                instance[name] = self.vars[classname].value[name].value
                        # if not specified, field is default value
                        except:
                            try:
                                instance[name] = var_obj.value[name].copy()
                            except:
                                instance[name] = var_obj[name].value
                        curr_arg_num += 1
                    return instance
                # gets an attribute of an instance of a class
                elif func == 'getattr':
                    vn = self.parse(0, line, f, sp, args)[2]
                    # vn must be a varname
                    self.check_varname(vn, line)
                    # get the attribute
                    return self.vars[vn].value[self.parse(1, line, f, sp, args)[2]]
                # sets an attribute of an instance of a class
                elif func == 'setattr':
                    # current working object
                    o = self.vars[self.parse(0, line, f, sp, args)[2]].value
                    # name of attribute to set
                    attr = self.parse(1, line, f, sp, args)[2]
                    # attr must be a string
                    self.type_err([(attr, (str,))], line, lines_ran)
                    # value to set
                    val = self.parse(2, line, f, sp, args)[2]
                    # set the value
                    o[attr] = val
                    return val

                # PRACTICAL FUNCTIONALITY
                # started 5/20/2023

                # starts and retrieves an instance
                # of an application on the local machine
                # only properly implemented for Windows
                # uses pywinauto to do all of this
                # one argument: path to application
                # any more arguments: the existing application doesn't close
                elif func == 'app':
                    global timings_set
                    # set timings if not already set
                    if not timings_set:
                        from pywinauto import timings
                        # pywinauto defaults
                        timings.Timings.after_clickinput_wait = 0.001
                        timings.Timings.after_setcursorpos_wait = 0.001
                        timings.Timings.after_sendkeys_key_wait = 0.001
                        timings.Timings.after_menu_wait = 0.001
                    # get the path to the application
                    path = self.parse(0, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    # if there is not second argument, we do not kill any
                    # existing instances of the application
                    name = None
                    extension = None
                    if len(args) == 1:
                        # get the name and extension of the application
                        _sp = path.split('\\')
                        name = _sp[-1].split('.')[0]
                        extension = _sp[-1].split('.')[1]
                        # use taskkill to kill the application
                        # taskkill should end the program by name, and should kill
                        # all child processes forcefully, it should also not print
                        # anything to the console
                        os.system(
                            f'taskkill /f /im {name}.{extension} >nul 2>&1')
                    # creates an App variable
                    return self.App(path=path, name=name, extension=extension)
                # connects to the first argument given that
                # the first argument is an instance of self.App
                elif func == 'connect':
                    from pywinauto.application import Application
                    appl = self.parse(0, line, f, sp, args)[2]
                    a = Application(backend="uia").connect(
                        process=appl.application.process)
                    # connect to the application
                    return self.App(path=appl.path)
                # starts and retrieves an instance of an Excel workbook
                # using the openpyxl library
                #
                # this method works better than the app() system call
                # as it utilizes an Excel-specific library for stability
                # and speed
                #
                # creation of a workbook can be done with the 'file' msn2 class
                elif func == 'excel':
                    # automating Excel
                    import openpyxl
                    path = self.parse(0, line, f, sp, args)[2]
                    # path must be str
                    self.type_err([(path, (str,))], line, lines_ran)
                    # creates and returns a Workbook
                    return self.Workbook(openpyxl.load_workbook(path), path)
                # quicker conditional operator as functional prefix
                elif len(func) > 0 and func[0] == '?':
                    func = func[1:]
                    ret = None
                    if self.interpret(func):
                        ret = self.interpret(args[0][0])
                    else:
                        # else block is optional
                        try:
                            ret = self.interpret(args[1][0])
                        except:
                            None
                    return ret
                # executes C code and retrieves the environment
                elif func == 'C':
                    # get the C code
                    c_code = self.msn2_replace(args[0][0])
                    # create a directory for the C code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                    # create a file for the C code
                    # and write the C code to it
                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/c{file_num}.c'
                    with open(file_name, 'w') as f:
                        f.write(c_code)
                    # creates a new process
                    # and executes the C code
                    # returns the environment
                    # including the out and variables

                    def retrieve_c_environment(c_code):
                        import subprocess
                        # executable
                        executable = f'{exec_folder_path}/c{file_num}.exe'
                        # create a new process
                        # and execute the C code
                        compiled_code = subprocess.run(
                            ['gcc', file_name, '-o', executable],
                            # capture the output
                            capture_output=True,
                            text=True
                        )
                        # if there's an error, print it
                        if len(compiled_code.stderr) > 0:
                            return {'out': '', 'err': compiled_code.stderr}
                        # run the executable
                        compiled_code = subprocess.run(
                            [executable],
                            # capture the output
                            capture_output=True,
                            text=True
                        )
                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr
                        # get the environment
                        # env = out.split('\n')[-2]
                        # env = env.replace('\'', '"')
                        # env = json.loads(env)
                        return {'out': out, 'err': err}
                    # execute the C code
                    return retrieve_c_environment(c_code)
                # executes JavaScript code and retrieves the environment
                # no compilation is needed, the code is executed via
                # node __filename__.js
                elif func == 'JS':
                    # get the JavaScript code
                    js_code = self.msn2_replace(args[0][0])
                    # create a directory for the JavaScript code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                    # create a file for the JavaScript code
                    # and write the JavaScript code to it
                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/js{file_num}.js'
                    # if JS() has two arguments, the second is the name of
                    # the file, excluding .js
                    if len(args) == 2:
                        file_name = f'{exec_folder_path}/{self.parse(1, line, f, sp, args)[2]}.js'
                    with open(file_name, 'w') as f:
                        f.write(js_code)
                    # creates a new process
                    # and executes the JavaScript code
                    # returns the environment
                    # including the out and variables

                    def retrieve_js_environment(js_code):
                        import subprocess
                        # executable
                        executable = file_name
                        # create a new process
                        # and execute the JavaScript code
                        compiled_code = subprocess.run(
                            ['node', file_name],
                            # capture the output
                            capture_output=True,
                            text=True
                        )
                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr
                        # if there is an error, print it
                        if len(err) > 0:
                            print(err)
                        # remove a succeeding newline
                        # if it exists
                        if len(out) > 0 and out[-1] == '\n':
                            out = out[:-1]
                        return {'out': out, 'err': err}
                    # execute the JavaScript code
                    return retrieve_js_environment(js_code)
                # compiles and executes Java code and retrieves the environment
                elif func == 'JAVA':
                    java_code = self.msn2_replace(args[0][0])
                    # create a directory for the Java code
                    # if it does not exist
                    exec_folder_path = '_exec'
                    # if the folder does not exist, create it
                    if not os.path.exists(exec_folder_path):
                        os.mkdir(exec_folder_path)
                    # create a file for the Java code
                    # and write the Java code to it
                    # get the amount of files in the directory
                    # and use that as the file name
                    file_num = len(os.listdir(exec_folder_path))
                    file_name = f'{exec_folder_path}/java{file_num}.java'
                    # if JAVA() has two arguments, the second is the name of
                    # the file, excluding .java
                    if len(args) == 2:
                        file_name = f'{exec_folder_path}/{self.parse(1, line, f, sp, args)[2]}.java'
                    with open(file_name, 'w') as f:
                        f.write(java_code)
                    # creates a new process
                    # and executes the Java code
                    # returns the environment
                    # including the out and variables

                    def retrieve_java_environment(java_code):
                        import subprocess
                        # create a new process
                        # and execute the Java code
                        compiled_code = subprocess.run(
                            ['javac', file_name],
                            # capture the output
                            capture_output=True,
                            text=True,
                            shell=True
                        )
                        # run the code
                        compiled_code = subprocess.run(
                            ['java', '-cp', exec_folder_path, f'{file_name}'],
                            # capture the output
                            capture_output=True,
                            text=True,
                            shell=True
                        )
                        # get the output and error
                        out = compiled_code.stdout
                        err = compiled_code.stderr
                        # if there is an error, print it
                        if len(err) > 0:
                            print(err)
                        # remove a succeeding newline
                        # if it exists
                        if len(out) > 0 and out[-1] == '\n':
                            out = out[:-1]
                        return {'out': out, 'err': err}
                    # execute the Java code
                    return retrieve_java_environment(java_code)
                # inline function, takes any amount of instructions
                # returns the result of the last instruction
                elif func == "=>" or (func == '' and objfunc == ''):
                    ret = None
                    for i in range(len(args)):
                        ret = self.parse(i, line, f, sp, args)[2]
                    return ret
                # if the function, when parsed, is an integer,
                # then it is a loop that runs func times
                elif (_i := self.get_int(func)) != None:
                    ret = None
                    for _ in range(_i):
                        for arguments in args:
                            ins_s = arguments[0]
                            line, ret = self.convert_arg(
                                ins_s, line, f, sp, args)
                    return ret
                # # if the function is a variable name
                elif func in self.vars:
                    # value
                    val = self.vars[func].value
                    # if the variable is an integer,
                    # run the arguments as blocks inside
                    # that many times
                    if isinstance(val, int):
                        ret = None
                        for _ in range(val):
                            for arguments in args:
                                ins_s = arguments[0]
                                line, ret = self.convert_arg(
                                    ins_s, line, f, sp, args)
                        return ret
                    # otherwise return the value
                    return val
                # mouse pointer operations
                elif obj.startswith('pointer'):
                    from pywinauto import mouse
                    # thread based action?
                    p_thread = False
                    # return
                    ret = '<msnint2 class>'
                    # determine if thread based pointer action has been requested
                    if obj.endswith(':lock'):
                        p_thread = True
                        pointer_lock.acquire()
                    # gets the current position of the mouse
                    if objfunc == 'getpos' or objfunc == 'pos' or objfunc == 'position':
                        # import win32api
                        import win32api
                        ret = win32api.GetCursorPos()
                    # moves the mouse to an x, y position
                    elif objfunc == 'move' or objfunc == 'hover':
                        ret = mouse.move(coords=(self.parse(0, line, f, sp, args)[
                                         2], self.parse(1, line, f, sp, args)[2]))
                    # right clicks the mouse
                    elif objfunc == 'click' or objfunc == 'left_click':
                        # if args are provided
                        if len(args) == 2:
                            ret = mouse.click(coords=(self.parse(0, line, f, sp, args)[
                                              2], self.parse(1, line, f, sp, args)[2]))
                        # if no args are provided
                        else:
                            # import win32api
                            import win32api
                            ret = mouse.click(coords=win32api.GetCursorPos())
                    # right clicks the mouse
                    elif objfunc == 'right_click':
                        # if args are provided
                        if len(args) == 2:
                            start = self.parse(0, line, f, sp, args)[2]
                            # start must be int
                            self.type_err([(start, (int,))], line, lines_ran)
                            end = self.parse(1, line, f, sp, args)[2]
                            # end must be int
                            self.type_err([(end, (int,))], line, lines_ran)
                            ret = mouse.right_click(coords=(start, end))
                        # if no args are provided
                        else:
                            # import win32api
                            import win32api
                            ret = mouse.right_click(
                                coords=win32api.GetCursorPos())
                    # double clicks the mouse
                    elif objfunc == 'double_click':
                        # if args are provided
                        if len(args) == 2:
                            start = self.parse(0, line, f, sp, args)[2]
                            # start must be int
                            self.type_err([(start, (int,))], line, lines_ran)
                            end = self.parse(1, line, f, sp, args)[2]
                            # end must be int
                            self.type_err([(end, (int,))], line, lines_ran)
                            ret = mouse.double_click(coords=(start, end))
                        # if no args are provided
                        else:
                            import win32api
                            ret = mouse.double_click(
                                coords=win32api.GetCursorPos())
                    # scrolls the mouse wheel to the bottom of the page
                    elif objfunc == 'scroll_bottom':
                        import win32api
                        ret = mouse.scroll(
                            wheel_dist=9999999, coords=win32api.GetCursorPos())
                    # scrolls the mouse wheel to the top of the page
                    elif objfunc == 'scroll_top':
                        import win32api
                        ret = mouse.scroll(
                            wheel_dist=-9999999, coords=win32api.GetCursorPos())
                    elif objfunc == 'scroll':
                        import win32api
                        dist = self.parse(0, line, f, sp, args)[2]
                        # dist must be int
                        self.type_err([(dist, (int,))], line, lines_ran)
                        ret = mouse.scroll(wheel_dist=dist, coords=win32api.GetCursorPos())

                    # determines if the left mouse button is down
                    elif objfunc == 'left_down':
                        import win32api
                        ret = win32api.GetKeyState(0x01) < 0
                    # determines if the right mouse button is down
                    elif objfunc == 'right_down':
                        import win32api
                        ret = win32api.GetKeyState(0x02) < 0
                    # waits for the left button to be pressed
                    elif objfunc == 'wait_left':
                        import win32api
                        while True:
                            if win32api.GetKeyState(0x01) < 0:
                                break
                        ret = True
                    # waits for the right button to be pressed
                    elif objfunc == 'wait_right':
                        import win32api
                        while True:
                            if win32api.GetKeyState(0x02) < 0:
                                break
                        ret = True
                    # waits for a click
                    # waits for the left button to be pressed down
                    # then waits for it to be released
                    elif objfunc == 'wait_left_click':
                        import win32api
                        while True:
                            if win32api.GetKeyState(0x01) < 0:
                                break
                        while True:
                            if win32api.GetKeyState(0x01) >= 0:
                                break
                        ret = True
                    # waits for the right button to be pressed down
                    # then waits for it to be released
                    elif objfunc == 'wait_right_click':
                        import win32api
                        while True:
                            if win32api.GetKeyState(0x02) < 0:
                                break
                        while True:
                            if win32api.GetKeyState(0x02) >= 0:
                                break
                        ret = True
                    # DIRECTIONAL MOVEMENTS
                    # moves the mouse down from its current location
                    elif objfunc == 'down':
                        import win32api
                        curr_x, curr_y = win32api.GetCursorPos()
                        moving = self.parse(0, line, f, sp, args)[2]
                        # moving must be int
                        self.type_err([(moving, (int,))], line, lines_ran)
                        ret = mouse.move(
                            coords=(curr_x, curr_y + moving))
                    # moves the mouse up from its current location
                    elif objfunc == 'up':
                        import win32api
                        curr_x, curr_y = win32api.GetCursorPos()
                        moving = self.parse(0, line, f, sp, args)[2]
                        # moving must be int
                        self.type_err([(moving, (int,))], line, lines_ran)
                        ret = mouse.move(
                            coords=(curr_x, curr_y - moving))
                    # moves the mouse left from its current location
                    elif objfunc == 'left':
                        import win32api
                        curr_x, curr_y = win32api.GetCursorPos()
                        moving = self.parse(0, line, f, sp, args)[2]
                        # moving must be int
                        self.type_err([(moving, (int,))], line, lines_ran)
                        ret = mouse.move(
                            coords=(curr_x - moving, curr_y))
                    # moves the mouse right from its current location
                    elif objfunc == 'right':
                        import win32api
                        curr_x, curr_y = win32api.GetCursorPos()
                        moving = self.parse(0, line, f, sp, args)[2]
                        # moving must be int
                        self.type_err([(moving, (int,))], line, lines_ran)
                        ret = mouse.move(
                            coords=(curr_x + moving, curr_y))
                    # drags the mouse
                    # takes 4 or 5 arguments
                    # the first two are the starting coordinates
                    # the second two are the ending coordinates
                    # 5th argument is speed from 0-100
                    elif objfunc == 'drag':
                        import time
                        start = (self.parse(0, line, f, sp, args)[2],
                                 self.parse(1, line, f, sp, args)[2])
                        # start[1] and [2] must be int
                        self.type_err([(start[0], (int,)),
                                       (start[1], (int,))], line, lines_ran)
                        end = (self.parse(2, line, f, sp, args)[2],
                               self.parse(3, line, f, sp, args)[2])
                        # end[1] and [2] must be int
                        self.type_err([(end[0], (int,)),
                                        (end[1], (int,))], line, lines_ran)
                        # presses the mouse down at the coordinates
                        mouse.press(coords=start)
                        # slowly moves the mouse to the end coordinates
                        # this is to prevent the mouse from moving too fast
                        # and not dragging the object
                        # the farther the distance, the longer it takes
                        # to move the mouse
                        speed = 50
                        if len(args) == 5:
                            speed = self.parse(4, line, f, sp, args)[2]
                            # speed must be int
                            self.type_err([(speed, (int,))], line, lines_ran)
                        # reverse the speed, so a speed of 50 gives
                        # end_range of 50, and a speed of 75 gives
                        # end_range of 25
                        end_range = 100 - speed
                        for i in range(0, end_range):
                            mouse.move(coords=(int(start[0] + (end[0] - start[0]) / 100 * i),
                                               int(start[1] + (end[1] - start[1]) / 100 * i)))
                            time.sleep(0.001)
                        # releases the mouse at the end coordinates
                        mouse.release(coords=end)
                        ret = True
                    # release the lock
                    if p_thread:
                        pointer_lock.release()
                    return ret
                # clipboard operations
                # if no arguments, the clipboard
                # is returned
                #
                # if one argument, the text is copied
                # uses pyperclip
                elif func == 'clipboard':
                    # import pyperclip
                    import pyperclip
                    # if no arguments
                    if args[0][0] == '':
                        return pyperclip.paste()
                    # if one argument
                    else:
                        copying = str(self.parse(0, line, f, sp, args)[2])
                        pyperclip.copy(copying)
                        return copying
                # functional syntax I decided to add to make loops a tiny bit faster,
                # cannot receive non literal arguments
                # syntax:     3|5|i (prnt(i))
                # prnts 3\n4\n5
                elif func.count('|') == 2:
                    loop_args = func.split('|')
                    start = self.interpret(loop_args[0])
                    end = self.interpret(loop_args[1])
                    loopvar = loop_args[2]
                    # prepare loop variable
                    self.vars[loopvar] = Var(loopvar, start)
                    if start < end:
                        for i in range(start, end):
                            self.vars[loopvar].value = i
                            self.interpret(args[0][0])

                    # reversed iteration
                    else:
                        for i in reversed(range(end, start)):
                            self.vars[loopvar].value = i
                            self.interpret(args[0][0])
                    return
                # fallback
                else:
                    try:
                        line = self.replace_vars2(line)
                        return eval(line, {}, {})
                    except:
                        # maybe its a variable?
                        try:
                            return self.vars[line].value
                        except:
                            # ok hopefully its a string lol
                            return line
            if obj != '':
                objfunc += c
            else:
                func += c
        # try a variable
        if line in self.vars:
            return self.vars[line].value
        # otherwise nothing
        # try replacing variables 2
        try:
            line = self.replace_vars2(line)
        except:
            None
        # get value of line
        try:
            return eval(line, {}, {})
        except:
            try:
                return eval(str(self.replace_vars(line)), {}, {})
            except:
                return None
    # adds a new program wide syntax

    def add_syntax(self, token, between, function):
        syntax[token] = [between, function]
        return [token, between, function]

    # class specific error messages
    def no_var_err(self, _vn, _type, _and, _locals, line):
        return self.err(
            f'No {_type} variable "{_vn}" found in Python environment',
            f'Current {_and} environment locals:\n{self._locals}\nsize: {len(self._locals)}\nsee globals with "py.globals()"',
            line, lines_ran
        )

    # replaces tokens in the string with certain
    # characters or values
    # TODO: implement linear interpretation
    def msn2_replace(self, script):

        # replace hashtag marker with a hashtag
        script = script.replace('<tag>', '#')
        script = script.replace('<nl>', '\n')
        script = script.replace('<rp>', ')')
        script = script.replace('<lp>', '(')
        script = script.replace('(,)', ',')
        script = script.replace('<or>', '||')

        tag = '<msn2element>'
        endtag = '</msn2element>'
        # replaces whats in between the tags
        # with the interpretation of whats between the tags
        #
        # interpretation  is with self.interpret(script)
        #
        # script(
        #     <msn2element>'hello1'</msn2element>

        #     <msn2element>
        #         cat('hello',
        #             <msn2element>'hi there'</msn2element>
        #         )
        #     </msn2element>
        # )
        #
        # correct output of script() = hello1hellohi there

        # interpret the tags similar to the way
        # parantheticals are interpreted
        # this is a recursive function
        # open paren = tag
        # close paren = endtag

        def recurse_tags(scr, force_string=False):
            # get the first tag
            # if there is no tag, return the script
            if (first := scr.find(tag)) == -1:
                return scr
            # find the matching end tag
            stack = []
            i = first + len(tag)
            while i < len(scr):
                if scr[i:i+len(endtag)] == endtag:
                    if len(stack) == 0:
                        break
                    stack.pop()
                    i += len(endtag)
                elif scr[i:i+len(tag)] == tag:
                    stack.append(tag)
                    i += len(tag)
                else:
                    i += 1
            # recursively interpret the code between the tags
            interpreted_code = self.interpret(
                recurse_tags(scr[first+len(tag):i]))
            if force_string:
                interpreted_code = f'"{interpreted_code}"'
            new_scr = f'{scr[:first]}{interpreted_code}{scr[i+len(endtag):]}'
            # recursively continue replacing tags in the remaining script
            return recurse_tags(new_scr)
        # applying <msn2element> tags
        with_msn2elements = recurse_tags(script)
        # switch tags
        tag = '<msn2>'
        endtag = '</msn2>'
        # applying <msn2> tags
        # for string based needs
        with_msn2 = recurse_tags(with_msn2elements, force_string=True)
        # applying '{=' '=}' tags
        # does the same thing as <msn2element> tags
        tag = '{='
        endtag = '=}'
        with_msn2 = recurse_tags(with_msn2)
        return with_msn2

    # determines the number of arguments based on the args array
    def arg_count(self, args):
        # no arguments supplied
        if args[0][0] == '':
            return 0
        else:
            return len(args)

    # returns True if should go to the next file to import
    def imp(self, i, line, f, sp, args, can_exit):
        path = self.parse(i, line, f, sp, args)[2]
        # path must be a string
        self.type_err([(path, (str,))], line, lines_ran)
        if not path.endswith('.msn2'):
            path += '.msn2'
        if path in can_exit:
            return False
        can_exit.add(path)
        contents = ''
        with open(path) as f:
            contents = f.readlines()
            script = ''
            for line in contents:
                script += line
        self.logg("importing library", path)
        return script

    def run_syntax(self, key, line):
        # get everything between after syntax and before next index of syntax
        # or end of line
        inside = line[len(key):line.rindex(key)]
        # variable name
        invarname = syntax[key][0]
        # function to be run
        function = syntax[key][1]
        # store the in between for the user function
        self.vars[invarname] = Var(invarname, inside)
        return self.interpret(function)

    def replace_vars2(self, line):
        for varname, var in self.vars.items():
            try:
                val = var.value
            except:
                try:
                    val = eval(str(var), {}, {})
                except:
                    val = str(var)
            if isinstance(val, str):
                val = f'"{val}"'
            line = line.replace(f'?{varname}?', str(val))
        return line
    # gets a thread by name

    def thread_by_name(self, name):
        try:
            # thread exists
            return self.env_by_name(name)[0]
        except:
            # thread does not exist (yet)
            return None
    # gets an environment by name

    def env_by_name(self, name):
        for threadname in self.threads.keys():
            if threadname == name:
                return self.threads[threadname]
        return None
    # splits a named argument in a function

    def split_named_arg(self, as_s, method, func_args):
        # name of argument
        meth_argname = ''
        # iterate up to the first '=' in as_s
        for j in range(1, len(as_s)):
            if as_s[j] == '=':
                last_ind = j + 1
                break
            # cannot be space
            if as_s[j] != ' ':
                meth_argname += as_s[j]
        # argument to interpret
        arg = as_s[last_ind:]
        # set the method arg at the specified index
        # to the corresponding argument name in method.args
        # similar to that of Python's argument picker func(arg1=None, arg2=None)
        # get the index of the named argument in method.args
        try:
            ind = method.args.index(meth_argname)
        except ValueError:
            # add the meth_argname to method.args
            method.args.append(meth_argname)
            ind = method.args.index(meth_argname)
        # adjust the func_args list to the correct size
        if len(func_args) < ind + 1:
            func_args += [None] * (ind + 1 - len(func_args))
        # set the argument at the index to the value
        func_args[ind] = self.interpret(arg)
        return func_args, meth_argname, arg, ind

    # parses an argument from a function
    def parse(self, arg_number, line, f, sp, args):
        as_s = args[arg_number][0]
        line, ret = self.convert_arg(as_s, line, f, sp, args)
        return line, as_s, ret

    def convert_arg(self, ins_s, line, f, sp, args):
        ret = self.interpret(ins_s)
        return f"{line[:f + sp + args[0][1] + 1]}{ret}{line[f + sp + args[0][2] + 1:]}", ret

    # gets the shortened version of a variable's value
    def shortened(self, needs_short):
        # if the string is greater than 200 chars
        if len(_t := str(needs_short)) > self.env_max_chars:
            return f'{_t[:self.env_max_chars]}...'
        # otherwise, return the string
        return _t
    # starts a user shell

    def shell(self):
        ip = None
        while ip != 'exit':
            ip = input(">>> ")
            self.interpret(ip)
    # logs something to the console

    def logg(self, msg, line):
        # self.log += "[*] " + msg + " : " + line + "\n"
        # do the above but with f strings
        self.log += f"[*] {msg} : {line}\n"

    # prints an array with styling
    # the only argument is an array of maps,
    # this is an example:
    #
    # [{'text': 'Hello', 'style': 'bold', 'fore': 'green', 'back': 'blue'},
    # {'text': ' there', 'style': 'italic'}]
    # version : 2.0.386
    def styled_print(self, array):
        global colors
        # set colors if not available for the entire
        # environment
        if not colors:
            # set the colors / styles
            colors = {
                '': '',
                'reset': '\033[0m',
                'black': '\033[30m',
                'red': '\033[31m',
                'green': '\033[32m',
                'yellow': '\033[33m',
                'blue': '\033[34m',
                'magenta': '\033[35m',
                'cyan': '\033[36m',
                'white': '\033[37m',
                'bold': '\033[1m',
                'italic': '\033[3m',
                'underline': '\033[4m',
                'blink': '\033[5m'
            }
        # for each map in the array
        for mp in array:
            # get all the parameters if they exist
            text = mp['text'] if 'text' in mp else ''
            # get the style
            style = mp['style'] if 'style' in mp else ''
            # get the foreground color
            fore = mp['fore'] if 'fore' in mp else ''
            # get the background color
            back = mp['back'] if 'back' in mp else ''
            # print the text with the style
            print(colors[style] + colors[fore] +
                  colors[back] + str(text) + colors['reset'], end='')
        # print a newline
        print()

    # checks for and throws a type error
    def type_err(self, values: list[(any, tuple)], line, lines_ran):
        # for each entry in values
        for value, permitted_types in values:
            # if the value is not in the permitted types
            if (current_type := type(value)) not in permitted_types:
                # create string for error
                error_s = ''
                # for each permitted type
                for i, permitted_type in enumerate(permitted_types):
                    # add the type to the error string
                    error_s += str(permitted_type)
                    # if this is not the last type
                    if i != len(permitted_types) - 1:
                        # add a comma
                        error_s += ' or '
                # throw error
                self.err(
                    'Incorrect type specified',
                    f"In value: {value}, expected {error_s} got {current_type}",
                    line, lines_ran
                )

    # general error printing
    def err(self, err, msg, line, lines_ran):
        # if we're not trying something, and there's an error,
        # print the error
        if not self.trying:
            # the total words printed for this error
            words_printed = ''
            # prints the error
            def print_err(array):
                # print the error
                self.styled_print(array)
                # add to words printed
                nonlocal words_printed
                words_printed += str(array)
            # printing the traceback
            print_err([
                {'text': 'MSN2 Traceback:\n', 'style': 'bold', 'fore': 'green'},
                {'text': (divider := '--------------'),
                 'style': 'bold', 'fore': 'green'},
            ])
            _branches = []
            root_nums = []            
            for k, (root_num, code_line) in inst_tree.items():
                root_nums.append(root_num)
            root_nums = list(set(root_nums))
            root_nums.sort()
            for root_num in root_nums:
                branches = []
                for k, (root_num2, code_line2) in inst_tree.items():
                    if root_num2 == root_num:
                        branches.append(code_line2)
                _branches.append(branches)
            # print the traceback
            # only the last 7 branches
            for i, _branch in enumerate((new_branches := _branches[-7:])):
                # color of the text
                _branch_color = 'black'
                # if this is the last branch
                if (is_caller := i == len(new_branches) - 1):
                    _branch_color = 'red'
                else:
                    _branch_color = 'white'
                # print the caller
                print_err([
                    {'text': '>> ', 'style': 'bold', 'fore': 'black'},
                    {'text': self.shortened(_branch[0].strip()), 'style': 'bold',
                     'fore': _branch_color},
                    {'text': ' <<< ' if is_caller else '',
                        'style': 'bold', 'fore': 'yellow'},
                    {'text': 'SOURCE' if is_caller else '',
                        'style': 'bold', 'fore': 'yellow'}
                ])
                # if branches more than 3
                if len(_branch) > 4 and not is_caller:
                    # print the lines branching off
                    print_err([
                        {'text': '    at   ', 'style': 'bold', 'fore': 'black'},
                        {'text': self.shortened(_branch[0].strip()), 'style': 'bold',
                         'fore': _branch_color}
                    ])
                    # print ...
                    print_err([
                        {'text': '    at   ', 'style': 'bold', 'fore': 'black'},
                        {'text': f'... ({len(_branch) - 4} more)',
                         'style': 'bold', 'fore': 'black'}
                    ])
                # if branches less than 3
                else:
                    if len(_branch) > 7:
                        # print the before elipses
                        print_err([
                            {'text': '    at   ', 'style': 'bold', 'fore': 'black'},
                            {'text': f'... ({len(_branch) - 7} more)',
                             'style': 'bold', 'fore': 'black'}
                        ])
                        for i, _branch2 in enumerate(_branch[len(_branch) - 7:]):
                            print_err([
                                {'text': '    at   ',
                                    'style': 'bold', 'fore': 'black'},
                                {'text': self.shortened(_branch2.strip()), 'style': 'bold',
                                 'fore': _branch_color}
                            ])
                    else:
                        for _branch2 in _branch[1:]:

                            print_err([
                                {'text': '    at   ',
                                    'style': 'bold', 'fore': 'black'},
                                {'text': self.shortened(_branch2.strip()), 'style': 'bold',
                                 'fore': _branch_color}
                            ])
            # print the finishing divider
            print_err([
                {'text': divider, 'style': 'bold', 'fore': 'green'}
            ])
            # print this error with print_err()
            print_err([
                {'text': '[-] ', 'style': 'bold', 'fore': 'red'},
                {'text': err, 'style': 'bold', 'fore': 'red'},
                {'text': '\n'},
                {'text': msg, 'style': 'bold', 'fore': 'red'},
            ])
            # add to log
            self.log += words_printed + "\n"
        raise self.MSN2Exception(
            'MSN2 Exception thrown, see above for details')

    # throws a keyerror
    def raise_key(self, key, line):
        return self.err(
            'Key not found in dictionary',
            f'Key "{key}" not found in dictionary',
            line, lines_ran
        )

    # throws msn2 error for a work in progress mechanism
    def raise_wip(self, what_is_wip, line, lines_ran):
        # get the version for this interpreter
        return self.err(
            'Work in progress',
            f'{what_is_wip} is a work in progress in MSN2 v{self.version} and not been fully implemented.',
            line, lines_ran
        )
    # verifies a type is iterable
    # iterable types include: list, tuple, dict, set, str
    def check_iterable(self, value, line):
        import collections.abc
        if not isinstance(value, collections.abc.Iterable):
            return self.type_err([(value, (collections.abc.Iterable,))], line, lines_ran)
        return True

    # checks for validity of a variable name being set
    # by its string representation
    def check_varname(self, varname, line):
        # varname must be a string
        self.type_err([(varname, (str,))], line, lines_ran)
        # varname must exist
        if not varname:
            self.raise_varname_chars(varname, line, lines_ran)

    # throws an operation error
    def raise_operation_err(self, vname, operator, line):
        # operation error
        return self.err(
            f'Operation error: "{operator}"',
            f'There was an error applying the operator "{operator}" to "{vname}" and its argument(s).\n',
            line, lines_ran
        )
    # throws an error asking for a variable name to at least contain characters

    def raise_varname_chars(self, varname, line, lines_ran):
        return self.err(
            'New variable name must contain characters.',
            f'Variable name "{varname}" is an invalid name.',
            line,
            lines_ran
        )
    # throws msn2 error for comparison error

    def raise_comp(self, comp_operation, vname, line):
        return self.err(
            'Comparison error',
            f'Could not perform comparison operation "{comp_operation}" on variable "{vname}" and the supplied argument(s)',
            line, lines_ran
        )
    # empty array error

    def raise_empty_array(self, line):
        return self.err(
            'Error computing average',
            f'Array is empty',
            line, lines_ran
        )
    # non_numeric value error

    def raise_avg(self, line):
        return self.err(
            'Error computing average',
            f'Array contains non-numeric values',
            line, lines_ran
        )
    # value is not in the list

    def raise_value(self, value, line):
        return self.err(
            'Value error',
            f'Value "{value}" is not in the list.',
            line, lines_ran
        )
    # throws msn2 error for Incorrect number of arguments

    def raise_incorrect_args(self, expected, actual, line, lines_ran, method):
        return self.err(f'Incorrect number of function arguments for {method.name}',
                        f'Expected {expected}, got {actual}', line, lines_ran)
    # throws msn2 error for Index out of bounds

    def raise_index_out_of_bounds(self, line, lines_ran, method):
        return self.err(f'Index out of bounds in body of {method.name}',
                        f'Index out of bounds', line, lines_ran)

    def __del__(self):
        None

    def method_args(self, line, j):
        argstring = ''
        instring = False
        for k in range(j + 1, len(line)):
            if line[k] == '"' and not instring:
                instring = True
            elif line[k] == '"' and instring:
                instring = False
            if not instring:
                if line[k] != ' ':
                    if line[k] == ')':
                        break
                    argstring += line[k]
            else:
                argstring += line[k]
        return argstring.split(','), k

    def thread_split(self, line):
        self.interpret(line)

    # splits a process
    def process_split(self, line):
        inter = Interpreter()
        inter.parent = self
        return inter.interpret(line)

    # executing Python scripts
    def exec_python(self, python_block):
        # get the python script with arguments inserted
        py_script = str(self.interpret(f"script({python_block})"))
        # try to execute the script
        try:
            exec(py_script, self._globals, self._locals)
        except Exception as e:
            # send an error
            self.err("Error running Python script",
                     str(e), py_script, lines_ran)
        return py_script

    def var_exists(self, varname):
        if varname in self.vars:
            return True
        return False

    def var_names(self):
        names = []
        for key in self.vars:
            names.append(key)
        return names

    def is_py_str(self, line):
        try:
            return line[0] == '"' and line[len(line) - 1] == '"'
        except:
            return False

    def is_str(self, line):
        return line[0] == '<' and line[len(line) - 1] == '>'

    # gets the variable value from the variable name
    def get_var(self, name):
        return self.vars[name].value

    # extracts the argument lines from the merged arguments passed
    def get_args(self, line):
        args = []
        l = len(line)
        arg = ''
        start = 0
        p = 0
        a = 0
        s = 0
        indouble = False
        s2 = 0
        insingle = False
        b = 0
        for i in range(l + 1):
            c = ''
            try:
                c = line[i]
            except:
                None
            if c == '[' and not s2 > 0 and not s > 0:
                a += 1
            if c == ']' and not s2 > 0 and not s > 0:
                a -= 1
            if c == '(' and not s2 > 0 and not s > 0:
                p += 1
            if c == ')' and not s2 > 0 and not s > 0:
                p -= 1
            if not self.in_string(s, s2):
                if c == '{':
                    b += 1
                if c == '}':
                    b -= 1
            if not indouble and not s2 > 0 and c == '"':
                s += 1
                indouble = True
            elif indouble and c == '"':
                s -= 1
                indouble = False
            if not insingle and not s > 0 and c == "'":
                s2 += 1
                insingle = True
            elif insingle and c == "'":
                s2 -= 1
                insingle = False
            # print(line)
            # print (f"""
            # c: {c}
            # p: {p}
            # a: {a}
            # s: {s}
            # s2: {s2}

            # """)
            if c == ',' and s == 0 and s2 == 0 and p == 0 and a == 0 and b == 0:
                args.append([arg, start, start + len(arg)])
                start = i + 1
                arg = ''
                continue
            elif i == l:
                args.append([arg, start, start + len(arg)])
            arg += c
        return args

    def in_string(self, s, s2):
        return s > 0 or s2 > 0

    def interpret_msnscript_1(self, line):
        element = ''
        variable = ''
        for i in range(0, len(line)):
            c = line[i]
            if c != ' ':
                if c == '+' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]

                    # if element is a number
                    if isinstance(element, float) or isinstance(element, int):
                        self.vars[variable].value += self.interpret(element)
                    # if element is a string
                    elif isinstance(element, str):
                        try:
                            self.vars[variable].value += self.interpret(
                                element)
                        except:
                            self.vars[variable].value += self.interpret(
                                element)
                    return self.vars[variable].value
                elif c == '-' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value -= self.interpret(element)
                    return self.vars[variable].value
                elif c == '*' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value *= self.interpret(element)
                    return self.vars[variable].value
                elif c == '/' and line[i + 1] == '=':
                    variable = element
                    element = ''
                    for j in range(i + 2, len(line)):
                        element += line[j]
                    self.vars[variable].value /= self.interpret(element)
                    return self.vars[variable].value
                elif c == '=':
                    variable = element
                    element = ''
                    string = False
                    array = False
                    for j in range(i+1, len(line)):
                        if line[j] == '"':
                            string = True
                        if line[j] == '[':
                            array = True
                        element += line[j]
                    self.vars[variable] = Var(
                        variable, self.interpret(element))
                    return self.vars[variable].value
                else:
                    element += c

    # scrapes all html elements from a URL

    def html_all_elements(self, url):
        import requests
        # web scraping
        from bs4 import BeautifulSoup
        # obtains a response from the URL
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html5lib')
        # obtains all html elements
        return soup.find_all()

    def me(self):
        return str(self).replace(' ', '').replace('<', '').replace('>', '').replace('Interpreter', '')

    # prints text with a box around it
    def bordered(text):
        lines = text.splitlines()
        width = max(len(s) for s in lines)
        res = ['┌' + '─' * width + '┐']
        for s in lines:
            res.append('│' + (s + ' ' * width)[:width] + '│')
        res.append('└' + '─' * width + '┘')
        return '\n'.join(res)

    # exception
    class MSN2Exception(Exception):
        pass
    class Method:
        def __init__(self, name, interpreter):
            self.name = name
            self.args = []
            self.body = []
            self.returns = []
            self.ended = False
            self.interpreter = interpreter

        def add_arg(self, arg):
            self.args.append(arg)

        def add_body(self, body):
            self.body.append(body)

        def add_return(self, ret):
            self.returns.append(ret)

        def default(self, func_insert, index):
            return not func_insert and isinstance(self.args[index], list)

        def run(self, args, inter, actual_args=None):
            # finds an argument in the list of arguments
            def find_arg(func_var):
                ind = -1
                # for each arg in self.args
                for j, argname in enumerate(self.args):
                    # using default value
                    if isinstance(argname, list):
                        if argname[0] == func_var:
                            ind = j
                            break
                    # not
                    else:
                        if argname == func_var:
                            ind = j
                            break
                return ind
            # tries to place a variable in its argument location
            # before interpretation

            def try_var(func_var, func_insert):
                try:
                    # if self.args[i] is a list
                    if self.default(func_insert, i):
                        raise Exception()
                    # find the index of self.args
                    ind = find_arg(func_var)
                    if self.default(func_insert, ind):
                        inter.vars[func_var] = Var(
                            self.args[ind][0], self.args[ind][1])
                    else:
                        inter.vars[func_var] = Var(func_var, func_insert)
                except:
                    try:
                        if self.default(func_insert, i):
                            raise TypeError()
                        inter.vars[func_var] = func_insert
                    except TypeError:
                        # using default value
                        func_var, default_value = func_var
                        ind = find_arg(func_var)
                        if args[ind]:
                            inter.vars[func_var] = Var(func_var, args[ind])
                        elif default_value:
                            inter.vars[func_var] = Var(func_var, default_value)
            # loop through arguments to set
            for i in range(len(self.args)):
                if actual_args:
                    try:
                        if self.is_str(actual_args[i][0]):
                            try_var(self.args[i], args[i])
                            continue
                    # index out of bounds
                    except IndexError:
                        pass
                try:
                    inter.vars[self.args[i]] = inter.vars[args[i]]
                except:
                    try:
                        try_var(self.args[i], args[i])
                    except IndexError:
                        # using default value
                        args.append(self.args[i][1])
                        try_var(self.args[i][0], args[i])
            # required number of arguments for this function
            # remove 'self' if this is a class function
            required_arg_num = len(self.args) - self.args.count('self')
            # adjust self.args to remove duplicate arguments
            # for each type list in self.args
            for lst in (with_list := [lst for lst in self.args if isinstance(lst, list)]):
                # for each entry in the list
                for entry in lst:
                    # if the entry is in self.args
                    if entry in self.args:
                        # remove the entry
                        self.args.remove(entry)
            actual_arg_num = len(actual_args)
            for arg in actual_args:
                arg = arg[0]
                if arg == '':
                    actual_arg_num -= 1
            named_args_num = len(with_list)
            # check for error with named arguments
            if actual_arg_num + named_args_num < required_arg_num:
                inter.raise_incorrect_args(
                    required_arg_num, actual_arg_num + named_args_num, self.body[0], lines_ran, self)
            for line in self.body:
                method_ret = inter.interpret(line)
            return method_ret

        def is_str(self, value):
            return (value[0] == '"' and value[-1] == '"') or (value[0] == "'" and value[-1] == "'")
    # class for an App

    class App:
        # constructor
        def __init__(self, path, application=None, name=None, extension=None):
            # path of application being launched
            self.path = path
            if name:
                self.name = name
                self.extension = extension
            else:
                _spl = path.split('\\')[-1].split('.')
                # extension of the application
                self.extension = _spl[-1]
                self.name = _spl[0]
            # pwinauto application object
            self.application = application

    # element for an application
    class AppElement:
        # constructor
        def __init__(self, window, name):

            # creates a modified window
            self.window = window
            # # set the window
            # self.window = window
            # set the name
            self.name = name

        # gets the text of the window
        def window_text(self):
            return self.name

        # gets all children of the window
        def children(self):
            return self.window.children()

        # sets the focus to the window
        def set_focus(self):
            self.window.set_focus()

        # gets the properties of the window
        def get_properties(self):
            return self.window.get_properties()

        # gets the highest level parent of this element
        def top_level_parent(self):
            return self.window.top_level_parent()

        # computes the height of the window
        def height(self):
            try:
                return self.window.get_properties()['rectangle'].bottom - self.window.get_properties()['rectangle'].top
            except:
                return
        # computes the width of the window

        def width(self):
            try:
                return self.window.get_properties()['rectangle'].right - self.window.get_properties()['rectangle'].left
            except:
                return
        # string

        def __str__(self):
            return Interpreter.bordered(f'Text: {self.name if self.name else "[No Text Found]"}\nSize:\
{f"{self.width()}x{self.height()}"}\nObject:\n{self.window}')

    # class for a button
    class Button(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
        # clicks the button

        def click(self):
            self.window.click()
        # right clicks the button

        def right_click(self):
            self.window.click_input(button='right')
    # class for a Link

    class Link(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for a Menu

    class Menu(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for a ToolBar

    class ToolBar(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for a scrollbar

    class ScrollBar(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for TabItems

    class TabItem(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for Hyperlink

    class Hyperlink(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # class for Inputs

    class Input(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)

        # types text into the input
        def type_keys(self, text):
            self.window.type_keys(text)
    # class for Tables

    class Table(AppElement):
        # constructor
        def __init__(self, window, name):
            # call super constructor
            super().__init__(window, name)
    # ------------------------------------
    # working with Excel

    class Workbook:
        # constructor
        def __init__(self, workbook, path) -> None:
            self.workbook = workbook
            self.path = path

    # sheet class
    class Sheet(Workbook):
        def __init__(self, sheet, title, workbook, path) -> None:
            super().__init__(workbook, path)
            self.sheet = sheet
            self.title = title
